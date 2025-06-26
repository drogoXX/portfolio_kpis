# Portfolio Performance Dashboard - ENHANCED VERSION WITH ALL FIXES
# Fixed Issues: 
# 1. Work Package Variance Logic (negative variances as opportunities)
# 2. Material Work Package Summary calculations
# 3. CM Variability Analysis using actual template data
# 4. CPI/SPI Traffic Lights fixed
# 5. Risk Contingency section removed
# 6. POC Velocity calculation corrected

import streamlit as st
import pandas as pd
import numpy as np
import plotly.express as px
import plotly.graph_objects as go
from plotly.subplots import make_subplots
import datetime
import openpyxl
from openpyxl import load_workbook

# ================================================================================
# CONFIGURATION & CONSTANTS
# ================================================================================

WASTE_TO_ENERGY_ROI_TARGET = 12.0
REVENUE_LABELS = ['Contract Price', 'Revenues', 'POC%', 'Cash IN', 'Cash OUT', 'Cash In %']
QUARTERS = ['Q1', 'Q2', 'Q3', 'Q4', 'Total']

# Executive Traffic Light Thresholds
EXECUTIVE_THRESHOLDS = {
    'cm2_margin': {'excellent': 15, 'good': 10, 'warning': 5, 'critical': 0},
    'cm1_margin': {'excellent': 25, 'good': 20, 'warning': 15, 'critical': 10},
    'poc_velocity': {'excellent': 10, 'good': 5, 'warning': 2, 'critical': 0},
    'cash_flow_efficiency': {'excellent': 1.1, 'good': 0.95, 'warning': 0.85, 'critical': 0.75},
    'revenue_growth': {'excellent': 15, 'good': 5, 'stable': -2, 'warning': -5, 'critical': -15},
    'committed_vs_budget': {'excellent': 0.9, 'good': 1.0, 'warning': 1.1, 'critical': 1.2},
    'cost_variance': {'excellent': -5, 'good': 5, 'warning': 15, 'critical': 25},
    'schedule_performance': {'excellent': 110, 'good': 95, 'warning': 80, 'critical': 70},
    # FIX #4: CPI thresholds (higher is better)
    'cost_performance_index': {'excellent': 1.1, 'good': 1.0, 'warning': 0.9, 'critical': 0.8},
    # Add SPI thresholds
    'schedule_performance_index': {'excellent': 1.1, 'good': 1.0, 'warning': 0.9, 'critical': 0.8}
}

# Enhanced Cost Categories
COST_CATEGORIES = {
    'EC': 'External Costs',
    'IC': 'Internal Costs', 
    'CM1': 'Contribution Margin 1',
    'CM2': 'Contribution Margin 2'
}

# Page configuration
st.set_page_config(
    page_title="Executive Portfolio Dashboard",
    page_icon="📊",
    layout="wide",
    initial_sidebar_state="expanded"
)

# ================================================================================
# ENHANCED STYLING FOR EXECUTIVE DASHBOARD
# ================================================================================

st.markdown("""
<style>
    .main-header {
        background: linear-gradient(135deg, #1f4e79 0%, #2d5aa0 50%, #1f4e79 100%);
        padding: 1.5rem;
        border-radius: 15px;
        margin-bottom: 2rem;
        box-shadow: 0 4px 15px rgba(0,0,0,0.2);
    }
    .main-header h1 {
        color: white;
        margin: 0;
        text-align: center;
        font-size: 2.5rem;
        font-weight: 700;
    }
    .main-header p {
        color: #e6f2ff;
        text-align: center;
        margin: 0.5rem 0 0 0;
        font-size: 1.1rem;
    }
    
    /* Executive KPI Cards */
    .executive-kpi {
        background: white;
        padding: 1.5rem;
        border-radius: 12px;
        box-shadow: 0 3px 10px rgba(0,0,0,0.1);
        border-left: 5px solid #2F5F8F;
        margin: 0.5rem 0;
        text-align: center;
    }
    .executive-kpi h3 {
        margin: 0 0 0.5rem 0;
        color: #1f4e79;
        font-size: 1.1rem;
        font-weight: 600;
    }
    .executive-kpi .kpi-value {
        font-size: 2rem;
        font-weight: 700;
        margin: 0.5rem 0;
    }
    .executive-kpi .kpi-trend {
        font-size: 0.9rem;
        margin: 0;
    }
    
    /* Traffic Light Indicators */
    .traffic-light-excellent { border-left-color: #28a745; }
    .traffic-light-good { border-left-color: #28a745; }
    .traffic-light-stable { border-left-color: #17a2b8; }
    .traffic-light-warning { border-left-color: #ffc107; }
    .traffic-light-critical { border-left-color: #dc3545; }
    
    .status-excellent { color: #28a745; }
    .status-good { color: #28a745; }
    .status-stable { color: #17a2b8; }
    .status-warning { color: #ffc107; }
    .status-critical { color: #dc3545; }
    
    /* Executive Summary Cards */
    .exec-summary {
        background: linear-gradient(135deg, #f8f9fa 0%, #e9ecef 100%);
        padding: 1.5rem;
        border-radius: 12px;
        margin: 1rem 0;
        border-left: 4px solid #2F5F8F;
    }
    
    /* Margin Analysis Cards */
    .margin-card {
        background: white;
        padding: 1rem;
        border-radius: 8px;
        border-left: 4px solid;
        margin: 0.5rem 0;
        box-shadow: 0 2px 5px rgba(0,0,0,0.1);
    }
    .margin-excellent { border-left-color: #28a745; }
    .margin-good { border-left-color: #17a2b8; }
    .margin-warning { border-left-color: #ffc107; }
    .margin-critical { border-left-color: #dc3545; }
    
    /* Risk Assessment Cards */
    .risk-card {
        background: white;
        padding: 1rem;
        border-radius: 8px;
        margin: 0.5rem 0;
        box-shadow: 0 2px 5px rgba(0,0,0,0.1);
    }
    .risk-low { border-left: 4px solid #28a745; }
    .risk-medium { border-left: 4px solid #ffc107; }
    .risk-high { border-left: 4px solid #ff6b35; }
    .risk-critical { border-left: 4px solid #dc3545; }
</style>
""", unsafe_allow_html=True)

# ================================================================================
# UTILITY FUNCTIONS
# ================================================================================

def calculate_contingency_metrics(work_packages, poc_current=0):
    """
    Calculate contingency efficiency metrics for a project
    
    Returns dict with:
    - contingency_as_sold: Original contingency amount
    - contingency_fct_n1: Previous forecast contingency
    - contingency_fct_n: Current contingency
    - consumed_amount: Amount consumed (as_sold - fct_n)
    - consumed_percentage: Percentage consumed
    - efficiency: Contingency efficiency score
    - trend: Consumption trend (accelerating/stable/improving)
    - status: Status classification
    """
    # Find risk contingency work packages
    risk_contingencies = [
        wp for wp in work_packages.values() 
        if 'risk' in wp.get('description', '').lower() and 
        'contingenc' in wp.get('description', '').lower()
    ]
    
    if not risk_contingencies:
        return {
            'has_contingency': False,
            'contingency_as_sold': 0,
            'contingency_fct_n1': 0,
            'contingency_fct_n': 0,
            'consumed_amount': 0,
            'consumed_percentage': 0,
            'efficiency': None,
            'trend': 'No Contingency',
            'status': 'N/A',
            'status_icon': '➖',
            'status_color': 'info'
        }
    
    # Aggregate contingency values
    contingency_as_sold = sum(rc.get('as_sold', 0) for rc in risk_contingencies)
    contingency_fct_n1 = sum(rc.get('fct_n1', 0) for rc in risk_contingencies)
    contingency_fct_n = sum(rc.get('fct_n', 0) for rc in risk_contingencies)
    
    # Calculate consumption
    consumed_amount = contingency_as_sold - contingency_fct_n
    consumed_percentage = (consumed_amount / contingency_as_sold * 100) if contingency_as_sold > 0 else 0
    
    # Calculate efficiency using the corrected formula
    if poc_current > 0:
        efficiency = (2 - (consumed_percentage / poc_current)) * 100
    else:
        efficiency = 200  # No progress yet, so no consumption expected
    
    # Cap efficiency at reasonable bounds
    efficiency = max(0, min(200, efficiency))
    
    # Calculate trend
    early_consumption = contingency_as_sold - contingency_fct_n1 if contingency_as_sold > 0 else 0
    recent_consumption = contingency_fct_n1 - contingency_fct_n if contingency_fct_n1 > 0 else 0
    
    if early_consumption > 0 and recent_consumption > early_consumption * 1.2:
        trend = 'Accelerating'
        trend_icon = '↗️'
    elif recent_consumption < early_consumption * 0.8:
        trend = 'Improving'
        trend_icon = '↘️'
    else:
        trend = 'Stable'
        trend_icon = '→'
    
    # Determine status
    if efficiency >= 150:
        status = 'Excellent'
        status_icon = '🟢'
        status_color = 'success'
    elif efficiency >= 120:
        status = 'Good'
        status_icon = '🟢'
        status_color = 'success'
    elif efficiency >= 80:
        status = 'On Track'
        status_icon = '🟦'
        status_color = 'info'
    elif efficiency >= 50:
        status = 'Warning'
        status_icon = '🟡'
        status_color = 'warning'
    else:
        status = 'Critical'
        status_icon = '🔴'
        status_color = 'error'
    
    return {
        'has_contingency': True,
        'contingency_as_sold': contingency_as_sold,
        'contingency_fct_n1': contingency_fct_n1,
        'contingency_fct_n': contingency_fct_n,
        'consumed_amount': consumed_amount,
        'consumed_percentage': consumed_percentage,
        'remaining_amount': contingency_fct_n,
        'remaining_percentage': (contingency_fct_n / contingency_as_sold * 100) if contingency_as_sold > 0 else 0,
        'efficiency': efficiency,
        'trend': trend,
        'trend_icon': trend_icon,
        'status': status,
        'status_icon': status_icon,
        'status_color': status_color,
        'early_consumption': early_consumption,
        'recent_consumption': recent_consumption
    }

def get_traffic_light_status(value, thresholds, reverse=False):
    """Get traffic light status based on thresholds"""
    if reverse:  # For metrics where lower is better (like cost overruns)
        if value <= thresholds['excellent']:
            return "🟢", "Excellent", "excellent"
        elif value <= thresholds['good']:
            return "🟢", "Good", "good"
        elif value <= thresholds['warning']:
            return "🟡", "Warning", "warning"
        else:
            return "🔴", "Critical", "critical"
    else:  # For metrics where higher is better
        if value >= thresholds['excellent']:
            return "🟢", "Excellent", "excellent"
        elif value >= thresholds['good']:
            return "🟢", "Good", "good"
        elif 'stable' in thresholds and value >= thresholds['stable']:
            return "🟦", "Stable", "stable"
        elif value >= thresholds['warning']:
            return "🟡", "Warning", "warning"
        else:
            return "🔴", "Critical", "critical"

def format_currency_millions(value):
    """Format currency values in millions with 1 decimal place"""
    if pd.isna(value) or value == 0:
        return "CHF 0.0M"
    millions = value / 1000
    return f"CHF {millions:.1f}M"

def format_currency_thousands(value):
    """Format currency values in thousands"""
    if pd.isna(value) or value == 0:
        return "CHF 0K"
    return f"CHF {value:,.0f}K"

def format_percentage(value, decimals=1):
    """Format percentage values"""
    if pd.isna(value):
        return "N/A"
    return f"{value:.{decimals}f}%"

def safe_get_value(data_dict, *keys, default=0):
    """Safely get nested dictionary values"""
    try:
        value = data_dict
        for key in keys:
            value = value[key]
        return float(value) if value is not None else default
    except (KeyError, TypeError, ValueError):
        return default

def safe_float(value):
    """Safely convert value to float"""
    try:
        return float(value) if value is not None else 0
    except (ValueError, TypeError):
        return 0

# FIX #6: Corrected POC Velocity Calculation
def calculate_poc_velocity(poc_current, poc_previous):
    """Calculate POC velocity as simple difference (already in %)"""
    try:
        current = float(poc_current) if poc_current is not None else 0
        previous = float(poc_previous) if poc_previous is not None else 0
        # POC is already in percentage, so velocity is just the difference
        return current - previous
    except (ValueError, TypeError):
        return 0


def get_threshold_summary():
    """Get a summary of active thresholds for display"""
    return {
        "CM2 Status Ranges": f"🟢 ≥{EXECUTIVE_THRESHOLDS['cm2_margin']['excellent']}% | 🟢 ≥{EXECUTIVE_THRESHOLDS['cm2_margin']['good']}% | 🟡 ≥{EXECUTIVE_THRESHOLDS['cm2_margin']['warning']}% | 🔴 <{EXECUTIVE_THRESHOLDS['cm2_margin']['warning']}%"
    }

# Then in the sidebar, add:
with st.sidebar.expander("📊 Custom Threshold Ranges", expanded=False):
    threshold_summary = get_threshold_summary()
    for metric, ranges in threshold_summary.items():
        st.markdown(f"**{metric}:**")
        st.markdown(f"<small>{ranges}</small>", unsafe_allow_html=True)
        st.markdown("")

def calculate_period_variance(current_value, previous_value):
    """Calculate period-over-period variance with proper handling of edge cases"""
    try:
        current_value = float(current_value) if current_value is not None else 0
        previous_value = float(previous_value) if previous_value is not None else 0
        
        if previous_value == 0:
            return 100 if current_value > 0 else 0
        return ((current_value - previous_value) / abs(previous_value)) * 100
    except (ValueError, TypeError):
        return 0

def calculate_earned_value_metrics(project_data):
    """Calculate comprehensive earned value management metrics"""
    try:
        # Basic values
        contract_value = safe_get_value(project_data, 'revenues', 'Contract Price', 'n_ptd')
        poc_current = safe_get_value(project_data, 'revenues', 'POC%', 'n_ptd')
        revenue_current = safe_get_value(project_data, 'revenues', 'Revenues', 'n_ptd')
        cost_analysis = project_data.get('cost_analysis', {})
        actual_costs = cost_analysis.get('total_actuals', 0)
        
        # Earned Value Calculations
        planned_value = contract_value  # Budget at completion
        earned_value = (poc_current / 100) * contract_value if poc_current > 0 else revenue_current
        actual_cost = actual_costs if actual_costs > 0 else revenue_current
        
        # Performance Indices
        cost_performance_index = earned_value / actual_cost if actual_cost > 0 else 1.0
        schedule_performance_index = earned_value / (revenue_current if revenue_current > 0 else earned_value) if revenue_current > 0 else 1.0
        
        # Variances
        cost_variance = earned_value - actual_cost
        schedule_variance = earned_value - revenue_current
        
        # Forecasting
        estimate_at_completion = actual_cost + ((contract_value - earned_value) / cost_performance_index) if cost_performance_index > 0 else contract_value
        variance_at_completion = contract_value - estimate_at_completion
        
        return {
            'planned_value': planned_value,
            'earned_value': earned_value,
            'actual_cost': actual_cost,
            'cost_performance_index': cost_performance_index,
            'schedule_performance_index': schedule_performance_index,
            'cost_variance': cost_variance,
            'schedule_variance': schedule_variance,
            'estimate_at_completion': estimate_at_completion,
            'variance_at_completion': variance_at_completion,
            'cost_efficiency': (cost_performance_index - 1) * 100,
            'schedule_efficiency': (schedule_performance_index - 1) * 100
        }
    except Exception:
        return {
            'planned_value': 0, 'earned_value': 0, 'actual_cost': 0,
            'cost_performance_index': 1.0, 'schedule_performance_index': 1.0,
            'cost_variance': 0, 'schedule_variance': 0,
            'estimate_at_completion': 0, 'variance_at_completion': 0,
            'cost_efficiency': 0, 'schedule_efficiency': 0
        }

def get_overall_project_status(cm2_class, committed_class, poc_class, poc_current=None, poc_velocity=None):
    """
    Determine overall project status based on key metrics with maturity consideration
    
    Parameters:
    - cm2_class: Status class for CM2 margin
    - committed_class: Status class for committed ratio
    - poc_class: Status class for POC velocity (raw, without maturity adjustment)
    - poc_current: Current POC percentage (optional, for maturity adjustment)
    - poc_velocity: Current POC velocity (optional, for maturity adjustment)
    """
    status_scores = {
        'excellent': 4, 'good': 3, 'warning': 2, 'critical': 1
    }
    
    # Get base scores
    cm2_score = status_scores.get(cm2_class, 1)
    committed_score = status_scores.get(committed_class, 1)
    poc_score = status_scores.get(poc_class, 1)
    
    # Adjust POC velocity score based on project maturity if data is available
    if poc_current is not None and poc_velocity is not None:
        # Recalculate POC status with maturity consideration
        _, _, adjusted_poc_class = get_poc_velocity_status_with_maturity(poc_velocity, poc_current)
        poc_score = status_scores.get(adjusted_poc_class, poc_score)
    
    # Calculate average with adjusted scores
    avg_score = (cm2_score + committed_score + poc_score) / 3
    
    if avg_score >= 3.5:
        return "🟢 Excellent"
    elif avg_score >= 2.5:
        return "🟢 Good"  # Changed from yellow to green for clarity
    elif avg_score >= 1.5:
        return "🟠 Warning"
    else:
        return "🔴 Critical"


def calculate_project_health_score(cpi, spi, cm2_pct, poc_velocity=None):
    """
    Calculate normalized project health score (0-100%)
    
    Components:
    - Cost Performance (30%): CPI where 1.0 = 100%, capped at 120%
    - Schedule Performance (30%): SPI where 1.0 = 100%, capped at 120%
    - Margin Health (25%): CM2% where 15% = 100%, scaled appropriately
    - Progress Velocity (15%): POC velocity where 5%/month = 100%
    """
    
    # Normalize CPI (0-100 scale, with bonus for overperformance)
    # CPI of 1.0 = 100 points, max 120 points at CPI 1.2+
    cpi_score = min(cpi * 100, 120)
    
    # Normalize SPI (0-100 scale, with bonus for overperformance)
    # SPI of 1.0 = 100 points, max 120 points at SPI 1.2+
    spi_score = min(spi * 100, 120)
    
    # Normalize CM2 Margin (0-100 scale)
    # 0% = 0 points, 15% = 100 points, 20%+ = 120 points
    if cm2_pct <= 0:
        cm2_score = 0
    elif cm2_pct <= 15:
        cm2_score = (cm2_pct / 15) * 100
    else:
        # Bonus points for margins above 15%, capped at 120
        cm2_score = min(100 + ((cm2_pct - 15) / 5) * 20, 120)
    
    # Normalize POC Velocity if provided (0-100 scale)
    # 0% = 0 points, 5%/month = 100 points, 7%+ = 120 points
    if poc_velocity is not None:
        if poc_velocity <= 0:
            velocity_score = 0
        elif poc_velocity <= 5:
            velocity_score = (poc_velocity / 5) * 100
        else:
            velocity_score = min(100 + ((poc_velocity - 5) / 2) * 20, 120)
    else:
        velocity_score = 100  # Default if not provided
    
    # Calculate weighted health score
    if poc_velocity is not None:
        health_score = (
            cpi_score * 0.30 +
            spi_score * 0.30 +
            cm2_score * 0.25 +
            velocity_score * 0.15
        )
    else:
        # Without velocity, redistribute weight
        health_score = (
            cpi_score * 0.35 +
            spi_score * 0.35 +
            cm2_score * 0.30
        )
    
    # Ensure final score is between 0-100
    return min(max(health_score, 0), 100)


def calculate_expected_poc_velocity(poc_current):
    """
    Calculate expected POC velocity based on project maturity
    Returns expected velocity in percentage points per month
    """
    if poc_current >= 95:
        return 1.0  # 1% per month is fine for nearly complete projects
    elif poc_current >= 90:
        return 2.0  # 2% per month expected
    elif poc_current >= 80:
        return 3.0  # 3% per month expected  
    elif poc_current >= 60:
        return 5.0  # 5% per month expected
    elif poc_current >= 40:
        return 7.0  # 7% per month expected
    else:
        return 10.0  # 10%+ per month expected for early stage projects

def get_poc_velocity_status_with_maturity(poc_velocity, poc_current):
    """
    Get POC velocity status considering project maturity
    Returns icon, status text, and status class
    """
    expected_velocity = calculate_expected_poc_velocity(poc_current)
    velocity_ratio = poc_velocity / expected_velocity if expected_velocity > 0 else 0
    
    # Determine status based on how actual velocity compares to expected
    if velocity_ratio >= 1.2:  # 20% above expected
        return "🟢", "Excellent", "excellent"
    elif velocity_ratio >= 0.8:  # Within 80% of expected
        return "🟢", "Good", "good"
    elif velocity_ratio >= 0.5:  # Within 50% of expected
        return "🟡", "Warning", "warning"
    else:  # Below 50% of expected
        return "🔴", "Critical", "critical"


# ================================================================================
# ENHANCED TEMPLATE PARSING FUNCTIONS
# ================================================================================

def parse_excel_template_v24(uploaded_file):
    """Parse Template_Simple v2.3/v2.4 with comprehensive data extraction"""
    try:
        workbook = load_workbook(uploaded_file, data_only=True)
        project_data = {
            'project_info': {},
            'revenues': {},
            'totals': {},
            'quarterly': {},
            'work_packages': {},
            'cash_flow_quarterly': {},
            'cash_flow_monthly': {},
            'cost_analysis': {},
            'earned_value': {},
            'risk_factors': []
        }
        
        st.info(f"📋 Parsing Template - Found sheets: {workbook.sheetnames}")
        
        # Parse Project Info (Sheet 1) - ENHANCED
        if '1_Project_Info' in workbook.sheetnames:
            ws_info = workbook['1_Project_Info']
            for row in range(2, 15):
                try:
                    field = ws_info.cell(row=row, column=1).value
                    value = ws_info.cell(row=row, column=2).value
                    if field and value is not None:
                        project_data['project_info'][str(field).strip()] = value
                except Exception:
                    continue
        
        # Parse Project Revenues (Sheet 2) - ENHANCED
        if '2_Project_Revenues' in workbook.sheetnames:
            ws_revenues = workbook['2_Project_Revenues']
            
            # PTD/MTD data (rows 2-7)
            for i, label in enumerate(REVENUE_LABELS, 2):
                try:
                    project_data['revenues'][label] = {
                        'n_ptd': safe_float(ws_revenues.cell(row=i, column=2).value),
                        'n1_ptd': safe_float(ws_revenues.cell(row=i, column=3).value),
                        'n_mtd': safe_float(ws_revenues.cell(row=i, column=4).value)
                    }
                except Exception:
                    project_data['revenues'][label] = {'n_ptd': 0, 'n1_ptd': 0, 'n_mtd': 0}
            
            # Calculate Cash In % if missing
            for period in ['n_ptd', 'n1_ptd', 'n_mtd']:
                try:
                    if project_data['revenues']['Cash In %'][period] == 0:
                        cash_in = project_data['revenues']['Cash IN'][period]
                        contract = project_data['revenues']['Contract Price'][period]
                        if contract > 0:
                            project_data['revenues']['Cash In %'][period] = (cash_in / contract) * 100
                except Exception:
                    continue

            # Parse quarterly revenue data if exists (rows 12-16)
            quarterly_row_mapping = {
            'Q1': 12,
            'Q2': 13,
            'Q3': 14,
            'Q4': 15,
            'Total': 16
        }

            # Debug flag for quarterly parsing
            quarterly_debug = st.checkbox("Show Quarterly Parsing Debug", value=False, key=f"quarterly_debug_{uploaded_file.name}")

            for quarter, expected_row in quarterly_row_mapping.items():
                try:
                    # Get the actual label from column A to verify we're reading the right row
                    actual_label = ws_revenues.cell(row=expected_row, column=1).value
        
                    if quarterly_debug:
                        st.write(f"Row {expected_row} - Expected: {quarter}, Found: {actual_label}")
        
                    # Check if this row contains quarterly data
                    if actual_label and (quarter in str(actual_label) or (quarter == 'Total' and 'total' in str(actual_label).lower())):
                        # Parse the quarterly values
                        actuals = safe_float(ws_revenues.cell(row=expected_row, column=2).value)
                        gap_to_close = safe_float(ws_revenues.cell(row=expected_row, column=3).value)
                        budget = safe_float(ws_revenues.cell(row=expected_row, column=4).value)
                        delta = safe_float(ws_revenues.cell(row=expected_row, column=5).value)
                        delta_pct = safe_float(ws_revenues.cell(row=expected_row, column=6).value)
            
                        # Store the data
                        project_data['quarterly'][quarter] = {
                            'actuals': actuals,
                            'gap_to_close': gap_to_close,
                            'budget': budget,
                            'delta': delta,
                            'delta_pct': delta_pct
                        }
            
                        # Debug output for verification
                        if quarterly_debug and quarter != 'Total':
                            st.write(f"✅ {quarter}: Actuals={actuals:,.0f}, Gap={gap_to_close:,.0f}, Budget={budget:,.0f}, Delta%={delta_pct:.1f}%")
                    else:
                        # Row doesn't match expected quarter - try to find it elsewhere
                        if quarterly_debug:
                            st.warning(f"⚠️ {quarter} not found at row {expected_row}, searching...")
            
                        # Search for the quarter in nearby rows (±2 rows)
                        found = False
                        for offset in [-2, -1, 1, 2]:
                            try:
                                search_row = expected_row + offset
                                search_label = ws_revenues.cell(row=search_row, column=1).value
                                if search_label and quarter in str(search_label):
                                    # Found the quarter at a different row
                                    project_data['quarterly'][quarter] = {
                                        'actuals': safe_float(ws_revenues.cell(row=search_row, column=2).value),
                                        'gap_to_close': safe_float(ws_revenues.cell(row=search_row, column=3).value),
                                        'budget': safe_float(ws_revenues.cell(row=search_row, column=4).value),
                                        'delta': safe_float(ws_revenues.cell(row=search_row, column=5).value),
                                        'delta_pct': safe_float(ws_revenues.cell(row=search_row, column=6).value)
                                    }
                                    if quarterly_debug:
                                        st.info(f"✅ Found {quarter} at row {search_row}")
                                    found = True
                                    break
                            except:
                                continue
            
                        if not found:
                            # Default to zeros if not found
                            project_data['quarterly'][quarter] = {
                                'actuals': 0, 'gap_to_close': 0, 'budget': 0, 'delta': 0, 'delta_pct': 0
                            }
                            if quarterly_debug:
                                st.error(f"❌ {quarter} data not found - using zeros")
                    
                except Exception as e:
                    if quarterly_debug:
                        st.error(f"Error parsing {quarter}: {str(e)}")
                    project_data['quarterly'][quarter] = {
                        'actuals': 0, 'gap_to_close': 0, 'budget': 0, 'delta': 0, 'delta_pct': 0
                    }

            # Validate quarterly data completeness
            if quarterly_debug:
                st.markdown("#### Quarterly Data Summary:")
                for q in ['Q1', 'Q2', 'Q3', 'Q4']:
                    q_data = project_data['quarterly'].get(q, {})
                    if q_data['actuals'] > 0 or q_data['budget'] > 0:
                        st.success(f"{q}: ✅ Has data")
                    else:
                        st.warning(f"{q}: ⚠️ No data")

        
        # Parse Cost Breakdown (Sheet 3) - COMPREHENSIVE WITH FIX #3
        if '3_Cost_Breakdown' in workbook.sheetnames:
            ws_costs = workbook['3_Cost_Breakdown']
            
            project_data['cost_analysis'] = {
                'total_as_sold': 0, 'total_committed': 0, 'total_fct_n': 0, 'total_actuals': 0,
                'ec_total_as_sold': 0, 'ec_total_fct_n': 0, 'ic_total_as_sold': 0, 'ic_total_fct_n': 0,
                'cm1_value_as_sold': 0, 'cm1_value_fct_n': 0, 'cm1_pct_as_sold': 0, 'cm1_pct_fct_n': 0,
                'cm2_value_as_sold': 0, 'cm2_value_fct_n': 0, 'cm2_pct_as_sold': 0, 'cm2_pct_fct_n': 0,
                'committed_ratio': 0, 'cost_variance_pct': 0,
                # FIX #3: Add FCT(n-1) values for CM variability analysis
                'cm1_value_fct_n1': 0, 'cm1_pct_fct_n1': 0,
                'cm2_value_fct_n1': 0, 'cm2_pct_fct_n1': 0,
                'ec_total_fct_n1': 0, 'ic_total_fct_n1': 0,
                # Add selling price for all periods
                'selling_price_as_sold': 0, 'selling_price_fct_n1': 0, 'selling_price_fct_n': 0,
                'total_fct_n1': 0
            }
            

            # Parse cost data with enhanced logic
            max_row = min(ws_costs.max_row, 50) if hasattr(ws_costs, 'max_row') else 30
            work_package_count = 0
            total_wp_value = 0
            
            # Debug flag to track what we're finding
            selling_price_found = False
            
            for excel_row in range(2, max_row + 1):
                try:
                    item_code = ws_costs.cell(row=excel_row, column=1).value
                    if not item_code:
                        continue
                        
                    item_str = str(item_code).strip()
                    item_upper = item_str.upper()
                    
                    # Get values for this row
                    description = str(ws_costs.cell(row=excel_row, column=2).value or '')
                    as_sold = safe_float(ws_costs.cell(row=excel_row, column=3).value)
                    committed = safe_float(ws_costs.cell(row=excel_row, column=4).value)
                    ctc = safe_float(ws_costs.cell(row=excel_row, column=5).value)
                    fct_n = safe_float(ws_costs.cell(row=excel_row, column=6).value)
                    fct_n1 = safe_float(ws_costs.cell(row=excel_row, column=7).value) if ws_costs.max_column >= 7 else 0
                    actuals = safe_float(ws_costs.cell(row=excel_row, column=9).value) if ws_costs.max_column >= 9 else 0
                    
                    # DEBUG: Show first few rows to see what's being parsed
                    if excel_row <= 5:
                        st.write(f"DEBUG Row {excel_row}: Item='{item_str}', Desc='{description[:30]}', AS={as_sold}")
                    
                    # Special handling for Selling Price - check both item code and description
                    if ('SELLING PRICE' in item_upper or 'SELLING PRICE' in description.upper()) and not selling_price_found:
                        project_data['cost_analysis']['selling_price_as_sold'] = as_sold
                        project_data['cost_analysis']['selling_price_fct_n1'] = fct_n1
                        project_data['cost_analysis']['selling_price_fct_n'] = fct_n
                        selling_price_found = True
                        st.success(f"✅ Found Selling Price: AS={as_sold}, FCT(n)={fct_n}, FCT(n-1)={fct_n1}")
                        continue
                    
                    # Identify and store work packages
                    if not any(keyword in item_upper for keyword in ['TOTAL', 'CM1', 'CM2', 'SELLING']):
                        is_risk_contingency = 'risk' in description.lower() and 'contingenc' in description.lower()
                        work_package = {
                            'code': item_str,
                            'description': description,
                            'as_sold': as_sold,
                            'committed': committed,
                            'ctc': ctc,
                            'fct_n': fct_n,
                            'fct_n1': fct_n1,
                            'actuals': actuals,
                            'variance_pct': calculate_period_variance(fct_n, as_sold) if as_sold > 0 else 0,
                            'commitment_ratio': committed / as_sold if as_sold > 0 else 0,
			    'is_risk_contingency': is_risk_contingency
                        }

                        project_data['work_packages'][item_str] = work_package
                        
                        # Track major work packages (>10% of total value)
                        if as_sold > 0:
                            work_package_count += 1
                            total_wp_value += as_sold
                    
                    # Parse summary rows with exact matching
                    if item_upper == 'TOTAL':
                        project_data['cost_analysis']['total_as_sold'] = as_sold
                        project_data['cost_analysis']['total_committed'] = committed
                        project_data['cost_analysis']['total_fct_n'] = fct_n
                        project_data['cost_analysis']['total_fct_n1'] = fct_n1
                        project_data['cost_analysis']['total_actuals'] = actuals
                        project_data['cost_analysis']['cost_variance_pct'] = calculate_period_variance(fct_n, as_sold)
                    
                    # Parse EC (External Costs) totals - check for "Total EC" specifically
                    elif ('TOTAL EC' in item_upper) or ('EC' in item_upper and 'TOTAL' in item_upper):
                        project_data['cost_analysis']['ec_total_as_sold'] = as_sold
                        project_data['cost_analysis']['ec_total_fct_n1'] = fct_n1
                        project_data['cost_analysis']['ec_total_fct_n'] = fct_n
                    
                    # Parse IC/IL (Internal Costs/Labour) totals
                    elif ('TOTAL IL' in item_upper) or ('IL' in item_upper and 'TOTAL' in item_upper) or ('IC' in item_upper and 'TOTAL' in item_upper):
                        project_data['cost_analysis']['ic_total_as_sold'] = as_sold
                        project_data['cost_analysis']['ic_total_fct_n1'] = fct_n1
                        project_data['cost_analysis']['ic_total_fct_n'] = fct_n
                        
                except Exception as e:
                    st.warning(f"Error parsing row {excel_row}: {str(e)}")
                    continue
            
            # Mark if selling price was found
            project_data['cost_analysis']['selling_price_found'] = selling_price_found

        # Enhanced cost analysis calculations
        contract_value = safe_get_value(project_data, 'revenues', 'Contract Price', 'n_ptd')

        # Get selling prices for all periods
        # Check if selling price was actually found in the template
        if not project_data['cost_analysis'].get('selling_price_found', False):
            # Selling price row was not found in the template
            if project_data['cost_analysis']['total_as_sold'] > 0:
                # We have cost data but no selling price - this is a problem
                st.error("❌ Selling Price row not found in template - please check your Excel template")
                st.info("Expected to find a row with 'Selling Price' in the cost breakdown sheet")
                # Use contract value as emergency fallback
                project_data['cost_analysis']['selling_price_as_sold'] = contract_value
                project_data['cost_analysis']['selling_price_fct_n1'] = contract_value
                project_data['cost_analysis']['selling_price_fct_n'] = contract_value
            else:
                # No cost data at all - use contract value as fallback
                project_data['cost_analysis']['selling_price_as_sold'] = contract_value
                project_data['cost_analysis']['selling_price_fct_n1'] = contract_value
                project_data['cost_analysis']['selling_price_fct_n'] = contract_value


# CALCULATE CM1 and CM2 for all periods
# CM1 = Selling Price - Total EC
# CM2 = CM1 - Total IC = Selling Price - Total EC - Total IC

# AS SOLD period
        selling_price_as = project_data['cost_analysis']['selling_price_as_sold']
        if selling_price_as > 0:
            ec_as = project_data['cost_analysis']['ec_total_as_sold']
            ic_as = project_data['cost_analysis']['ic_total_as_sold']
    
            # Calculate CM1 AS
            project_data['cost_analysis']['cm1_value_as_sold'] = selling_price_as - ec_as
            project_data['cost_analysis']['cm1_pct_as_sold'] = ((selling_price_as - ec_as) / selling_price_as * 100) if selling_price_as > 0 else 0
    
            # Calculate CM2 AS
            project_data['cost_analysis']['cm2_value_as_sold'] = selling_price_as - ec_as - ic_as
            project_data['cost_analysis']['cm2_pct_as_sold'] = ((selling_price_as - ec_as - ic_as) / selling_price_as * 100) if selling_price_as > 0 else 0

        # FCT(n-1) period
        selling_price_n1 = project_data['cost_analysis']['selling_price_fct_n1']
        if selling_price_n1 > 0:
            ec_n1 = project_data['cost_analysis']['ec_total_fct_n1']
            ic_n1 = project_data['cost_analysis']['ic_total_fct_n1']
    
            # Calculate CM1 FCT(n-1)
            project_data['cost_analysis']['cm1_value_fct_n1'] = selling_price_n1 - ec_n1
            project_data['cost_analysis']['cm1_pct_fct_n1'] = ((selling_price_n1 - ec_n1) / selling_price_n1 * 100) if selling_price_n1 > 0 else 0
    
            # Calculate CM2 FCT(n-1)
            project_data['cost_analysis']['cm2_value_fct_n1'] = selling_price_n1 - ec_n1 - ic_n1
            project_data['cost_analysis']['cm2_pct_fct_n1'] = ((selling_price_n1 - ec_n1 - ic_n1) / selling_price_n1 * 100) if selling_price_n1 > 0 else 0

        # FCT(n) period
        selling_price_n = project_data['cost_analysis']['selling_price_fct_n']
        if selling_price_n > 0:
            ec_n = project_data['cost_analysis']['ec_total_fct_n']
            ic_n = project_data['cost_analysis']['ic_total_fct_n']
    
            # Calculate CM1 FCT(n)
            project_data['cost_analysis']['cm1_value_fct_n'] = selling_price_n - ec_n
            project_data['cost_analysis']['cm1_pct_fct_n'] = ((selling_price_n - ec_n) / selling_price_n * 100) if selling_price_n > 0 else 0
    
            # Calculate CM2 FCT(n)
            project_data['cost_analysis']['cm2_value_fct_n'] = selling_price_n - ec_n - ic_n
            project_data['cost_analysis']['cm2_pct_fct_n'] = ((selling_price_n - ec_n - ic_n) / selling_price_n * 100) if selling_price_n > 0 else 0

        if project_data['cost_analysis']['total_as_sold'] > 0:
            project_data['cost_analysis']['committed_ratio'] = (
                project_data['cost_analysis']['total_committed'] / 
                project_data['cost_analysis']['total_as_sold']
            )
        
        # Parse Cash Flow (Sheet 4) - QUARTERLY ONLY
        if '4_Cash_Flow' in workbook.sheetnames:
            ws_cashflow = workbook['4_Cash_Flow']
            
            # Parse quarterly cash flow data only
            max_row = min(ws_cashflow.max_row, 30) if hasattr(ws_cashflow, 'max_row') else 30
            
            for row in range(2, max_row + 1):
                try:
                    period = ws_cashflow.cell(row=row, column=1).value
                    if not period:
                        continue
                        
                    period_str = str(period).strip()
                    
                    # Quarterly data (FY format only)
                    if 'FY' in period_str:
                        project_data['cash_flow_quarterly'][period_str] = {
                            'as_sold': safe_float(ws_cashflow.cell(row=row, column=2).value),
                            'fct_n1': safe_float(ws_cashflow.cell(row=row, column=3).value),
                            'fct_n': safe_float(ws_cashflow.cell(row=row, column=4).value),
                            'variance_n_vs_sold': 0,
                            'variance_n_vs_n1': 0
                        }
                except Exception:
                    continue
            
            # Calculate variances for quarterly data
            for quarter, data in project_data['cash_flow_quarterly'].items():
                if data['as_sold'] != 0:
                    data['variance_n_vs_sold'] = calculate_period_variance(data['fct_n'], data['as_sold'])
                if data['fct_n1'] != 0:
                    data['variance_n_vs_n1'] = calculate_period_variance(data['fct_n'], data['fct_n1'])
        
        # Calculate Earned Value Metrics
        project_data['earned_value'] = calculate_earned_value_metrics(project_data)
        
        # Risk Assessment
        project_data['risk_factors'] = assess_project_risks(project_data)
        
        # Success summary
        contract_value = safe_get_value(project_data, 'revenues', 'Contract Price', 'n_ptd')
        cm2_pct = project_data['cost_analysis']['cm2_pct_fct_n']
        
        st.success(f"✅ Template Analysis Complete:")
        st.info(f"• **Contract Value:** CHF {contract_value:,.0f}")
        st.info(f"• **CM2 Margin:** {cm2_pct:.1f}%")
        st.info(f"• **Quarterly Cash Flow:** {len(project_data['cash_flow_quarterly'])} periods")
        st.info(f"• **Work Packages:** {len(project_data['work_packages'])} items")
        st.info(f"• **Risk Factors:** {len(project_data['risk_factors'])} identified")
        
        # Enhanced CM data verification
        st.markdown("#### 📊 Margin Data Verification:")
        cm_data_found = False

        # Debug: Show what was parsed
        if st.checkbox("Show detailed parsing results", value=False, key=f"debug_parsing_{project_data['project_info'].get('Project No.', 'unknown')}"):
            st.write("**Raw Parsed Values:**")
            st.write(f"• Selling Price - AS: {project_data['cost_analysis']['selling_price_as_sold']:,.0f}, FCT(n): {project_data['cost_analysis']['selling_price_fct_n']:,.0f}, FCT(n-1): {project_data['cost_analysis']['selling_price_fct_n1']:,.0f}")
            st.write(f"• EC Total - AS: {project_data['cost_analysis']['ec_total_as_sold']:,.0f}, FCT(n): {project_data['cost_analysis']['ec_total_fct_n']:,.0f}, FCT(n-1): {project_data['cost_analysis']['ec_total_fct_n1']:,.0f}")
            st.write(f"• IC/IL Total - AS: {project_data['cost_analysis']['ic_total_as_sold']:,.0f}, FCT(n): {project_data['cost_analysis']['ic_total_fct_n']:,.0f}, FCT(n-1): {project_data['cost_analysis']['ic_total_fct_n1']:,.0f}")
            st.write(f"• Total Costs - AS: {project_data['cost_analysis']['total_as_sold']:,.0f}, FCT(n): {project_data['cost_analysis']['total_fct_n']:,.0f}, FCT(n-1): {project_data['cost_analysis']['total_fct_n1']:,.0f}")
            st.write(f"• Contract Value: {contract_value:,.0f}")
        
        # Check CM1 data
        if project_data['cost_analysis']['cm1_pct_as_sold'] != 0 or project_data['cost_analysis']['cm1_pct_fct_n'] != 0:
            st.success(f"✅ **CM1 Data Calculated:**")
            st.write(f"   • CM1%: AS={project_data['cost_analysis']['cm1_pct_as_sold']:.2f}%, FCT(n)={project_data['cost_analysis']['cm1_pct_fct_n']:.2f}%, FCT(n-1)={project_data['cost_analysis']['cm1_pct_fct_n1']:.2f}%")
            st.write(f"   • CM1 Value: AS={project_data['cost_analysis']['cm1_value_as_sold']:,.0f}, FCT(n)={project_data['cost_analysis']['cm1_value_fct_n']:,.0f}, FCT(n-1)={project_data['cost_analysis']['cm1_value_fct_n1']:,.0f}")
            cm_data_found = True
        else:
            st.warning("⚠️ CM1 data could not be calculated - missing EC data")
            
        # Check CM2 data
        if project_data['cost_analysis']['cm2_pct_as_sold'] != 0 or project_data['cost_analysis']['cm2_pct_fct_n'] != 0:
            st.success(f"✅ **CM2 Data Calculated:**")
            st.write(f"   • CM2%: AS={project_data['cost_analysis']['cm2_pct_as_sold']:.2f}%, FCT(n)={project_data['cost_analysis']['cm2_pct_fct_n']:.2f}%, FCT(n-1)={project_data['cost_analysis']['cm2_pct_fct_n1']:.2f}%")
            st.write(f"   • CM2 Value: AS={project_data['cost_analysis']['cm2_value_as_sold']:,.0f}, FCT(n)={project_data['cost_analysis']['cm2_value_fct_n']:,.0f}, FCT(n-1)={project_data['cost_analysis']['cm2_value_fct_n1']:,.0f}")
        # Display EC/IC breakdown
            st.markdown("**Cost Breakdown Verification:**")
            st.write(f"   • EC: AS={project_data['cost_analysis']['ec_total_as_sold']:,.0f}, FCT(n)={project_data['cost_analysis']['ec_total_fct_n']:,.0f}, FCT(n-1)={project_data['cost_analysis']['ec_total_fct_n1']:,.0f}")
            st.write(f"   • IC: AS={project_data['cost_analysis']['ic_total_as_sold']:,.0f}, FCT(n)={project_data['cost_analysis']['ic_total_fct_n']:,.0f}, FCT(n-1)={project_data['cost_analysis']['ic_total_fct_n1']:,.0f}")
            st.write(f"   • Selling Price: AS={project_data['cost_analysis']['selling_price_as_sold']:,.0f}, FCT(n)={project_data['cost_analysis']['selling_price_fct_n']:,.0f}, FCT(n-1)={project_data['cost_analysis']['selling_price_fct_n1']:,.0f}")
            cm_data_found = True
        else:
            st.warning("⚠️ CM2 data could not be calculated - missing EC/IC data")
            
        if cm_data_found:
            st.info("💡 Margin Variability Analysis will be available for this project")
        else:
            st.warning("❌ Margin Variability Analysis will NOT be available - missing cost breakdown data")
        
        return project_data
        
    except Exception as e:
        st.error(f"❌ Error parsing template: {str(e)}")
        st.exception(e)
        return None

def assess_project_risks(project_data):
    """Simplified project risk assessment with dynamic thresholds"""
    risk_factors = []
    
    try:
        # Cost and margin risks
        cost_analysis = project_data.get('cost_analysis', {})
        cm2_pct = cost_analysis.get('cm2_pct_fct_n', 0)
        committed_ratio = cost_analysis.get('committed_ratio', 0)
        cost_variance = cost_analysis.get('cost_variance_pct', 0)
        
        # Get current CM2 thresholds from EXECUTIVE_THRESHOLDS
        cm2_excellent = EXECUTIVE_THRESHOLDS['cm2_margin']['excellent']
        cm2_good = EXECUTIVE_THRESHOLDS['cm2_margin']['good']
        cm2_warning = EXECUTIVE_THRESHOLDS['cm2_margin']['warning']
        
        # Margin risks using dynamic thresholds
        if cm2_pct < cm2_warning:  # Below warning threshold is critical
            risk_factors.append({
                'type': 'Margin Risk',
                'severity': 'Critical',
                'description': f'CM2 margin critically low at {cm2_pct:.1f}% (below warning threshold: {cm2_warning}%)',
                'impact': 'High',
                'recommendation': 'Immediate cost reduction and revenue optimization required'
            })
        elif cm2_pct < cm2_good:  # Between warning and good is high risk
            risk_factors.append({
                'type': 'Margin Risk',
                'severity': 'High',
                'description': f'CM2 margin below target at {cm2_pct:.1f}% (target: {cm2_good}%)',
                'impact': 'Medium',
                'recommendation': 'Review cost structure and identify optimization opportunities'
            })
        elif cm2_pct < cm2_excellent:  # Between good and excellent is medium risk
            risk_factors.append({
                'type': 'Margin Risk',
                'severity': 'Medium',
                'description': f'CM2 margin at {cm2_pct:.1f}% - room for improvement (excellent: {cm2_excellent}%)',
                'impact': 'Low',
                'recommendation': 'Continue monitoring and seek margin enhancement opportunities'
            })
        # If cm2_pct >= cm2_excellent, no margin risk is added
        
        # Cost commitment risks (keeping existing logic)
        if committed_ratio > 1.2:
            risk_factors.append({
                'type': 'Cost Commitment',
                'severity': 'Critical',
                'description': f'Severe cost overcommitment: {committed_ratio:.2f} ratio',
                'impact': 'High',
                'recommendation': 'Emergency cost review and procurement controls'
            })
        elif committed_ratio > 1.1:
            risk_factors.append({
                'type': 'Cost Commitment',
                'severity': 'High',
                'description': f'High cost commitment: {committed_ratio:.2f} ratio',
                'impact': 'Medium',
                'recommendation': 'Enhanced cost monitoring and approval processes'
            })
        
        # Cost variance risks (keeping existing logic)
        if cost_variance > 25:
            risk_factors.append({
                'type': 'Cost Variance',
                'severity': 'Critical',
                'description': f'Extreme cost variance: {cost_variance:+.1f}%',
                'impact': 'High',
                'recommendation': 'Comprehensive cost baseline review required'
            })
        elif cost_variance > 15:
            risk_factors.append({
                'type': 'Cost Variance',
                'severity': 'High',
                'description': f'High cost variance: {cost_variance:+.1f}%',
                'impact': 'Medium',
                'recommendation': 'Detailed variance analysis and corrective action plan'
            })
        
        # Schedule and POC risks (keeping existing logic)
        poc_current = safe_get_value(project_data, 'revenues', 'POC%', 'n_ptd')
        poc_previous = safe_get_value(project_data, 'revenues', 'POC%', 'n1_ptd')
        poc_velocity = calculate_poc_velocity(poc_current, poc_previous)
        
        if poc_velocity < 2 and poc_current < 90:
            risk_factors.append({
                'type': 'Schedule Risk',
                'severity': 'High',
                'description': f'Low POC velocity: {poc_velocity:.1f}%/month',
                'impact': 'Medium',
                'recommendation': 'Resource reallocation and schedule acceleration'
            })
        
        # Cash flow risks (keeping existing logic)
        quarterly_data = project_data.get('cash_flow_quarterly', {})
        if quarterly_data:
            negative_quarters = sum(1 for q in quarterly_data.values() if q['fct_n'] < 0)
            total_quarters = len(quarterly_data)
            if negative_quarters > total_quarters * 0.3:
                risk_factors.append({
                    'type': 'Cash Flow',
                    'severity': 'High',
                    'description': f'Multiple negative cash flow quarters: {negative_quarters}/{total_quarters}',
                    'impact': 'High',
                    'recommendation': 'Cash flow optimization and milestone acceleration'
                })
        
        # Revenue risks (keeping existing logic)
        revenue_current = safe_get_value(project_data, 'revenues', 'Revenues', 'n_ptd')
        revenue_previous = safe_get_value(project_data, 'revenues', 'Revenues', 'n1_ptd')
        revenue_variance = calculate_period_variance(revenue_current, revenue_previous)
        
        if revenue_variance < -15:
            risk_factors.append({
                'type': 'Revenue Risk',
                'severity': 'Critical',
                'description': f'Significant revenue decline: {revenue_variance:.1f}%',
                'impact': 'High',
                'recommendation': 'Revenue recovery plan and stakeholder engagement'
            })
        
        # NEW: Risk Contingency Adequacy Assessment (REPLACES Work Package Risk)
        work_packages = project_data.get('work_packages', {})
        risk_contingencies = [wp for wp in work_packages.values() 
                             if 'risk' in wp.get('description', '').lower() and 
                             'contingenc' in wp.get('description', '').lower()]
        
        contract_value = safe_get_value(project_data, 'revenues', 'Contract Price', 'n_ptd')
        
        if risk_contingencies:
            # Sum all risk contingency values
            total_risk_contingency = sum(rc.get('fct_n', 0) for rc in risk_contingencies)
            contingency_percentage = (total_risk_contingency / contract_value * 100) if contract_value > 0 else 0
            
            # Check if contingency is being consumed
            original_contingency = sum(rc.get('as_sold', 0) for rc in risk_contingencies)
            contingency_consumption = ((original_contingency - total_risk_contingency) / original_contingency * 100) if original_contingency > 0 else 0
            
            # Risk Contingency Adequacy Thresholds
            if contingency_percentage < 1:  # Less than 1% contingency remaining
                risk_factors.append({
                    'type': 'Contingency Risk',
                    'severity': 'Medium',
                    'description': f'Insufficient risk contingency: {contingency_percentage:.1f}% of contract value',
                    'impact': 'Medium',
                    'recommendation': 'Review risk register and consider contingency replenishment'
                })
            elif contingency_percentage < 5 and contingency_consumption > 80:  # <5% remaining AND >80% consumed
                risk_factors.append({
                    'type': 'Contingency Risk', 
                    'severity': 'High',
                    'description': f'Low contingency: {contingency_percentage:.1f}% remaining, {contingency_consumption:.0f}% consumed',
                    'impact': 'High',
                    'recommendation': 'Monitor emerging risks closely, prepare contingency plan'
                })

        else:
            # No risk contingency found at all
            risk_factors.append({
                'type': 'Contingency Risk',
                'severity': 'High',
                'description': 'No risk contingency identified in project structure',
                'impact': 'High',
                'recommendation': 'Establish risk contingency budget for unforeseen events'
            })
        
        # OPTIONAL: Add Work Package Performance Concentration Risk
        high_variance_wps = [wp for wp in work_packages.values() 
                            if wp.get('variance_pct', 0) > 15 and
                            not ('risk' in wp.get('description', '').lower() and 
                                 'contingenc' in wp.get('description', '').lower())]
        
        if len(work_packages) > 0 and len(high_variance_wps) / len(work_packages) > 0.3:  # >30% of WPs have issues
            risk_factors.append({
                'type': 'WP Performance Risk',
                'severity': 'High',
                'description': f'{len(high_variance_wps)} of {len(work_packages)} work packages exceeding budget by >15%',
                'impact': 'High',
                'recommendation': 'Systemic issue - review estimation or execution processes'
            })
        
        # OPTIONAL: Add Financial Buffer Risk
        if risk_contingencies and contingency_percentage < 3 and cm2_pct < 10:
            risk_factors.append({
                'type': 'Financial Buffer Risk',
                'severity': 'Critical',
                'description': f'Low contingency ({contingency_percentage:.1f}%) combined with thin margins ({cm2_pct:.1f}%)',
                'impact': 'High',
                'recommendation': 'Project has minimal financial buffer for risks'
            })
        
    except Exception as e:
        risk_factors.append({
            'type': 'Assessment Error',
            'severity': 'Medium',
            'description': f'Risk assessment incomplete: {str(e)}',
            'impact': 'Low',
            'recommendation': 'Manual risk review recommended'
        })
    
    return risk_factors

# ================================================================================
# COMPREHENSIVE DASHBOARD RENDERING FUNCTIONS (SIMPLIFIED)
# ================================================================================

def render_executive_kpi_dashboard(portfolio_summary):
    """Render comprehensive executive-level KPI dashboard"""
    st.markdown("## 🎯 Executive Performance Dashboard")
    
    # Calculate contingency efficiency for portfolio if not already in summary
    if 'portfolio_contingency_efficiency' not in portfolio_summary:
        # This would need to be calculated in create_enhanced_portfolio_summary
        portfolio_contingency_efficiency = portfolio_summary.get('portfolio_contingency_efficiency', 100)
    else:
        portfolio_contingency_efficiency = portfolio_summary['portfolio_contingency_efficiency']
    
    # Top-level Executive KPIs
    col1, col2, col3, col4, col5 = st.columns(5)
    
    with col1:
        # Portfolio Value KPI
        contract_var = portfolio_summary.get('contract_variance', 0)
        value_icon, value_status, value_class = get_traffic_light_status(
            contract_var, EXECUTIVE_THRESHOLDS['revenue_growth']
        )
        
        st.markdown(f"""
        <div class="executive-kpi traffic-light-{value_class}">
            <h3>Portfolio Value</h3>
            <div class="kpi-value">{format_currency_millions(portfolio_summary['total_contract_value'])}</div>
            <div class="kpi-trend status-{value_class}">{value_icon} {value_status} ({contract_var:+.1f}%)</div>
        </div>
        """, unsafe_allow_html=True)
    
    with col2:
        # CM2 Margin KPI
        avg_cm2 = portfolio_summary.get('avg_cm2_pct', 0)
        cm2_icon, cm2_status, cm2_class = get_traffic_light_status(
            avg_cm2, EXECUTIVE_THRESHOLDS['cm2_margin']
        )
        
        st.markdown(f"""
        <div class="executive-kpi traffic-light-{cm2_class}">
            <h3>CM2 Margin</h3>
            <div class="kpi-value">{avg_cm2:.1f}%</div>
            <div class="kpi-trend status-{cm2_class}">{cm2_icon} {cm2_status}</div>
        </div>
        """, unsafe_allow_html=True)
    
    with col3:
        # Cost Performance Index
        avg_cpi = portfolio_summary.get('avg_cost_performance_index', 1.0)
        cpi_icon, cpi_status, cpi_class = get_traffic_light_status(
            avg_cpi, EXECUTIVE_THRESHOLDS['cost_performance_index']
        )
        
        st.markdown(f"""
        <div class="executive-kpi traffic-light-{cpi_class}">
            <h3>Cost Performance</h3>
            <div class="kpi-value">{avg_cpi:.2f}</div>
            <div class="kpi-trend status-{cpi_class}">{cpi_icon} {cpi_status}</div>
        </div>
        """, unsafe_allow_html=True)
    
    with col4:
        # POC Velocity
        poc_velocity = portfolio_summary.get('weighted_poc_velocity', 0)
        poc_icon, poc_status, poc_class = get_traffic_light_status(
            poc_velocity, EXECUTIVE_THRESHOLDS['poc_velocity']
        )
        
        st.markdown(f"""
        <div class="executive-kpi traffic-light-{poc_class}">
            <h3>POC Velocity</h3>
            <div class="kpi-value">{poc_velocity:.1f}%/mo</div>
            <div class="kpi-trend status-{poc_class}">{poc_icon} {poc_status}</div>
        </div>
        """, unsafe_allow_html=True)
    
    with col5:
        # Contingency Efficiency (replacing Risk Score)
        if portfolio_contingency_efficiency is not None:
            if portfolio_contingency_efficiency >= 150:
                cont_icon = "🟢"
                cont_status = "Excellent"
                cont_class = "excellent"
            elif portfolio_contingency_efficiency >= 120:
                cont_icon = "🟢"
                cont_status = "Good"
                cont_class = "good"
            elif portfolio_contingency_efficiency >= 80:
                cont_icon = "🟦"
                cont_status = "On Track"
                cont_class = "good"
            elif portfolio_contingency_efficiency >= 50:
                cont_icon = "🟡"
                cont_status = "Warning"
                cont_class = "warning"
            else:
                cont_icon = "🔴"
                cont_status = "Critical"
                cont_class = "critical"
            
            st.markdown(f"""
            <div class="executive-kpi traffic-light-{cont_class}">
                <h3>Contingency Efficiency</h3>
                <div class="kpi-value">{portfolio_contingency_efficiency:.0f}%</div>
                <div class="kpi-trend status-{cont_class}">{cont_icon} {cont_status}</div>
            </div>
            """, unsafe_allow_html=True)
        else:
            # Fallback to Risk Score if no contingency data
            avg_risk_score = portfolio_summary.get('average_risk_score', 0)
            risk_icon = "🟢" if avg_risk_score <= 3 else "🟡" if avg_risk_score <= 6 else "🔴"
            risk_status = "Low" if avg_risk_score <= 3 else "Medium" if avg_risk_score <= 6 else "High"
            
            st.markdown(f"""
            <div class="executive-kpi traffic-light-{'good' if avg_risk_score <= 3 else 'warning' if avg_risk_score <= 6 else 'critical'}">
                <h3>Portfolio Risk</h3>
                <div class="kpi-value">{avg_risk_score:.1f}/10</div>
                <div class="kpi-trend">{risk_icon} {risk_status} Risk</div>
            </div>
            """, unsafe_allow_html=True)

def render_enhanced_margin_analysis(portfolio_data):
    """Render enhanced margin analysis with EC/IC breakdown - CLEAN VERSION"""
    st.markdown("## 📊 Enhanced Margin Analysis")
    
    # Calculate portfolio-level margin metrics
    total_contract = 0
    total_ec_fct_n = 0
    total_ic_fct_n = 0
    total_cm1_value_fct_n = 0
    total_cm2_value_fct_n = 0
    projects_with_margin_data = 0
    
    # Add contingency tracking
    total_contingency_as_sold = 0
    total_contingency_fct_n = 0
    total_poc_weighted = 0
    projects_with_contingency = 0
    
    margin_projects = []
    
    # Process each project
    for project_id, project in portfolio_data.items():
        try:
            if 'cost_analysis' in project['data']:
                cost_data = project['data']['cost_analysis']
                contract_value = safe_get_value(project['data'], 'revenues', 'Contract Price', 'n_ptd')
                poc_current = safe_get_value(project['data'], 'revenues', 'POC%', 'n_ptd')
                
                if contract_value > 0:
                    total_contract += contract_value
                    total_ec_fct_n += cost_data.get('ec_total_fct_n', 0)
                    total_ic_fct_n += cost_data.get('ic_total_fct_n', 0)
                    total_cm1_value_fct_n += cost_data.get('cm1_value_fct_n', 0)
                    total_cm2_value_fct_n += cost_data.get('cm2_value_fct_n', 0)
                    projects_with_margin_data += 1
                    
                    # Calculate contingency metrics for this project
                    work_packages = project['data'].get('work_packages', {})
                    contingency_metrics = calculate_contingency_metrics(work_packages, poc_current)
                    
                    if contingency_metrics['has_contingency']:
                        total_contingency_as_sold += contingency_metrics['contingency_as_sold']
                        total_contingency_fct_n += contingency_metrics['contingency_fct_n']
                        total_poc_weighted += poc_current * contract_value
                        projects_with_contingency += 1
                    
                    margin_projects.append({
                        'project_id': project_id,
                        'project_name': project['name'],
                        'contract_value': contract_value,
                        'ec_total': cost_data.get('ec_total_fct_n', 0),
                        'ic_total': cost_data.get('ic_total_fct_n', 0),
                        'cm1_pct': cost_data.get('cm1_pct_fct_n', 0),
                        'cm2_pct': cost_data.get('cm2_pct_fct_n', 0),
                        'committed_ratio': cost_data.get('committed_ratio', 0),
                        'cost_variance_pct': cost_data.get('cost_variance_pct', 0),
                        'contingency_efficiency': contingency_metrics['efficiency'] if contingency_metrics['has_contingency'] else None,
                        'contingency_status': contingency_metrics['status_icon'] if contingency_metrics['has_contingency'] else '➖'
                    })
        except Exception as e:
            continue  # Skip problematic projects
    
    if projects_with_margin_data == 0:
        st.warning("📊 No margin analysis data available.")
        return
    
    # Portfolio margin summary
    portfolio_cm1_pct = (total_cm1_value_fct_n / total_contract * 100) if total_contract > 0 else 0
    portfolio_cm2_pct = (total_cm2_value_fct_n / total_contract * 100) if total_contract > 0 else 0
    portfolio_ec_pct = (total_ec_fct_n / total_contract * 100) if total_contract > 0 else 0
    portfolio_ic_pct = (total_ic_fct_n / total_contract * 100) if total_contract > 0 else 0
    
    # Calculate portfolio contingency efficiency
    portfolio_contingency_consumed = total_contingency_as_sold - total_contingency_fct_n
    portfolio_contingency_consumed_pct = (portfolio_contingency_consumed / total_contingency_as_sold * 100) if total_contingency_as_sold > 0 else 0
    portfolio_avg_poc = (total_poc_weighted / total_contract) if total_contract > 0 else 0
    
    if portfolio_avg_poc > 0 and total_contingency_as_sold > 0:
        portfolio_contingency_efficiency = (2 - (portfolio_contingency_consumed_pct / portfolio_avg_poc)) * 100
        portfolio_contingency_efficiency = max(0, min(200, portfolio_contingency_efficiency))
    else:
        portfolio_contingency_efficiency = None
    
    # Display margin analysis cards
    col1, col2, col3 = st.columns(3)
    
    with col1:
        # CM1 Analysis
        cm1_icon, cm1_status, cm1_class = get_traffic_light_status(
            portfolio_cm1_pct, EXECUTIVE_THRESHOLDS['cm1_margin']
        )
        
        st.markdown(f"""
        <div class="margin-card margin-{cm1_class}">
            <h4>CM1 (After External Costs)</h4>
            <h2>{portfolio_cm1_pct:.1f}% {cm1_icon}</h2>
            <p><strong>Status:</strong> {cm1_status}</p>
            <p><strong>Value:</strong> {format_currency_millions(total_cm1_value_fct_n)}</p>
            <p><strong>External Costs:</strong> {portfolio_ec_pct:.1f}% of contract</p>
        </div>
        """, unsafe_allow_html=True)
    
    with col2:
        # CM2 Analysis
        cm2_icon, cm2_status, cm2_class = get_traffic_light_status(
            portfolio_cm2_pct, EXECUTIVE_THRESHOLDS['cm2_margin']
        )
        
        st.markdown(f"""
        <div class="margin-card margin-{cm2_class}">
            <h4>CM2 (After All Costs)</h4>
            <h2>{portfolio_cm2_pct:.1f}% {cm2_icon}</h2>
            <p><strong>Status:</strong> {cm2_status}</p>
            <p><strong>Value:</strong> {format_currency_millions(total_cm2_value_fct_n)}</p>
            <p><strong>Internal Costs:</strong> {portfolio_ic_pct:.1f}% of contract</p>
        </div>
        """, unsafe_allow_html=True)
    
    with col3:
        # Contingency Efficiency Analysis (replacing Cost Structure)
        if portfolio_contingency_efficiency is not None:
            # Determine status for portfolio contingency
            if portfolio_contingency_efficiency >= 150:
                cont_class = 'excellent'
                cont_icon = '🟢'
                cont_status = 'Excellent'
            elif portfolio_contingency_efficiency >= 120:
                cont_class = 'good'
                cont_icon = '🟢'
                cont_status = 'Good'
            elif portfolio_contingency_efficiency >= 80:
                cont_class = 'good'  # Using 'good' for blue since we don't have 'info' style
                cont_icon = '🟦'
                cont_status = 'On Track'
            elif portfolio_contingency_efficiency >= 50:
                cont_class = 'warning'
                cont_icon = '🟡'
                cont_status = 'Warning'
            else:
                cont_class = 'critical'
                cont_icon = '🔴'
                cont_status = 'Critical'
            
            st.markdown(f"""
            <div class="margin-card margin-{cont_class}">
                <h4>Contingency Efficiency</h4>
                <h2>{portfolio_contingency_efficiency:.0f}% {cont_icon}</h2>
                <p><strong>Status:</strong> {cont_status}</p>
                <p><strong>Consumed:</strong> {portfolio_contingency_consumed_pct:.1f}% of {format_currency_millions(total_contingency_as_sold)}</p>
                <p><strong>Remaining:</strong> {format_currency_millions(total_contingency_fct_n)}</p>
                <p><strong>Projects with Contingency:</strong> {projects_with_contingency}</p>
            </div>
            """, unsafe_allow_html=True)
        else:
            st.markdown(f"""
            <div class="margin-card margin-good">
                <h4>Contingency Efficiency</h4>
                <h2>No Data ➖</h2>
                <p><strong>Status:</strong> No contingency allocated</p>
                <p><strong>Total Portfolio:</strong> {format_currency_millions(total_contract)}</p>
                <p><strong>Projects:</strong> {projects_with_margin_data}</p>
            </div>
            """, unsafe_allow_html=True)
    
    # Enhanced Margin Analysis Chart
    if margin_projects:
        try:
            fig = create_comprehensive_margin_chart(margin_projects)
            st.plotly_chart(fig, use_container_width=True)
        except Exception as e:
            st.warning(f"Chart creation failed: {str(e)}")
    
    # Updated project margin performance table with contingency efficiency
    st.markdown("### 📋 Project Margin Performance")
    
    margin_summary = []
    for project in margin_projects:
        try:
            cm2_icon, _, cm2_class = get_traffic_light_status(project['cm2_pct'], EXECUTIVE_THRESHOLDS['cm2_margin'])
            committed_icon, _, _ = get_traffic_light_status(project['committed_ratio'], EXECUTIVE_THRESHOLDS['committed_vs_budget'], reverse=True)
            
            cont_eff_display = f"{project['contingency_efficiency']:.0f}% {project['contingency_status']}" if project['contingency_efficiency'] is not None else "N/A ➖"
            
            margin_summary.append({
                'Project': project['project_id'],
                'Contract Value': format_currency_millions(project['contract_value']),
                'CM1 %': f"{project['cm1_pct']:.1f}%",
                'CM2 %': f"{project['cm2_pct']:.1f}%",
                'CM2 Status': f"{cm2_icon}",
                'Cost Variance': f"{project['cost_variance_pct']:+.1f}%",
                'Committed Ratio': f"{project['committed_ratio']:.2f} {committed_icon}",
                'Contingency Eff.': cont_eff_display
            })
        except Exception as e:
            continue  # Skip problematic entries
    
    if margin_summary:
        df_margin = pd.DataFrame(margin_summary)
        st.dataframe(df_margin, use_container_width=True)
    else:
        st.warning("No margin data available for display.")


def debug_value_impact_calculation(project_id, cm2_as_sold, cm2_fct_n, cm2_total_erosion, 
                                  selling_price_fct_n, cm2_erosion_value, contract_value):
    """Debug function to trace value impact calculations"""
    print("\n" + "="*80)
    print(f"DEBUG: Value Impact Calculation for Project {project_id}")
    print("="*80)
    print(f"Input Values:")
    print(f"  - CM2 AS SOLD:        {cm2_as_sold:.2f}%")
    print(f"  - CM2 FCT(n):         {cm2_fct_n:.2f}%")
    print(f"  - CM2 Total Erosion:  {cm2_total_erosion:+.2f}pp (FCT(n) - AS SOLD)")
    print(f"  - Selling Price FCT(n): CHF {selling_price_fct_n:,.0f}")
    print(f"  - Contract Value:      CHF {contract_value:,.0f}")
    print(f"\nCalculation:")
    print(f"  Value Impact = (Total Erosion / 100) × Selling Price FCT(n)")
    print(f"  Value Impact = ({cm2_total_erosion:+.2f} / 100) × {selling_price_fct_n:,.0f}")
    print(f"  Value Impact = {cm2_total_erosion/100:+.4f} × {selling_price_fct_n:,.0f}")
    print(f"  Value Impact = CHF {cm2_erosion_value:,.0f}")
    print(f"\nResult:")
    print(f"  - Value Impact: CHF {cm2_erosion_value:,.0f} ({cm2_erosion_value/1000:.0f}K)")
    print(f"  - Direction: {'Improvement' if cm2_erosion_value > 0 else 'Erosion' if cm2_erosion_value < 0 else 'No Change'}")
    print("="*80 + "\n")

# FIX #3: Enhanced CM Variability Analysis
def calculate_margin_variability_metrics(project_data):
    """Calculate comprehensive margin variability metrics for a project"""
    
    # Extract margin data across periods from cost_analysis
    cost_analysis = project_data.get('cost_analysis', {})
    contract_value = safe_get_value(project_data, 'revenues', 'Contract Price', 'n_ptd')
    
    # CM2 across periods
    cm2_as_sold = cost_analysis.get('cm2_pct_as_sold', 0)
    cm2_fct_n1 = cost_analysis.get('cm2_pct_fct_n1', 0) 
    cm2_fct_n = cost_analysis.get('cm2_pct_fct_n', 0)
    
    # CM1 across periods  
    cm1_as_sold = cost_analysis.get('cm1_pct_as_sold', 0)
    cm1_fct_n1 = cost_analysis.get('cm1_pct_fct_n1', 0)
    cm1_fct_n = cost_analysis.get('cm1_pct_fct_n', 0)
    
    # Get selling prices for value calculations
    selling_price_as_sold = cost_analysis.get('selling_price_as_sold', contract_value)
    selling_price_fct_n = cost_analysis.get('selling_price_fct_n', contract_value)
    
    # Check if we have any historical data at all
    has_historical_data = False
    if (cm2_as_sold != 0 and cm2_as_sold != cm2_fct_n) or (cm2_fct_n1 != 0 and cm2_fct_n1 != cm2_fct_n):
        has_historical_data = True
    elif (cm1_as_sold != 0 and cm1_as_sold != cm1_fct_n) or (cm1_fct_n1 != 0 and cm1_fct_n1 != cm1_fct_n):
        has_historical_data = True
    
    # If no historical data, return empty metrics
    if not has_historical_data:
        return {
            'has_historical_data': False,
            'cm2_current': cm2_fct_n,
            'cm1_current': cm1_fct_n,
            'message': 'No historical margin data found for variability analysis'
        }
    
    # Calculate total erosion/improvement (in percentage points)
    cm2_total_erosion = cm2_fct_n - cm2_as_sold  # Positive = improvement, Negative = erosion
    cm1_total_erosion = cm1_fct_n - cm1_as_sold
    
    # SIMPLIFIED: Calculate value impact by multiplying total variance by contract value
    # Use the current contract value (selling price FCT n) as the base
    # Total erosion is already in percentage points, so divide by 100 to get the factor
    cm2_erosion_value = (cm2_total_erosion / 100) * selling_price_fct_n
    cm1_erosion_value = (cm1_total_erosion / 100) * selling_price_fct_n
    
    # DEBUG: Enable/disable debug output
    DEBUG_VALUE_IMPACT = True  # Set to False to disable debug output
    if DEBUG_VALUE_IMPACT:
        # Try to get project ID from project_data
        project_id = project_data.get('project_info', {}).get('Project No.', 'Unknown')
        debug_value_impact_calculation(
            project_id=project_id,
            cm2_as_sold=cm2_as_sold,
            cm2_fct_n=cm2_fct_n,
            cm2_total_erosion=cm2_total_erosion,
            selling_price_fct_n=selling_price_fct_n,
            cm2_erosion_value=cm2_erosion_value,
            contract_value=contract_value
        )
    
    return {
        # CM2 Variability Metrics
        'cm2_total_erosion': cm2_total_erosion,  # Total margin change from baseline (pp)
        'cm2_recent_change': cm2_fct_n - cm2_fct_n1,   # Recent margin change (pp)
        'cm2_forecast_accuracy': abs(cm2_fct_n1 - cm2_fct_n) if cm2_fct_n1 != 0 else 0,
        'cm2_volatility_index': calculate_volatility_index([cm2_as_sold, cm2_fct_n1, cm2_fct_n]),
        
        # CM1 Variability Metrics  
        'cm1_total_erosion': cm1_total_erosion,
        'cm1_recent_change': cm1_fct_n - cm1_fct_n1,
        'cm1_forecast_accuracy': abs(cm1_fct_n1 - cm1_fct_n) if cm1_fct_n1 != 0 else 0,
        'cm1_volatility_index': calculate_volatility_index([cm1_as_sold, cm1_fct_n1, cm1_fct_n]),
        
        # Margin Stability Assessment
        'margin_trend': assess_margin_trend(cm2_as_sold, cm2_fct_n1, cm2_fct_n),
        'margin_risk_level': assess_margin_risk(cm2_fct_n, cm2_total_erosion),
        'forecast_reliability': assess_forecast_reliability(cm2_fct_n1, cm2_fct_n, cm1_fct_n1, cm1_fct_n),
        
        # Value Impact (absolute CHF impact)
        'cm2_erosion_value': cm2_erosion_value,
        'cm1_erosion_value': cm1_erosion_value,
        
        # Current margin values for display
        'cm2_current': cm2_fct_n,
        'cm1_current': cm1_fct_n,
        
        # Additional data for analysis
        'has_historical_data': has_historical_data,
        'cm2_as_sold': cm2_as_sold,
        'cm2_fct_n1': cm2_fct_n1,
        'cm1_as_sold': cm1_as_sold,
        'cm1_fct_n1': cm1_fct_n1,
        
        # Debug info
        'selling_price_as_sold': selling_price_as_sold,
        'selling_price_fct_n': selling_price_fct_n
    }

def calculate_volatility_index(margin_values):
    """Calculate margin volatility index (standard deviation of margin changes)"""
    if len(margin_values) < 2:
        return 0
    
    # Remove zeros and calculate standard deviation of percentage points
    valid_values = [v for v in margin_values if v != 0]
    if len(valid_values) < 2:
        return 0
        
    try:
        return np.std(valid_values)
    except:
        return 0

def assess_margin_trend(cm2_as_sold, cm2_fct_n1, cm2_fct_n):
    """Assess overall margin trend direction"""
    if cm2_as_sold == 0:
        return "📊 Unknown"
    
    total_change = cm2_fct_n - cm2_as_sold
    recent_change = cm2_fct_n - cm2_fct_n1 if cm2_fct_n1 != 0 else 0
    
    if total_change > 2:
        return "📈 Improving"
    elif total_change < -5:
        return "📉 Severely Declining" 
    elif total_change < -2:
        return "📉 Declining"
    elif abs(recent_change) <= 1:
        return "📊 Stable"
    else:
        return "🌊 Volatile"

def assess_margin_risk(current_cm2, cm2_total_erosion):
    """Assess margin risk level based on current margin and erosion/improvement"""
    
    # Get current thresholds
    cm2_excellent = EXECUTIVE_THRESHOLDS['cm2_margin']['excellent']
    cm2_good = EXECUTIVE_THRESHOLDS['cm2_margin']['good']
    cm2_warning = EXECUTIVE_THRESHOLDS['cm2_margin']['warning']

    # For improving margins (positive variance)
    if cm2_total_erosion > 2:  # Improvement of more than 2pp
        if current_cm2 >= cm2_excellent:
            return "🟢 Low"
        elif current_cm2 >= cm2_good:  # Relaxed threshold for improving projects
            return "🟢 Low"  # Good margin with positive trend
        elif current_cm2 >= cm2_warning:
            return "🟡 Medium"  # Acceptable margin, improving
        else:
            return "🔴 High"  # Low margin but improving
    
    # For stable margins (-2 to +2 pp change)
    elif cm2_total_erosion >= -2:
        if current_cm2 >= cm2_excellent:
            return "🟢 Low"
        elif current_cm2 >= cm2_good:
            return "🟡 Medium"
        elif current_cm2 >= cm2_warning:
            return "🟠 High"
        else:
            return "🔴 Critical"
    
    # For deteriorating margins (negative erosion < -2pp)
    else:
        if current_cm2 >= cm2_excellent and cm2_total_erosion > -5:
            return "🟡 Medium"  # High margin but deteriorating
        elif current_cm2 >= cm2_good and cm2_total_erosion > -5:
            return "🟠 High"  # Good margin but deteriorating
        elif current_cm2 >= cm2_warning and cm2_total_erosion > -10:
            return "🟠 High"
        else:
            return "🔴 Critical"  # Low margin and/or severe deterioration

def assess_forecast_reliability(cm2_n1, cm2_n, cm1_n1, cm1_n):
    """Assess how reliable margin forecasts are based on recent changes"""
    cm2_change = abs(cm2_n - cm2_n1) if cm2_n1 != 0 else 0
    cm1_change = abs(cm1_n - cm1_n1) if cm1_n1 != 0 else 0
    
    avg_change = (cm2_change + cm1_change) / 2
    
    if avg_change <= 1:
        return "🎯 Highly Reliable"
    elif avg_change <= 3:
        return "✅ Reliable" 
    elif avg_change <= 5:
        return "⚠️ Moderately Reliable"
    else:
        return "❌ Unreliable"

def render_margin_variability_analysis(portfolio_data):
    """Render comprehensive margin variability analysis dashboard"""
    st.markdown("## 📊 Contribution Margin Variability Analysis")
    
    # Add debug toggle
    debug_mode = st.checkbox("🔍 Show Value Impact Debug Information", value=False, key="margin_debug")
    
    # Process all projects for margin variability
    project_margin_metrics = {}
    portfolio_metrics = {
        'total_projects': 0,
        'projects_with_data': 0,
        'severely_declining': 0,
        'high_volatility': 0,
        'unreliable_forecasts': 0,
        'total_cm2_erosion_value': 0,
        'avg_cm2_volatility': 0,
        'margin_risk_projects': 0
    }
    
    for project_id, project in portfolio_data.items():
        try:
            metrics = calculate_margin_variability_metrics(project['data'])
            
            # Only include projects with historical data for variability analysis
            if metrics['has_historical_data']:
                project_margin_metrics[project_id] = {
                    'name': project['name'],
                    'metrics': metrics
                }
                
                # Debug display for this project if enabled
                if debug_mode:
                    with st.expander(f"Debug: {project_id} - {project['name']}", expanded=False):
                        st.markdown("**Value Impact Calculation Debug:**")
                        st.code(f"""
Project: {project_id}
CM2 AS SOLD: {metrics['cm2_as_sold']:.2f}%
CM2 FCT(n): {metrics['cm2_current']:.2f}%
CM2 Total Erosion: {metrics['cm2_total_erosion']:+.2f}pp

Selling Price FCT(n): CHF {metrics['selling_price_fct_n']:,.0f}

Calculation:
Value Impact = (Total Erosion / 100) × Selling Price FCT(n)
Value Impact = ({metrics['cm2_total_erosion']:+.2f} / 100) × {metrics['selling_price_fct_n']:,.0f}
Value Impact = CHF {metrics['cm2_erosion_value']:,.0f} ({metrics['cm2_erosion_value']/1000:.0f}K)

Direction: {'Improvement' if metrics['cm2_erosion_value'] > 0 else 'Erosion' if metrics['cm2_erosion_value'] < 0 else 'No Change'}
                        """)
                
                # Aggregate portfolio metrics
                portfolio_metrics['projects_with_data'] += 1
                portfolio_metrics['total_cm2_erosion_value'] += metrics['cm2_erosion_value']
                portfolio_metrics['avg_cm2_volatility'] += metrics['cm2_volatility_index']
                
                if "Severely Declining" in metrics['margin_trend']:
                    portfolio_metrics['severely_declining'] += 1
                if metrics['cm2_volatility_index'] > 5:
                    portfolio_metrics['high_volatility'] += 1
                if "Unreliable" in metrics['forecast_reliability']:
                    portfolio_metrics['unreliable_forecasts'] += 1
                if "Critical" in metrics['margin_risk_level'] or "High" in metrics['margin_risk_level']:
                    portfolio_metrics['margin_risk_projects'] += 1
            
            portfolio_metrics['total_projects'] += 1
            
        except Exception as e:
            st.warning(f"Could not process margin variability for project {project_id}: {str(e)}")
            continue
    
    if portfolio_metrics['projects_with_data'] == 0:
        st.warning("📊 No historical margin data available for variability analysis.")
        st.info("**Required data for each project:**")
        st.write("• CM1/CM2 percentages at 'As Sold' (original baseline)")
        st.write("• CM1/CM2 percentages at 'FCT(n-1)' (previous forecast)")
        st.write("• CM1/CM2 percentages at 'FCT(n)' (current forecast)")
        st.write("")
        st.write("Please ensure your Excel template includes these values in the '3_Cost_Breakdown' sheet.")
        return
    
    # Calculate averages
    portfolio_metrics['avg_cm2_volatility'] /= portfolio_metrics['projects_with_data']
    
    # Portfolio-level KPIs
    st.markdown("### 🎯 Portfolio Margin Variability KPIs")
    
    col1, col2, col3, col4, col5 = st.columns(5)
    
    with col1:
        erosion_pct = (portfolio_metrics['severely_declining'] / portfolio_metrics['projects_with_data'] * 100)
        erosion_icon = "🔴" if erosion_pct > 20 else "🟡" if erosion_pct > 10 else "🟢"
        st.metric("Margin Erosion Risk", f"{erosion_pct:.0f}%", f"{erosion_icon} {portfolio_metrics['severely_declining']} projects")
    
    with col2:
        volatility_icon = "🔴" if portfolio_metrics['avg_cm2_volatility'] > 5 else "🟡" if portfolio_metrics['avg_cm2_volatility'] > 3 else "🟢"
        st.metric("Avg Margin Volatility", f"{portfolio_metrics['avg_cm2_volatility']:.1f}pp", f"{volatility_icon}")
    
    with col3:
        forecast_reliability_pct = (1 - portfolio_metrics['unreliable_forecasts'] / portfolio_metrics['projects_with_data']) * 100
        reliability_icon = "🟢" if forecast_reliability_pct > 80 else "🟡" if forecast_reliability_pct > 60 else "🔴"
        st.metric("Forecast Reliability", f"{forecast_reliability_pct:.0f}%", f"{reliability_icon}")
    
    with col4:
        total_impact = portfolio_metrics['total_cm2_erosion_value']
        if total_impact > 0:
            impact_icon = "🟢"
            impact_text = f"Improvement"
        else:
            impact_icon = "🔴"
            impact_text = f"Erosion"
        st.metric("Total Margin Impact", format_currency_millions(total_impact), f"{impact_icon} {impact_text}")

    with col5:
        risk_pct = (portfolio_metrics['margin_risk_projects'] / portfolio_metrics['projects_with_data'] * 100)
        risk_icon = "🔴" if risk_pct > 30 else "🟡" if risk_pct > 15 else "🟢"
        st.metric("High Risk Projects", f"{risk_pct:.0f}%", f"{risk_icon}")
    
    # Add note about data availability
    if portfolio_metrics['projects_with_data'] < portfolio_metrics['total_projects']:
        st.info(f"ℹ️ Variability analysis includes {portfolio_metrics['projects_with_data']} of {portfolio_metrics['total_projects']} projects (only those with historical data)")
    
    # Project-level margin variability table
    st.markdown("### 📋 Project Margin Variability Performance")
    
    margin_summary = []
    for project_id, data in project_margin_metrics.items():
        metrics = data['metrics']
        
        # Create detailed period breakdown
        period_breakdown = f"AS: {metrics['cm2_as_sold']:.1f}% → N-1: {metrics['cm2_fct_n1']:.1f}% → N: {metrics['cm2_current']:.1f}%"
        
        margin_summary.append({
            'Project': project_id,
            'Name': data['name'][:25] + "..." if len(data['name']) > 25 else data['name'],
            'CM2 Evolution': period_breakdown,
            'Total Erosion': f"{metrics['cm2_total_erosion']:+.1f}pp",
            'Recent Change': f"{metrics['cm2_recent_change']:+.1f}pp", 
            'Volatility': f"{metrics['cm2_volatility_index']:.1f}pp",
            'Trend': metrics['margin_trend'],
            'Risk Level': metrics['margin_risk_level'],
            'Forecast Reliability': metrics['forecast_reliability'],
            'Value Impact': format_currency_thousands(metrics['cm2_erosion_value'])
        })
    
    # Sort by absolute value impact
    try:
        margin_summary.sort(key=lambda x: abs(float(x['Value Impact'].replace('CHF ', '').replace('K', '').replace(',', '').replace('M', '000'))), reverse=True)
    except:
        pass  # Keep original order if sorting fails
    
    df_margin_var = pd.DataFrame(margin_summary)
    st.dataframe(df_margin_var, use_container_width=True)
    
    # Margin variability insights and recommendations
    render_margin_variability_insights(portfolio_metrics, project_margin_metrics)

def render_margin_variability_insights(portfolio_metrics, project_margin_metrics):
    """Render margin variability insights and strategic recommendations"""
    st.markdown("### 💡 Margin Variability Insights & Strategic Actions")
    
    col1, col2 = st.columns(2)
    
    with col1:
        # Portfolio margin health assessment
        total_projects = portfolio_metrics['projects_with_data']
        severely_declining = portfolio_metrics['severely_declining'] 
        high_volatility = portfolio_metrics['high_volatility']
        unreliable_forecasts = portfolio_metrics['unreliable_forecasts']
        
        # Calculate portfolio margin health score
        health_score = max(0, 100 - 
                          (severely_declining/total_projects * 40) - 
                          (high_volatility/total_projects * 30) - 
                          (unreliable_forecasts/total_projects * 30)) if total_projects > 0 else 0
        
        health_status = "Excellent" if health_score >= 80 else "Good" if health_score >= 60 else "Concerning" if health_score >= 40 else "Critical"
        
        
        total_impact = portfolio_metrics['total_cm2_erosion_value']
        if total_impact < 0:
            impact_text = f"{format_currency_millions(abs(total_impact))} margin erosion"
            impact_color = "red"
        else:
            impact_text = f"{format_currency_millions(total_impact)} margin improvement"
            impact_color = "green"

        st.markdown(f"""
        <div class="exec-summary">
            <h4>📊 Portfolio Margin Health Assessment</h4>
            <ul>
                <li><strong>Overall Health Score:</strong> {health_score:.0f}/100 - <span style="color: {'green' if health_score >= 80 else 'orange' if health_score >= 60 else 'red'}">{health_status}</span></li>
                <li><strong>Projects at Risk:</strong> {severely_declining} with severe margin decline</li>
                <li><strong>Volatile Margins:</strong> {high_volatility} projects with high volatility</li>
                <li><strong>Forecast Issues:</strong> {unreliable_forecasts} projects with unreliable forecasts</li>
                <li><strong>Portfolio Impact:</strong> <span style="color: {impact_color}">{impact_text}</span></li>
            </ul>
        </div>
        """, unsafe_allow_html=True)

    
    with col2:
        # Strategic recommendations based on margin patterns
        recommendations = []
        
        if total_impact > 0:
            recommendations.extend([
                "✅ **Portfolio Improvement:** Net margin gain of " + format_currency_millions(total_impact),
                "📈 **Momentum Building:** Continue current margin management practices"
        ])
        # Remove the severe decline recommendation if overall is positive
        if severely_declining > 0:
            recommendations[0] = f"⚠️ **Mixed Performance:** Despite {severely_declining} declining project(s), portfolio shows net improvement"


        # Risk-based recommendations
        if severely_declining > 0:
            recommendations.extend([
                f"🚨 **Immediate Action:** {severely_declining} projects with severe margin decline need intervention",
                "🔍 **Root Cause Analysis:** Investigate cost drivers causing margin erosion"
            ])
        
        if high_volatility > total_projects * 0.3:
            recommendations.extend([
                "📊 **Enhanced Monitoring:** High margin volatility across portfolio",
                "🎯 **Standardization:** Implement consistent cost estimation practices"
            ])
        
        if unreliable_forecasts > total_projects * 0.2:
            recommendations.extend([
                "📈 **Forecast Improvement:** Enhance margin forecasting accuracy",
                "🔄 **Regular Updates:** Increase frequency of margin reviews"
            ])
        
        # Positive recommendations
        if health_score >= 80:
            recommendations.extend([
                "✅ **Maintain Excellence:** Strong margin management practices",
                "📚 **Best Practice Sharing:** Document and replicate successful approaches"
            ])
        
        # General recommendations
        recommendations.extend([
            "🎯 **Focus Areas:** Projects with >5pp margin erosion need priority attention",
            "📊 **Dashboard Integration:** Monitor margin variability as key portfolio KPI",
            "💼 **Executive Reporting:** Include margin trends in senior management reviews"
        ])
        
        st.markdown(f"""
        <div class="exec-summary">
            <h4>🎯 Strategic Margin Management Actions</h4>
            <ul>
                {''.join([f'<li>{rec}</li>' for rec in recommendations])}
            </ul>
        </div>
        """, unsafe_allow_html=True)

def create_comprehensive_margin_chart(margin_projects):
    """Create comprehensive margin analysis chart with multiple perspectives including IL/EC ratio"""
    fig = make_subplots(
        rows=2, cols=2,
        subplot_titles=('CM1 vs CM2 Performance Matrix', 'Cost Structure with IL/EC Ratio Analysis', 
                       'Cost Variance vs Committed Ratio', 'IL/EC Ratio Distribution & Benchmarks'),
        specs=[[{"type": "scatter"}, {"secondary_y": True}], 
               [{"type": "scatter"}, {"type": "scatter"}]],
        vertical_spacing=0.15,
        horizontal_spacing=0.12
    )
    
    # Get current CM2 thresholds
    cm2_excellent = EXECUTIVE_THRESHOLDS['cm2_margin']['excellent']
    cm2_good = EXECUTIVE_THRESHOLDS['cm2_margin']['good']
    cm2_warning = EXECUTIVE_THRESHOLDS['cm2_margin']['warning']
    
    # Extract data for charts
    project_names = [p['project_id'] for p in margin_projects]
    cm1_values = [p['cm1_pct'] for p in margin_projects]
    cm2_values = [p['cm2_pct'] for p in margin_projects]
    ec_values = [p['ec_total']/1000 for p in margin_projects]
    ic_values = [p['ic_total']/1000 for p in margin_projects]
    contract_values = [p['contract_value']/1000 for p in margin_projects]
    committed_ratios = [p['committed_ratio'] for p in margin_projects]
    cost_variances = [p['cost_variance_pct'] for p in margin_projects]
    
    # Calculate IL/EC ratios
    il_ec_ratios = []
    for i in range(len(ec_values)):
        if ec_values[i] > 0:
            ratio = ic_values[i] / ec_values[i]
            il_ec_ratios.append(ratio)
        else:
            il_ec_ratios.append(0)
    
    # Industry benchmarks for IL/EC ratio
    INDUSTRY_BENCHMARKS = {
        'engineering': {'min': 0.15, 'target': 0.25, 'max': 0.35},
        'construction': {'min': 0.10, 'target': 0.20, 'max': 0.30},
        'software': {'min': 0.30, 'target': 0.50, 'max': 0.70},
        'consulting': {'min': 0.60, 'target': 0.80, 'max': 1.00},
        'default': {'min': 0.20, 'target': 0.30, 'max': 0.40}
    }
    
    # Use default benchmark
    benchmark = INDUSTRY_BENCHMARKS['construction']
    
    # 1. CM1 vs CM2 Performance Matrix
    colors = []
    for cm2 in cm2_values:
        if cm2 >= cm2_excellent:
            colors.append('darkgreen')
        elif cm2 >= cm2_good:
            colors.append('green')
        elif cm2 >= cm2_warning:
            colors.append('orange')
        else:
            colors.append('red')
    
    fig.add_trace(go.Scatter(
        x=cm1_values,
        y=cm2_values,
        mode='markers+text',
        marker=dict(size=[max(10, cv/50) for cv in contract_values], color=colors, opacity=0.7),
        text=project_names,
        textposition="top center",
        name='Projects',
        showlegend=False
    ), row=1, col=1)
    
    # Add threshold lines for CM2
    fig.add_hline(y=cm2_excellent, line_dash="solid", line_color="darkgreen", row=1, col=1)
    fig.add_hline(y=cm2_good, line_dash="dash", line_color="green", row=1, col=1)
    fig.add_hline(y=cm2_warning, line_dash="dot", line_color="orange", row=1, col=1)
    fig.add_vline(x=20, line_dash="dash", line_color="blue", row=1, col=1)
    
    # 2. Enhanced Cost Structure with IL/EC Ratio
    fig.add_trace(go.Bar(
        name='External Costs',
        x=project_names,
        y=ec_values,
        marker_color='#FF6B6B',
        opacity=0.8,
        showlegend=True,
        yaxis='y'
    ), row=1, col=2)
    
    fig.add_trace(go.Bar(
        name='Internal Costs',
        x=project_names,
        y=ic_values,
        marker_color='#4ECDC4',
        opacity=0.8,
        showlegend=True,
        yaxis='y'
    ), row=1, col=2)
    
    # Add IL/EC ratio line with colors
    ratio_colors = []
    for ratio in il_ec_ratios:
        if benchmark['min'] <= ratio <= benchmark['max']:
            ratio_colors.append('green')
        elif ratio < benchmark['min']:
            ratio_colors.append('orange')
        else:
            ratio_colors.append('red')
    
    fig.add_trace(go.Scatter(
        name='IL/EC Ratio',
        x=project_names,
        y=il_ec_ratios,
        mode='lines+markers+text',
        line=dict(color='darkblue', width=3),
        marker=dict(size=10, color=ratio_colors),
        text=[f"{r:.2f}" for r in il_ec_ratios],
        textposition="top center",
        yaxis='y2',
        showlegend=True
    ), row=1, col=2, secondary_y=True)
    
    # Add benchmark lines
    fig.add_hline(y=benchmark['target'], line_dash="dash", line_color="green", 
                  annotation_text=f"Target ({benchmark['target']:.2f})",
                  annotation_position="right", secondary_y=True, row=1, col=2)
    fig.add_hline(y=benchmark['min'], line_dash="dot", line_color="orange", 
                  annotation_text=f"Min ({benchmark['min']:.2f})",
                  annotation_position="right", secondary_y=True, row=1, col=2)
    fig.add_hline(y=benchmark['max'], line_dash="dot", line_color="orange", 
                  annotation_text=f"Max ({benchmark['max']:.2f})",
                  annotation_position="right", secondary_y=True, row=1, col=2)
    
    # 3. Cost Variance vs Committed Ratio Risk Matrix
    risk_colors = ['green' if (cv <= 10 and cr <= 1.1) else 'orange' if (cv <= 20 and cr <= 1.2) else 'red' 
                   for cv, cr in zip(cost_variances, committed_ratios)]
    
    fig.add_trace(go.Scatter(
        x=cost_variances,
        y=committed_ratios,
        mode='markers+text',
        marker=dict(size=[max(8, cv/50) for cv in contract_values], color=risk_colors, opacity=0.7),
        text=project_names,
        textposition="top center",
        name='Risk Matrix',
        showlegend=False
    ), row=2, col=1)
    
    fig.add_hline(y=1.1, line_dash="dash", line_color="orange", row=2, col=1)
    fig.add_vline(x=10, line_dash="dash", line_color="orange", row=2, col=1)
    
    # 4. IMPROVED IL/EC Ratio Distribution Analysis
    # Calculate statistics first
    within_range_count = sum(1 for r in il_ec_ratios if benchmark['min'] <= r <= benchmark['max'])
    avg_ratio = np.mean(il_ec_ratios) if il_ec_ratios else 0
    
    # Sort data for better visualization
    sorted_data = sorted(zip(project_names, il_ec_ratios), key=lambda x: x[1])
    sorted_projects = [d[0] for d in sorted_data]
    sorted_ratios = [d[1] for d in sorted_data]
    
    # Determine colors for sorted data
    sorted_colors = []
    for ratio in sorted_ratios:
        if benchmark['min'] <= ratio <= benchmark['max']:
            sorted_colors.append('green')
        elif ratio < benchmark['min']:
            sorted_colors.append('orange')
        else:
            sorted_colors.append('red')
    
    # Create horizontal lollipop chart
    fig.add_trace(go.Scatter(
        x=sorted_ratios,
        y=sorted_projects,
        mode='markers',
        marker=dict(
            size=12,
            color=sorted_colors,
            symbol='circle',
            line=dict(width=2, color='white')
        ),
        name='IL/EC Ratios',
        showlegend=False,
        hovertemplate='<b>%{y}</b><br>IL/EC Ratio: %{x:.2f}<extra></extra>'
    ), row=2, col=2)
    
    # Add lines from y-axis to points
    for i, (proj, ratio) in enumerate(zip(sorted_projects, sorted_ratios)):
        fig.add_trace(go.Scatter(
            x=[0, ratio],
            y=[proj, proj],
            mode='lines',
            line=dict(color='lightgray', width=1),
            showlegend=False,
            hoverinfo='skip'
        ), row=2, col=2)
    
    # Add the optimal range as a shaded vertical band
    max_x = max(max(sorted_ratios) * 1.1, benchmark['max'] * 1.2) if sorted_ratios else 1.0
    
    # Add background shading for different zones
    # Below minimum (orange zone)
    fig.add_shape(
        type="rect",
        x0=0, x1=benchmark['min'],
        y0=-0.5, y1=len(sorted_projects)-0.5,
        fillcolor="rgba(255, 200, 200, 0.2)",
        line=dict(width=0),
        layer="below",
        row=2, col=2
    )
    
    # Optimal range (green zone) - WITH DOTTED BORDER
    fig.add_shape(
        type="rect",
        x0=benchmark['min'], x1=benchmark['max'],
        y0=-0.5, y1=len(sorted_projects)-0.5,
        fillcolor="rgba(200, 255, 200, 0.3)",
        line=dict(color="darkgreen", width=2, dash="dot"),  # Changed to dotted line
        layer="below",
        row=2, col=2
    )
    
    # Above maximum (orange zone)
    fig.add_shape(
        type="rect",
        x0=benchmark['max'], x1=max_x,
        y0=-0.5, y1=len(sorted_projects)-0.5,
        fillcolor="rgba(255, 200, 200, 0.2)",
        line=dict(width=0),
        layer="below",
        row=2, col=2
    )
    
    # Add vertical reference lines
    # TARGET LINE - THIN RED LINE
    fig.add_vline(x=benchmark['target'], line_dash="solid", line_color="red", 
                  line_width=1, row=2, col=2)  # Changed to red and thin (width=1)
    
    # Min and Max lines
    fig.add_vline(x=benchmark['min'], line_dash="dash", line_color="darkorange", 
                  line_width=2, row=2, col=2)
    fig.add_vline(x=benchmark['max'], line_dash="dash", line_color="darkorange", 
                  line_width=2, row=2, col=2)
    
    # Add annotations
    fig.add_annotation(
        x=(benchmark['min'] + benchmark['max']) / 2,
        y=len(sorted_projects),
        text="<b>OPTIMAL RANGE</b>",
        showarrow=False,
        font=dict(size=12, color="darkgreen"),
        bgcolor="rgba(255, 255, 255, 0.8)",
        bordercolor="darkgreen",
        borderwidth=2,
        xref="x4",
        yref="y4"
    )
    
    # Add summary box
    fig.add_annotation(
        x=0.98,
        y=0.02,
        xref='x4 domain',
        yref='y4 domain',
        text=f"<b>Summary</b><br>" +
             f"Within Range: {within_range_count}/{len(il_ec_ratios)}<br>" +
             f"Average: {avg_ratio:.2f}<br>" +
             f"Target: {benchmark['target']:.2f}",
        showarrow=False,
        align='right',
        font=dict(size=10),
        bgcolor="rgba(255, 255, 255, 0.9)",
        bordercolor="gray",
        borderwidth=1,
        xanchor='right',
        yanchor='bottom'
    )
    
    # Update layout
    fig.update_layout(
        height=900, 
        showlegend=True, 
        title_text="Comprehensive Portfolio Margin & Cost Efficiency Analysis",
        barmode='stack'
    )
    
    # Update axes
    fig.update_xaxes(title_text="CM1 %", row=1, col=1)
    fig.update_yaxes(title_text="CM2 %", row=1, col=1)
    fig.update_xaxes(title_text="Projects", tickangle=-45, row=1, col=2)
    fig.update_yaxes(title_text="Costs (CHF M)", row=1, col=2)
    fig.update_yaxes(title_text="IL/EC Ratio", secondary_y=True, row=1, col=2)
    fig.update_xaxes(title_text="Cost Variance %", row=2, col=1)
    
    # CHANGE 1: Set y-axis range for Cost Variance vs Committed Ratio
    fig.update_yaxes(title_text="Committed Ratio", range=[0, 1.5], row=2, col=1)
    
    fig.update_xaxes(title_text="IL/EC Ratio", range=[0, max_x], row=2, col=2)
    fig.update_yaxes(title_text="Projects", row=2, col=2)
    
    return fig

def render_work_package_analysis(portfolio_data):
    """Render work package cost variance analysis with materiality and commodity filters"""
    st.markdown("## 📦 Work Package Cost Variance Analysis")
    
    # Aggregate work package data across portfolio
    all_work_packages = []
    risk_contingencies = []
    total_wp_value = 0
    high_variance_count = 0
    critical_variance_count = 0
    
    # Process each project
    for project_id, project in portfolio_data.items():
        work_packages = project['data'].get('work_packages', {})
        
        # Get project total as sold value for materiality calculation
        project_total_as_sold = safe_get_value(project['data'], 'cost_analysis', 'total_as_sold', default=0)
        if project_total_as_sold == 0:
            # Fallback to contract value if total_as_sold not available
            project_total_as_sold = safe_get_value(project['data'], 'revenues', 'Contract Price', 'n_ptd')
        
        for wp_code, wp_data in work_packages.items():
            variance_pct = wp_data.get('variance_pct', 0)
            as_sold = wp_data.get('as_sold', 0)
            description = wp_data.get('description', '')
            
            if as_sold > 0:  # Only include work packages with value
                
                # Calculate materiality percentage
                materiality_pct = (as_sold / project_total_as_sold * 100) if project_total_as_sold > 0 else 0
                
                wp_entry = {
                    'project_id': project_id,
                    'project_name': project['name'],
                    'wp_code': wp_code,
                    'description': description[:50] + "..." if len(description) > 50 else description,
                    'as_sold': as_sold,
                    'fct_n': wp_data.get('fct_n', 0),
                    'variance_pct': variance_pct,
                    'commitment_ratio': wp_data.get('commitment_ratio', 0),
                    'materiality_pct': materiality_pct,
                    'project_total_as_sold': project_total_as_sold
                }
                
                # Separate Risk Contingencies from commodity work packages
                if 'risk' in description.lower() and 'contingenc' in description.lower():
                    risk_contingencies.append(wp_entry)
                else:
                    # Only include commodity work packages that are >2.5% of project value
                    if materiality_pct > 2.5:
                        all_work_packages.append(wp_entry)
                        total_wp_value += as_sold
                        
                        # FIX #1 & #2: Only count negative variances as issues
                        if variance_pct > 25:  # Positive variance = cost increase = bad
                            critical_variance_count += 1
                        elif variance_pct > 15:  # Positive variance = cost increase = warning
                            high_variance_count += 1
    
    if not all_work_packages and not risk_contingencies:
        st.warning("📦 No work package data available for analysis.")
        return
    
    # Portfolio work package summary
    st.markdown("### 📊 Material Work Package Summary (>2.5% of Project Value)")
    
    col1, col2, col3, col4, col5 = st.columns(5)
    
    with col1:
        st.metric("Material Work Packages", len(all_work_packages))
    
    with col2:
        if all_work_packages:
            # Calculate average of positive variances only (cost increases)
            positive_variances = [wp['variance_pct'] for wp in all_work_packages if wp['variance_pct'] > 0]
            if positive_variances:
                avg_variance = np.mean(positive_variances)
                variance_icon = "🟢" if avg_variance <= 10 else "🟡" if avg_variance <= 20 else "🔴"
                st.metric("Avg Cost Increase", f"{avg_variance:+.1f}%", f"{variance_icon}")
            else:
                st.metric("Avg Cost Increase", "0.0%", "🟢")
        else:
            st.metric("Avg Cost Increase", "N/A")
    
    with col3:
        st.metric("High Variance (>15%)", high_variance_count, 
                 "🟠" if high_variance_count > 0 else "🟢")
    
    with col4:
        st.metric("Critical Variance (>25%)", critical_variance_count,
                 "🔴" if critical_variance_count > 0 else "🟢")
    
    with col5:
        total_materiality = sum([wp['as_sold'] for wp in all_work_packages])
        total_portfolio_value = sum([wp['project_total_as_sold'] for wp in all_work_packages])
        portfolio_coverage = (total_materiality / total_portfolio_value * 100) if total_portfolio_value > 0 else 0
        st.metric("Portfolio Coverage", f"{portfolio_coverage:.1f}%")
    
    # FIX #2: Updated variance distribution to focus on cost increases
    if all_work_packages:
        st.markdown("### 📊 Material Work Package Variance Distribution")
        
        variance_ranges = {
            'Cost Reduction (< -5%)': len([wp for wp in all_work_packages if wp['variance_pct'] < -5]),
            'Stable (±5%)': len([wp for wp in all_work_packages if -5 <= wp['variance_pct'] <= 5]),
            'Moderate Increase (5-15%)': len([wp for wp in all_work_packages if 5 < wp['variance_pct'] <= 15]),
            'High Increase (15-25%)': len([wp for wp in all_work_packages if 15 < wp['variance_pct'] <= 25]),
            'Critical Increase (>25%)': len([wp for wp in all_work_packages if wp['variance_pct'] > 25])
        }
        
        fig = go.Figure(data=[
            go.Pie(
                labels=list(variance_ranges.keys()),
                values=list(variance_ranges.values()),
                marker_colors=['darkgreen', 'green', 'yellow', 'orange', 'red'],
                hole=0.4
            )
        ])
        
        fig.update_layout(
            title='Material Work Package Cost Variance Distribution',
            height=400
        )
        
        st.plotly_chart(fig, use_container_width=True)
    
    # FIX #1: Focus on cost increases for high variance work packages
    st.markdown("### ⚠️ High Cost Increase Material Work Packages (>15% increase & >2.5% of Project)")
    
    high_cost_increase_wps = [wp for wp in all_work_packages if wp['variance_pct'] > 15]
    
    if high_cost_increase_wps:
        # Sort by variance (highest cost increases first)
        high_cost_increase_wps.sort(key=lambda x: x['variance_pct'], reverse=True)
        
        variance_summary = []
        for wp in high_cost_increase_wps:
            severity = "🔴 Critical" if wp['variance_pct'] > 25 else "🟠 High"
            
            variance_summary.append({
                'Project': wp['project_id'],
                'Work Package': wp['wp_code'],
                'Description': wp['description'],
                'As Sold': format_currency_thousands(wp['as_sold']),
                'FCT (n)': format_currency_thousands(wp['fct_n']),
                'Cost Increase': f"+{wp['variance_pct']:.1f}%",
                'Materiality': f"{wp['materiality_pct']:.1f}%",
                'Severity': severity,
                'Committed Ratio': f"{wp['commitment_ratio']:.2f}"
            })
        
        df_variance = pd.DataFrame(variance_summary)
        st.dataframe(df_variance, use_container_width=True)
    else:
        st.success("✅ No material work packages with significant cost increases.")
    
    # Show cost reductions as opportunities
    cost_reduction_wps = [wp for wp in all_work_packages if wp['variance_pct'] < -10]
    if cost_reduction_wps:
        st.markdown("### 💚 Cost Reduction Opportunities (>10% reduction & >2.5% of Project)")
        
        reduction_summary = []
        for wp in cost_reduction_wps:
            reduction_summary.append({
                'Project': wp['project_id'],
                'Work Package': wp['wp_code'],
                'Description': wp['description'],
                'As Sold': format_currency_thousands(wp['as_sold']),
                'FCT (n)': format_currency_thousands(wp['fct_n']),
                'Cost Reduction': f"{wp['variance_pct']:.1f}%",
                'Savings': format_currency_thousands(wp['as_sold'] - wp['fct_n']),
                'Materiality': f"{wp['materiality_pct']:.1f}%"
            })
        
        df_reduction = pd.DataFrame(reduction_summary)
        st.dataframe(df_reduction, use_container_width=True)
    
    # FIX #5: Remove Risk Contingencies section (simplified display)
    # Risk contingencies are still tracked but not displayed in detail
    
    # Enhanced recommendations section
    st.markdown("### 💡 Strategic Work Package Recommendations")
    
    col1, col2 = st.columns(2)
    
    with col1:
        # Material work package recommendations
        if high_cost_increase_wps:
            critical_wps = len([wp for wp in high_cost_increase_wps if wp['variance_pct'] > 25])
            total_critical_value = sum([wp['as_sold'] for wp in high_cost_increase_wps if wp['variance_pct'] > 25])
            
            st.markdown(f"""
            <div class="exec-summary">
                <h4>🎯 Material Work Package Actions</h4>
                <ul>
                    <li><strong>Critical WPs:</strong> {critical_wps} with >25% cost increase</li>
                    <li><strong>Value at Risk:</strong> {format_currency_millions(total_critical_value)}</li>
                    <li><strong>Focus Projects:</strong> {len(set([wp['project_id'] for wp in high_cost_increase_wps]))} projects need attention</li>
                    <li><strong>Materiality Threshold:</strong> Only analyzing WPs >2.5% of project value</li>
                </ul>
            </div>
            """, unsafe_allow_html=True)
        else:
            st.markdown(f"""
            <div class="exec-summary">
                <h4>✅ Material Work Package Status</h4>
                <ul>
                    <li><strong>Performance:</strong> All material WPs within acceptable variance</li>
                    <li><strong>Monitored WPs:</strong> {len(all_work_packages)} work packages >2.5% materiality</li>
                    <li><strong>Portfolio Coverage:</strong> {portfolio_coverage:.1f}% of total portfolio value</li>
                </ul>
            </div>
            """, unsafe_allow_html=True)
    
    with col2:
        # Strategic recommendations
        recommendations = [
            "🔍 **Focus on Materiality:** Analysis limited to WPs >2.5% of project value",
            "📊 **Cost Increase Focus:** Prioritize work packages with cost increases",
            "💚 **Capture Savings:** Realize cost reduction opportunities",
            "🎯 **Root Cause Analysis:** Investigate drivers for high cost increases",
            "📞 **Project Management:** Enhance monitoring for critical variance WPs"
        ]
        
        if cost_reduction_wps:
            recommendations.append(f"💰 **Opportunity:** {len(cost_reduction_wps)} WPs with cost savings potential")
            
        if len(all_work_packages) < 5:
            recommendations.append("⚠️ **Coverage Alert:** Few material WPs - consider lowering materiality threshold")
        
        st.markdown(f"""
        <div class="exec-summary">
            <h4>📋 Strategic Recommendations</h4>
            <ul>
                {''.join([f'<li>{rec}</li>' for rec in recommendations])}
            </ul>
        </div>
        """, unsafe_allow_html=True)

def render_quarterly_cash_flow_analysis(portfolio_data):
    """Render comprehensive quarterly cash flow analysis"""
    st.markdown("## 📈 Quarterly Cash Flow Analysis")
    
    # Aggregate quarterly data across all projects
    portfolio_quarters = {}
    projects_with_quarterly_data = 0
    
    for project_id, project in portfolio_data.items():
        quarterly_data = project['data'].get('cash_flow_quarterly', {})
        if quarterly_data:
            projects_with_quarterly_data += 1
            for quarter, data in quarterly_data.items():
                if quarter not in portfolio_quarters:
                    portfolio_quarters[quarter] = {
                        'as_sold': 0, 'fct_n1': 0, 'fct_n': 0, 'project_count': 0
                    }
                portfolio_quarters[quarter]['as_sold'] += data['as_sold']
                portfolio_quarters[quarter]['fct_n1'] += data['fct_n1']
                portfolio_quarters[quarter]['fct_n'] += data['fct_n']
                portfolio_quarters[quarter]['project_count'] += 1
    
    if projects_with_quarterly_data == 0:
        st.warning("📊 No quarterly cash flow data available.")
        return
    
    # Enhanced quarterly analysis with traffic lights
    col1, col2, col3, col4 = st.columns(4)
    
    with col1:
        total_as_sold = sum([q['as_sold'] for q in portfolio_quarters.values()])
        st.metric("Total As Sold Plan", format_currency_millions(total_as_sold))
    
    with col2:
        total_fct_n = sum([q['fct_n'] for q in portfolio_quarters.values()])
        st.metric("Total FCT (n)", format_currency_millions(total_fct_n))
    
    with col3:
        overall_variance = calculate_period_variance(total_fct_n, total_as_sold)
        variance_icon, _, _ = get_traffic_light_status(overall_variance, EXECUTIVE_THRESHOLDS['revenue_growth'])
        st.metric("Overall Variance", f"{overall_variance:+.1f}%", f"{variance_icon}")
    
    with col4:
        cash_flow_trend = calculate_cash_flow_trend(portfolio_quarters)
        trend_icon = "📈" if cash_flow_trend > 5 else "📊" if cash_flow_trend > -5 else "📉"
        st.metric("Trend", f"{cash_flow_trend:+.1f}%", f"{trend_icon}")
    
    # Create enhanced quarterly cash flow chart
    if portfolio_quarters:
        fig = create_enhanced_quarterly_cash_flow_chart(portfolio_quarters)
        st.plotly_chart(fig, use_container_width=True)
    
    # Quarterly performance table with insights
    st.markdown("### 📋 Quarterly Performance Analysis")
    
    quarterly_summary = []
    for quarter in sorted(portfolio_quarters.keys()):
        data = portfolio_quarters[quarter]
        variance_vs_sold = calculate_period_variance(data['fct_n'], data['as_sold'])
        variance_vs_n1 = calculate_period_variance(data['fct_n'], data['fct_n1'])
        
        # Enhanced status assessment
        if variance_vs_sold >= 10:
            status = "🟢 Excellent"
        elif variance_vs_sold >= 0:
            status = "🟢 On/Above Plan"
        elif variance_vs_sold >= -10:
            status = "🟡 Slight Variance"
        elif variance_vs_sold >= -20:
            status = "🟠 Concerning"
        else:
            status = "🔴 Critical"
        
        quarterly_summary.append({
            'Quarter': quarter,
            'As Sold Plan': format_currency_millions(data['as_sold']),
            'FCT (n-1)': format_currency_millions(data['fct_n1']),
            'FCT (n)': format_currency_millions(data['fct_n']),
            'vs Plan': f"{variance_vs_sold:+.1f}%",
            'vs Previous': f"{variance_vs_n1:+.1f}%",
            'Status': status,
            'Projects': data['project_count']
        })
    
    df_quarterly = pd.DataFrame(quarterly_summary)
    st.dataframe(df_quarterly, use_container_width=True)
    
    # Cash flow insights and recommendations
    render_cash_flow_insights(quarterly_summary, total_as_sold, total_fct_n, overall_variance)

def calculate_cash_flow_trend(portfolio_quarters):
    """Calculate cash flow trend across quarters"""
    if len(portfolio_quarters) < 2:
        return 0
    
    sorted_quarters = sorted(portfolio_quarters.keys())
    values = [portfolio_quarters[q]['fct_n'] for q in sorted_quarters]
    
    if len(values) < 2:
        return 0
    
    # Simple trend calculation (last vs first)
    return calculate_period_variance(values[-1], values[0]) / len(values)

def create_enhanced_quarterly_cash_flow_chart(portfolio_quarters):
    """Create enhanced quarterly cash flow visualization with trend analysis"""
    quarters = sorted(portfolio_quarters.keys())
    as_sold_values = [portfolio_quarters[q]['as_sold']/1000 for q in quarters]
    fct_n1_values = [portfolio_quarters[q]['fct_n1']/1000 for q in quarters]
    fct_n_values = [portfolio_quarters[q]['fct_n']/1000 for q in quarters]
    
    # Calculate variances for color coding
    variances = [calculate_period_variance(portfolio_quarters[q]['fct_n'], portfolio_quarters[q]['as_sold']) for q in quarters]
    colors = ['green' if v >= 0 else 'orange' if v >= -10 else 'red' for v in variances]
    
    fig = make_subplots(
        rows=2, cols=1,
        subplot_titles=('Portfolio Quarterly Cash Flow Analysis', 'Variance Analysis & Trends'),
        specs=[[{"secondary_y": False}], [{"secondary_y": True}]],
        vertical_spacing=0.1
    )
    
    # Main cash flow chart
    fig.add_trace(go.Bar(
        name='As Sold Plan',
        x=quarters,
        y=as_sold_values,
        marker_color='lightblue',
        opacity=0.7
    ), row=1, col=1)
    
    fig.add_trace(go.Bar(
        name='FCT (n-1)',
        x=quarters,
        y=fct_n1_values,
        marker_color='orange',
        opacity=0.7
    ), row=1, col=1)
    
    fig.add_trace(go.Bar(
        name='FCT (n)',
        x=quarters,
        y=fct_n_values,
        marker_color=colors,
        opacity=0.8
    ), row=1, col=1)
    
    # Add trend line for FCT (n)
    fig.add_trace(go.Scatter(
        name='FCT (n) Trend',
        x=quarters,
        y=fct_n_values,
        mode='lines+markers',
        line=dict(color='darkred', width=3),
        marker=dict(size=8, color='darkred')
    ), row=1, col=1)
    
    # Variance chart
    fig.add_trace(go.Bar(
        name='Variance vs Plan',
        x=quarters,
        y=variances,
        marker_color=colors,
        opacity=0.8,
        showlegend=False
    ), row=2, col=1)
    
    # Add cumulative variance trend
    cumulative_variance = []
    running_total = 0
    for v in variances:
        running_total += v
        cumulative_variance.append(running_total / len(cumulative_variance + [0]))
    
    fig.add_trace(go.Scatter(
        name='Cumulative Trend',
        x=quarters,
        y=cumulative_variance,
        mode='lines+markers',
        line=dict(color='purple', width=2),
        yaxis='y2'
    ), row=2, col=1)
    
    # Add threshold lines
    fig.add_hline(y=0, line_dash="dash", line_color="black", row=2, col=1)
    fig.add_hline(y=-10, line_dash="dot", line_color="orange", row=2, col=1)
    fig.add_hline(y=-20, line_dash="dot", line_color="red", row=2, col=1)
    
    fig.update_layout(
        height=800,
        showlegend=True,
        title_text="Enhanced Portfolio Quarterly Cash Flow Analysis"
    )
    
    # Update axes
    fig.update_xaxes(title_text="Financial Year Quarters", row=1, col=1)
    fig.update_yaxes(title_text="Cash Flow (CHF Thousands)", row=1, col=1)
    fig.update_xaxes(title_text="Financial Year Quarters", row=2, col=1)
    fig.update_yaxes(title_text="Variance %", row=2, col=1)
    fig.update_yaxes(title_text="Cumulative Variance %", secondary_y=True, row=2, col=1)
    
    return fig

def render_cash_flow_insights(quarterly_summary, total_as_sold, total_fct_n, overall_variance):
    """Render cash flow insights and recommendations"""
    st.markdown("### 💡 Cash Flow Insights & Recommendations")
    
    col1, col2 = st.columns(2)
    
    with col1:
        # Performance trend analysis
        excellent_quarters = len([q for q in quarterly_summary if "🟢 Excellent" in q['Status']])
        critical_quarters = len([q for q in quarterly_summary if "🔴" in q['Status']])
        total_quarters = len(quarterly_summary)
        
        st.markdown(f"""
        <div class="exec-summary">
            <h4>📊 Quarterly Performance Breakdown</h4>
            <ul>
                <li><strong>Total Quarters Analyzed:</strong> {total_quarters}</li>
                <li><strong>Excellent Performance:</strong> {excellent_quarters} quarters</li>
                <li><strong>Critical Performance:</strong> {critical_quarters} quarters</li>
                <li><strong>Success Rate:</strong> {(total_quarters - critical_quarters)/total_quarters*100:.0f}%</li>
                <li><strong>Portfolio Efficiency:</strong> {(total_fct_n / total_as_sold * 100):.1f}% of plan</li>
            </ul>
        </div>
        """, unsafe_allow_html=True)
    
    with col2:
        # Recommendations based on performance
        recommendations = []
        
        if overall_variance < -20:
            recommendations.append("🔴 **Critical Action Required:** Emergency cash flow review and intervention")
            recommendations.append("📞 **Immediate:** Schedule executive cash flow crisis meeting")
            recommendations.append("💰 **Finance:** Secure additional funding or credit facilities")
        elif overall_variance < -10:
            recommendations.append("🟡 **Monitor Closely:** Cash flow below plan - investigate root causes")
            recommendations.append("📊 **Enhanced Tracking:** Implement weekly cash flow monitoring")
            recommendations.append("🎯 **Action Plans:** Develop quarter-specific recovery strategies")
        elif overall_variance > 15:
            recommendations.append("🟢 **Excellent Performance:** Cash flow significantly ahead of plan")
            recommendations.append("💡 **Opportunity:** Consider acceleration of pipeline projects")
            recommendations.append("📈 **Investment:** Evaluate growth opportunities")
        else:
            recommendations.append("✅ **Stable Performance:** Cash flow within acceptable variance")
            recommendations.append("🎯 **Continue:** Maintain current cash flow management practices")
            recommendations.append("📊 **Optimize:** Look for efficiency improvements")
        
        # Quarter-specific insights
        if critical_quarters > 0:
            recommendations.append(f"⚠️ **Focus Areas:** {critical_quarters} quarters need immediate attention")
        
        if excellent_quarters > total_quarters * 0.6:
            recommendations.append("🏆 **Best Practices:** Document and replicate success factors")
        
        st.markdown(f"""
        <div class="exec-summary">
            <h4>🎯 Strategic Recommendations</h4>
            <ul>
                {''.join([f'<li>{rec}</li>' for rec in recommendations])}
            </ul>
        </div>
        """, unsafe_allow_html=True)

def render_comprehensive_risk_assessment(portfolio_data):
    """Render comprehensive portfolio risk assessment"""
    st.markdown("## ⚠️ Comprehensive Risk Assessment")
    
    # Aggregate risk data
    all_risks = []
    risk_summary = {
        'critical': 0, 'high': 0, 'medium': 0, 'low': 0,
        'total_projects': 0, 'total_risk_value': 0
    }
    
    for project_id, project in portfolio_data.items():
        risk_factors = project['data'].get('risk_factors', [])
        contract_value = safe_get_value(project['data'], 'revenues', 'Contract Price', 'n_ptd')
        
        project_risk_score = 0
        critical_risks = 0
        high_risks = 0
        
        for risk in risk_factors:
            risk['project_id'] = project_id
            risk['project_name'] = project['name']
            risk['contract_value'] = contract_value
            all_risks.append(risk)
            
            if risk['severity'] == 'Critical':
                critical_risks += 1
                project_risk_score += 4
            elif risk['severity'] == 'High':
                high_risks += 1
                project_risk_score += 3
            elif risk['severity'] == 'Medium':
                project_risk_score += 2
            else:
                project_risk_score += 1
        
        # Update summary
        if critical_risks > 0:
            risk_summary['critical'] += 1
        elif high_risks > 0:
            risk_summary['high'] += 1
        elif project_risk_score > 3:
            risk_summary['medium'] += 1
        else:
            risk_summary['low'] += 1
        
        risk_summary['total_projects'] += 1
        if critical_risks > 0 or high_risks > 0:
            risk_summary['total_risk_value'] += contract_value
    
    # Risk overview dashboard
    col1, col2, col3, col4, col5 = st.columns(5)
    
    with col1:
        critical_pct = (risk_summary['critical'] / risk_summary['total_projects'] * 100) if risk_summary['total_projects'] > 0 else 0
        st.markdown(f"""
        <div class="risk-card risk-critical">
            <h4>🔴 Critical Risk</h4>
            <h2>{risk_summary['critical']}</h2>
            <p>{critical_pct:.1f}% of portfolio</p>
        </div>
        """, unsafe_allow_html=True)
    
    with col2:
        high_pct = (risk_summary['high'] / risk_summary['total_projects'] * 100) if risk_summary['total_projects'] > 0 else 0
        st.markdown(f"""
        <div class="risk-card risk-high">
            <h4>🟠 High Risk</h4>
            <h2>{risk_summary['high']}</h2>
            <p>{high_pct:.1f}% of portfolio</p>
        </div>
        """, unsafe_allow_html=True)
    
    with col3:
        medium_pct = (risk_summary['medium'] / risk_summary['total_projects'] * 100) if risk_summary['total_projects'] > 0 else 0
        st.markdown(f"""
        <div class="risk-card risk-medium">
            <h4>🟡 Medium Risk</h4>
            <h2>{risk_summary['medium']}</h2>
            <p>{medium_pct:.1f}% of portfolio</p>
        </div>
        """, unsafe_allow_html=True)
    
    with col4:
        low_pct = (risk_summary['low'] / risk_summary['total_projects'] * 100) if risk_summary['total_projects'] > 0 else 0
        st.markdown(f"""
        <div class="risk-card risk-low">
            <h4>🟢 Low Risk</h4>
            <h2>{risk_summary['low']}</h2>
            <p>{low_pct:.1f}% of portfolio</p>
        </div>
        """, unsafe_allow_html=True)
    
    with col5:
        risk_value_pct = (risk_summary['total_risk_value'] / sum(safe_get_value(p['data'], 'revenues', 'Contract Price', 'n_ptd') for p in portfolio_data.values()) * 100) if portfolio_data else 0
        st.markdown(f"""
        <div class="risk-card risk-medium">
            <h4>💰 Value at Risk</h4>
            <h2>{risk_value_pct:.1f}%</h2>
            <p>{format_currency_millions(risk_summary['total_risk_value'])}</p>
        </div>
        """, unsafe_allow_html=True)
    
    # Risk details table
    if all_risks:
        st.markdown("### 📋 Detailed Risk Register")
        
        risk_details = []
        for risk in all_risks:
            severity_icon = "🔴" if risk['severity'] == 'Critical' else "🟠" if risk['severity'] == 'High' else "🟡" if risk['severity'] == 'Medium' else "🟢"
            impact_icon = "⚡" if risk['impact'] == 'High' else "⚡" if risk['impact'] == 'Medium' else "💧"
            
            risk_details.append({
                'Project': risk['project_id'],
                'Risk Type': risk['type'],
                'Severity': f"{severity_icon} {risk['severity']}",
                'Impact': f"{impact_icon} {risk['impact']}",
                'Description': risk['description'][:60] + "..." if len(risk['description']) > 60 else risk['description'],
                'Recommendation': risk['recommendation'][:50] + "..." if len(risk['recommendation']) > 50 else risk['recommendation']
            })
        
        df_risks = pd.DataFrame(risk_details)
        st.dataframe(df_risks, use_container_width=True)
    
    # Risk mitigation recommendations
    render_risk_mitigation_recommendations(all_risks, risk_summary)

def render_risk_mitigation_recommendations(all_risks, risk_summary):
    """Render risk mitigation recommendations"""
    st.markdown("### 💡 Risk Mitigation Strategy")
    
    # Categorize risks by type
    risk_categories = {}
    for risk in all_risks:
        category = risk['type']
        if category not in risk_categories:
            risk_categories[category] = {'critical': 0, 'high': 0, 'medium': 0, 'low': 0}
        risk_categories[category][risk['severity'].lower()] += 1
    
    col1, col2 = st.columns(2)
    
    with col1:
        st.markdown("#### 🎯 Immediate Actions Required")
        
        critical_actions = []
        high_actions = []
        
        for risk in all_risks:
            if risk['severity'] == 'Critical':
                critical_actions.append(f"**{risk['project_id']}** - {risk['type']}: {risk['recommendation']}")
            elif risk['severity'] == 'High':
                high_actions.append(f"**{risk['project_id']}** - {risk['type']}: {risk['recommendation']}")
        
        if critical_actions:
            st.error("🔴 **Critical Actions:**")
            for action in critical_actions[:5]:  # Show top 5
                st.write(f"• {action}")
        
        if high_actions:
            st.warning("🟠 **High Priority Actions:**")
            for action in high_actions[:5]:  # Show top 5
                st.write(f"• {action}")
    
    with col2:
        st.markdown("#### 📊 Risk Category Analysis")
        
        for category, counts in risk_categories.items():
            total_risks = sum(counts.values())
            severity_score = (counts['critical'] * 4 + counts['high'] * 3 + 
                            counts['medium'] * 2 + counts['low'] * 1)
            # Add specific insights for new risk categories
            if category == 'Contingency Risk':
                st.write(f"**{category}:** (Buffer Management)")
            elif category == 'WP Performance Risk':
                st.write(f"**{category}:** (Execution Quality)")
            elif category == 'Financial Buffer Risk':
                st.write(f"**{category}:** (Combined Risk)")
            else:
                st.write(f"**{category}:**")       

            st.write(f"**{category}:**")
            st.write(f"• Total: {total_risks} risks (Score: {severity_score})")
            st.write(f"• Critical: {counts['critical']}, High: {counts['high']}, Medium: {counts['medium']}, Low: {counts['low']}")
            st.write("")
    
    # Portfolio-level recommendations
    st.markdown("#### 🏢 Portfolio-Level Recommendations")
    
    portfolio_recommendations = []
    
    if risk_summary['critical'] > 0:
        portfolio_recommendations.append("🔴 **Executive Escalation Required:** Immediate intervention needed for critical risk projects")
    
    if (risk_summary['critical'] + risk_summary['high']) / risk_summary['total_projects'] > 0.3:
        portfolio_recommendations.append("⚠️ **Portfolio Risk Alert:** >30% of projects at high/critical risk - review resource allocation")
    
    if 'Margin Risk' in risk_categories and sum(risk_categories['Margin Risk'].values()) > 2:
        portfolio_recommendations.append("💰 **Margin Protection Program:** Implement enhanced cost control and revenue optimization")
    
    if 'Cost Commitment' in risk_categories and sum(risk_categories['Cost Commitment'].values()) > 2:
        portfolio_recommendations.append("🏗️ **Procurement Controls:** Strengthen approval processes and cost commitment monitoring")

    if 'Contingency Risk' in risk_categories and risk_categories['Contingency Risk']['critical'] > 0:
        portfolio_recommendations.append("💰 **Contingency Crisis:** Multiple projects with depleted contingencies - portfolio-wide risk review needed")

    if 'WP Performance Risk' in risk_categories and sum(risk_categories['WP Performance Risk'].values()) > 2:
        portfolio_recommendations.append("📊 **Systemic Estimation Issue:** Multiple projects with WP overruns - review estimation methodology")

    if 'Financial Buffer Risk' in risk_categories:
        portfolio_recommendations.append("🛡️ **Buffer Protection:** Projects operating with minimal financial cushion - prioritize risk mitigation")
    
    portfolio_recommendations.append("📊 **Enhanced Monitoring:** Implement weekly risk reviews for high-risk projects")
    portfolio_recommendations.append("🎯 **Best Practice Sharing:** Transfer successful risk mitigation strategies across portfolio")
    
    for recommendation in portfolio_recommendations:
        st.info(recommendation)

def render_portfolio_revenue_analytics(portfolio_data):
    """Render comprehensive portfolio revenue analytics - ENHANCED VERSION"""
    st.markdown("## 📊 Portfolio Revenue Analytics Dashboard")
    
    # Debug mode for data verification
    debug_mode = st.checkbox("Show Revenue Data Debug Info", value=False, key="revenue_debug")
    
    # Collect available quarters across all projects
    available_quarters = set()
    for project_id, project in portfolio_data.items():
        quarterly_data = project['data'].get('quarterly', {})
        for quarter in quarterly_data.keys():
            if quarter != 'Total':  # Exclude 'Total' entry
                available_quarters.add(quarter)
    
    available_quarters = sorted(list(available_quarters))
    
    if not available_quarters:
        st.warning("⚠️ No quarterly revenue data found in the uploaded files.")
        st.info("Please ensure your Excel templates contain quarterly revenue data in the 'Project Revenues' sheet.")
        return
    
    # Quarter selector for focused analysis
    st.markdown("### 🎯 Quarter Selection")
    col_q1, col_q2, col_q3 = st.columns([2, 3, 5])
    
    with col_q1:
        selected_quarter = st.selectbox(
            "Select Quarter for Analysis",
            available_quarters,
            index=len(available_quarters)-1 if available_quarters else 0,  # Default to latest quarter
            help="Choose which quarter to analyze for revenue performance"
        )
    
    with col_q2:
        st.info(f"📅 Analyzing: **{selected_quarter}**")
    
    # Collect and validate quarterly data
    portfolio_quarters = {'Q1': [], 'Q2': [], 'Q3': [], 'Q4': []}
    project_performance = []
    projects_with_data = []
    
    for project_id, project in portfolio_data.items():
        quarterly_data = project['data'].get('quarterly', {})
        revenues_data = project['data'].get('revenues', {})
        contract_value = safe_get_value(project['data'], 'revenues', 'Contract Price', 'n_ptd')
        
        if debug_mode:
            st.write(f"**Project {project_id}:**")
            st.write(f"- Contract Value: {contract_value:,.0f}")
            st.write(f"- Quarterly Data Available: {bool(quarterly_data)}")
            if quarterly_data:
                st.json(quarterly_data)
        
        # Validate data exists and is meaningful
        if quarterly_data and contract_value > 0:
            total_actual = 0
            total_budget = 0
            has_valid_data = False
            
            # Collect data for all quarters (for other visualizations)
            for quarter in ['Q1', 'Q2', 'Q3', 'Q4']:
                q_data = quarterly_data.get(quarter, {})
                
                # Try different field names based on template version
                actual = q_data.get('actuals', 0) or q_data.get('actual', 0) or q_data.get('revenue', 0)
                budget = q_data.get('budget', 0) or q_data.get('planned', 0)
                
                # If no budget, use gap_to_close + actuals as approximation
                if budget == 0 and 'gap_to_close' in q_data:
                    budget = actual + q_data.get('gap_to_close', 0)
                
                if actual > 0 or budget > 0:
                    has_valid_data = True
                
                portfolio_quarters[quarter].append({
                    'project_id': project_id,
                    'project_name': project['name'][:25],
                    'actual': actual,
                    'budget': budget,
                    'variance': q_data.get('delta_pct', 0),
                    'contract_value': contract_value
                })
                
                total_actual += actual
                total_budget += budget
            
            # Get specific quarter data for performance calculation
            selected_q_data = quarterly_data.get(selected_quarter, {})
            q_actual = selected_q_data.get('actuals', 0) or selected_q_data.get('actual', 0) or selected_q_data.get('revenue', 0)
            q_budget = selected_q_data.get('budget', 0) or selected_q_data.get('planned', 0)
            
            # If no budget, use gap_to_close + actuals as approximation
            if q_budget == 0 and 'gap_to_close' in selected_q_data:
                q_budget = q_actual + selected_q_data.get('gap_to_close', 0)
            
            # Only include projects with valid data for selected quarter
            if q_budget > 0:  # Only include if there's a budget for the quarter
                projects_with_data.append(project_id)
                
                # Calculate quarterly performance
                q_performance = (q_actual / q_budget * 100) if q_budget > 0 else 0
                
                project_performance.append({
                    'project_id': project_id,
                    'project_name': project['name'],
                    'contract_value': contract_value,
                    'quarterly_actual': q_actual,
                    'quarterly_budget': q_budget,
                    'quarterly_performance': q_performance,
                    'total_actual': total_actual,  # Keep for other charts
                    'total_budget': total_budget,   # Keep for other charts
                    'quarter': selected_quarter
                })
    
    if not projects_with_data:
        st.warning(f"⚠️ No valid revenue data found for {selected_quarter}.")
        st.info("Please select a different quarter or ensure your Excel templates contain quarterly budget data.")
        return
    
    # Create visualizations with enhanced styling
    col1, col2 = st.columns(2)
    
    with col1:
        # 1. Enhanced Stacked Bar Chart (keep existing for all quarters overview)
        st.markdown("#### 📊 Quarterly Revenue Distribution by Project")
        
        fig1 = go.Figure()
        
        # Create stacked bars for each project
        for i, quarter in enumerate(['Q1', 'Q2', 'Q3', 'Q4']):
            q_revenues = [item for item in portfolio_quarters[quarter] if item['project_id'] in projects_with_data]
            
            # Use different shades of blue for quarters
            colors = ['#084594', '#2171b5', '#4292c6', '#6baed6']
            
            fig1.add_trace(go.Bar(
                name=quarter,
                x=[item['project_id'] for item in q_revenues],
                y=[item['actual']/1000 for item in q_revenues],
                text=[f"{item['actual']/1000:.0f}" if item['actual'] > 0 else "" for item in q_revenues],
                textposition='inside',
                marker_color=colors[i],
                hovertemplate='%{x}<br>%{fullData.name}: CHF %{y:.1f}K<br>Budget: CHF %{customdata:.1f}K<extra></extra>',
                customdata=[item['budget']/1000 for item in q_revenues]
            ))
        
        fig1.update_layout(
            barmode='stack',
            height=450,
            xaxis_title='Projects',
            yaxis_title='Revenue (CHF Thousands)',
            showlegend=True,
            legend=dict(orientation="h", yanchor="bottom", y=1.02, xanchor="right", x=1),
            plot_bgcolor='rgba(0,0,0,0)',
            bargap=0.15,
            bargroupgap=0.1
        )
        
        fig1.update_xaxes(tickangle=-45)
        
        st.plotly_chart(fig1, use_container_width=True)
    
    with col2:
        # 2. ENHANCED Quarterly Performance Scatter Plot
        st.markdown(f"#### 🎯 {selected_quarter} Revenue Performance vs Contract Size")
        
        fig2 = go.Figure()
        
        # Filter for projects with valid quarterly performance
        valid_projects = [p for p in project_performance if p['quarterly_performance'] > 0 or p['quarterly_budget'] > 0]
        
        if valid_projects:
            performances = [p['quarterly_performance'] for p in valid_projects]
            colors = ['#00a651' if p >= 95 else '#ff9900' if p >= 85 else '#ee2724' for p in performances]
            
            # Calculate bubble sizes based on quarterly revenue (not contract value)
            max_q_revenue = max([p['quarterly_actual'] for p in valid_projects]) if valid_projects else 1
            
            # Prevent division by zero
            if max_q_revenue == 0:
                # If all projects have zero actual revenue, use budget for sizing
                max_q_budget = max([p['quarterly_budget'] for p in valid_projects]) if valid_projects else 1
                if max_q_budget > 0:
                    bubble_sizes = [max(15, min(50, (p['quarterly_budget']/max_q_budget)*50)) for p in valid_projects]
                    size_note = "Bubble size = Quarterly Budget (no actuals yet)"
                else:
                    # If both actual and budget are zero, use uniform size
                    bubble_sizes = [25 for p in valid_projects]  # Default size
                    size_note = "Uniform bubble size (no data)"
            else:
                bubble_sizes = [max(15, min(50, (p['quarterly_actual']/max_q_revenue)*50)) for p in valid_projects]
                size_note = "Bubble size = Quarterly Revenue"
            
            fig2.add_trace(go.Scatter(
                x=[p['contract_value']/1000 for p in valid_projects],
                y=[p['quarterly_performance'] for p in valid_projects],
                mode='markers+text',
                marker=dict(
                    size=bubble_sizes,
                    color=colors,
                    opacity=0.7,
                    line=dict(width=2, color='white')
                ),
                text=[p['project_id'] for p in valid_projects],
                textposition='top center',
                textfont=dict(size=10),
                customdata=[[p['quarterly_actual']/1000, p['quarterly_budget']/1000, p['quarterly_performance']] 
                           for p in valid_projects],
                hovertemplate='<b>%{text}</b><br>Contract: CHF %{x:.2f}M<br>' + 
                             f'{selected_quarter} Performance: %{{customdata[2]:.1f}}%<br>' +
                             f'{selected_quarter} Actual: CHF %{{customdata[0]:.1f}}K<br>' +
                             f'{selected_quarter} Budget: CHF %{{customdata[1]:.1f}}K<br>' +
                             f'{size_note}<extra></extra>',
                name=''
            ))
            
            # Add reference lines
            fig2.add_hline(y=100, line_dash="dash", line_color="green", 
                          annotation_text="Target", annotation_position="right")
            fig2.add_hline(y=90, line_dash="dot", line_color="orange", 
                          annotation_text="Warning", annotation_position="right")
            
            # Add quadrant shading
            fig2.add_hrect(y0=95, y1=130, fillcolor="green", opacity=0.1, line_width=0)
            fig2.add_hrect(y0=85, y1=95, fillcolor="orange", opacity=0.1, line_width=0)
            fig2.add_hrect(y0=0, y1=85, fillcolor="red", opacity=0.1, line_width=0)
	    
        else:
            # No valid projects for this quarter
            st.info(f"No projects with budget data for {selected_quarter}")
        
        fig2.update_layout(
            height=450,
            xaxis_title='Contract Value (CHF Millions)',
            yaxis_title=f'{selected_quarter} Revenue Performance %',
            showlegend=False,
            plot_bgcolor='rgba(0,0,0,0)',
            xaxis=dict(showgrid=True, gridcolor='lightgray', zeroline=False),
            yaxis=dict(showgrid=True, gridcolor='lightgray', zeroline=False, range=[0, 130]),
            title_font_size=14
        )
        
        st.plotly_chart(fig2, use_container_width=True)
    
    # 3. Enhanced Time Series with quarterly focus
    st.markdown("#### 📈 Portfolio Revenue Trend Analysis")
    
    # Add quarter performance summary
    selected_q_metrics = [p for p in project_performance if p['quarter'] == selected_quarter]
    if selected_q_metrics:
        total_q_actual = sum([p['quarterly_actual'] for p in selected_q_metrics])
        total_q_budget = sum([p['quarterly_budget'] for p in selected_q_metrics])
        overall_q_performance = (total_q_actual / total_q_budget * 100) if total_q_budget > 0 else 0
        
        col_m1, col_m2, col_m3, col_m4 = st.columns(4)
        with col_m1:
            st.metric(f"{selected_quarter} Budget", format_currency_millions(total_q_budget))
        with col_m2:
            st.metric(f"{selected_quarter} Actual", format_currency_millions(total_q_actual))
        with col_m3:
            perf_delta = overall_q_performance - 100
            st.metric(f"{selected_quarter} Performance", f"{overall_q_performance:.1f}%", 
                     f"{perf_delta:+.1f}%")
        with col_m4:
            on_target_count = len([p for p in selected_q_metrics if p['quarterly_performance'] >= 95])
            st.metric("Projects On Target", f"{on_target_count}/{len(selected_q_metrics)}")
    
    quarterly_totals = []
    for quarter in ['Q1', 'Q2', 'Q3', 'Q4']:
        q_data = [item for item in portfolio_quarters[quarter] if item['project_id'] in projects_with_data]
        total_actual = sum([item['actual'] for item in q_data])
        total_budget = sum([item['budget'] for item in q_data])
        
        # Calculate variance properly
        if total_budget > 0:
            variance = ((total_actual - total_budget) / total_budget * 100)
        else:
            variance = 0
            
        quarterly_totals.append({
            'quarter': quarter,
            'actual': total_actual,
            'budget': total_budget,
            'variance': variance,
            'is_selected': quarter == selected_quarter
        })
    
    # Create enhanced subplot figure
    fig3 = make_subplots(
        rows=1, cols=2,
        subplot_titles=('Portfolio Revenue Trend', f'{selected_quarter} Performance Distribution'),
        specs=[[{"secondary_y": True}, {"type": "bar"}]],
        horizontal_spacing=0.12
    )
    
    # Revenue trend with improved styling and selected quarter highlight
    quarters_list = [q['quarter'] for q in quarterly_totals]
    budget_colors = ['lightblue' if not q['is_selected'] else 'darkblue' for q in quarterly_totals]
    actual_colors = ['darkblue' if not q['is_selected'] else 'darkgreen' for q in quarterly_totals]
    
    fig3.add_trace(go.Bar(
        name='Budget',
        x=quarters_list,
        y=[q['budget']/1000000 for q in quarterly_totals],
        marker_color=budget_colors,
        opacity=0.7,
        text=[f"{q['budget']/1000000:.2f}" for q in quarterly_totals],
        textposition='outside',
        texttemplate='%{text}M'
    ), row=1, col=1)
    
    fig3.add_trace(go.Bar(
        name='Actual',
        x=quarters_list,
        y=[q['actual']/1000000 for q in quarterly_totals],
        marker_color=actual_colors,
        text=[f"{q['actual']/1000000:.2f}" for q in quarterly_totals],
        textposition='outside',
        texttemplate='%{text}M'
    ), row=1, col=1)
    
    # Add cumulative line with markers
    cumulative_actual = []
    cumulative_sum = 0
    for q in quarterly_totals:
        cumulative_sum += q['actual']/1000000
        cumulative_actual.append(cumulative_sum)
    
    fig3.add_trace(go.Scatter(
        name='Cumulative Actual',
        x=quarters_list,
        y=cumulative_actual,
        mode='lines+markers+text',
        line=dict(color='red', width=3),
        marker=dict(size=10),
        text=[f"{v:.1f}M" for v in cumulative_actual],
        textposition='top center',
        yaxis='y2'
    ), row=1, col=1, secondary_y=True)
    
    # Performance distribution for selected quarter
    if selected_q_metrics:
        performance_ranges = {
            '0-50%': len([p for p in selected_q_metrics if 0 <= p['quarterly_performance'] < 50]),
            '50-85%': len([p for p in selected_q_metrics if 50 <= p['quarterly_performance'] < 85]),
            '85-95%': len([p for p in selected_q_metrics if 85 <= p['quarterly_performance'] < 95]),
            '95-100%': len([p for p in selected_q_metrics if 95 <= p['quarterly_performance'] <= 100]),
            '>100%': len([p for p in selected_q_metrics if p['quarterly_performance'] > 100])
        }
        
        fig3.add_trace(go.Bar(
            name='Projects',
            x=list(performance_ranges.keys()),
            y=list(performance_ranges.values()),
            marker_color=['red', 'orange', 'yellow', 'lightgreen', 'darkgreen'],
            text=list(performance_ranges.values()),
            textposition='outside',
            showlegend=False
        ), row=1, col=2)
    
    fig3.update_layout(
        height=450,
        showlegend=True,
        legend=dict(orientation="h", yanchor="bottom", y=-0.2),
        plot_bgcolor='rgba(0,0,0,0)',
        title_text=f"Revenue Analysis with {selected_quarter} Focus"
    )
    
    fig3.update_xaxes(title_text="Quarter", row=1, col=1)
    fig3.update_xaxes(title_text="Performance Range", row=1, col=2)
    fig3.update_yaxes(title_text="Revenue (CHF M)", row=1, col=1)
    fig3.update_yaxes(title_text="Cumulative (CHF M)", secondary_y=True, row=1, col=1)
    fig3.update_yaxes(title_text="Number of Projects", row=1, col=2)
    
    st.plotly_chart(fig3, use_container_width=True)
    
    # 4. Enhanced Project Ranking Table - Focused on Selected Quarter
    st.markdown(f"#### 🏆 {selected_quarter} Project Performance Ranking")
    
    if selected_q_metrics:
        # Create enhanced ranking dataframe
        ranking_data = []
        for i, p in enumerate(sorted(selected_q_metrics, key=lambda x: x['quarterly_performance'], reverse=True)):
            # Determine performance status
            if p['quarterly_performance'] >= 100:
                perf_icon = '🌟'  # Exceeding
            elif p['quarterly_performance'] >= 95:
                perf_icon = '🟢'  # On target
            elif p['quarterly_performance'] >= 85:
                perf_icon = '🟡'  # Slightly below
            else:
                perf_icon = '🔴'  # Significantly below
            
            ranking_data.append({
                'Rank': i + 1,
                'Project': p['project_id'],
                'Name': p['project_name'][:30] + '...' if len(p['project_name']) > 30 else p['project_name'],
                'Contract Value': format_currency_millions(p['contract_value']),
                f'{selected_quarter} Budget': format_currency_thousands(p['quarterly_budget']),
                f'{selected_quarter} Actual': format_currency_thousands(p['quarterly_actual']),
		f'{selected_quarter} Performance': p['quarterly_performance'],
                'Variance': format_currency_thousands(p['quarterly_actual'] - p['quarterly_budget']),
                'Status': perf_icon
            })
        
        df_ranking = pd.DataFrame(ranking_data)
        
        # Apply conditional formatting
        st.dataframe(
            df_ranking,
            use_container_width=True,
            hide_index=True,
            column_config={
                "Rank": st.column_config.NumberColumn(width="small"),
                "Status": st.column_config.TextColumn(width="small"),
                f"{selected_quarter} Performance": st.column_config.ProgressColumn(
                    help="Revenue achievement rate for selected quarter",
                    format="%.1f%%",
                    min_value=0,
                    max_value=150,
                ),
            }
        )
        
        # Performance insights for selected quarter
        st.markdown(f"### 💡 {selected_quarter} Performance Insights")
        
        col_i1, col_i2 = st.columns(2)
        
        with col_i1:
            # Performance distribution
            excellent_count = len([p for p in selected_q_metrics if p['quarterly_performance'] >= 100])
            on_target_count = len([p for p in selected_q_metrics if 95 <= p['quarterly_performance'] < 100])
            below_target_count = len([p for p in selected_q_metrics if p['quarterly_performance'] < 95])
            
            st.markdown(f"""
            <div class="exec-summary">
                <h4>📊 {selected_quarter} Performance Summary</h4>
                <ul>
                    <li><strong>Exceeding Target (≥100%):</strong> {excellent_count} projects</li>
                    <li><strong>On Target (95-100%):</strong> {on_target_count} projects</li>
                    <li><strong>Below Target (<95%):</strong> {below_target_count} projects</li>
                    <li><strong>Overall Achievement:</strong> {overall_q_performance:.1f}%</li>
                    <li><strong>Revenue Gap:</strong> {format_currency_thousands(total_q_actual - total_q_budget)}</li>
                </ul>
            </div>
            """, unsafe_allow_html=True)
        
        with col_i2:
            # Top and bottom performers
            top_performers = sorted(selected_q_metrics, key=lambda x: x['quarterly_performance'], reverse=True)[:3]
            bottom_performers = sorted(selected_q_metrics, key=lambda x: x['quarterly_performance'])[:3]
            
            recommendations = []
            if overall_q_performance < 95:
                recommendations.append(f"🔴 **Revenue Gap Alert:** {selected_quarter} is {95-overall_q_performance:.1f}% below target")
            if below_target_count > len(selected_q_metrics) * 0.5:
                recommendations.append(f"⚠️ **Portfolio Risk:** Over 50% of projects below target in {selected_quarter}")
            if excellent_count > 0:
                recommendations.append(f"🌟 **Best Practices:** Study {excellent_count} exceeding projects for lessons learned")
            
            recommendations.extend([
                "📊 **Focus Areas:** Prioritize bottom 3 performers for intervention",
                "📈 **Revenue Recovery:** Implement catch-up plans for below-target projects"
            ])
            
            st.markdown(f"""
            <div class="exec-summary">
                <h4>🎯 Strategic Recommendations</h4>
                <ul>
                    {''.join([f'<li>{rec}</li>' for rec in recommendations])}
                </ul>
            </div>
            """, unsafe_allow_html=True)



def render_executive_project_table(portfolio_data):
    """Render comprehensive executive project summary table"""
    st.markdown("## 📋 Executive Project Summary")
    
    projects_summary = []
    
    for project_id, project in portfolio_data.items():
        try:
            data = project['data']
            
            # Key metrics
            contract_value = safe_get_value(data, 'revenues', 'Contract Price', 'n_ptd')
            revenue_ptd = safe_get_value(data, 'revenues', 'Revenues', 'n_ptd')
            poc_current = safe_get_value(data, 'revenues', 'POC%', 'n_ptd')
            poc_previous = safe_get_value(data, 'revenues', 'POC%', 'n1_ptd')
            
            # Enhanced cost metrics
            cost_analysis = data.get('cost_analysis', {})
            cm1_pct = cost_analysis.get('cm1_pct_fct_n', 0)
            cm2_pct = cost_analysis.get('cm2_pct_fct_n', 0)
            committed_ratio = cost_analysis.get('committed_ratio', 0)
            cost_variance_pct = cost_analysis.get('cost_variance_pct', 0)
            
            # Earned value metrics
            earned_value = data.get('earned_value', {})
            cpi = earned_value.get('cost_performance_index', 1.0)
            spi = earned_value.get('schedule_performance_index', 1.0)
            
            # Risk assessment
            risk_factors = data.get('risk_factors', [])
            risk_score = len([r for r in risk_factors if r['severity'] in ['Critical', 'High']])
            
            # Status indicators
            cm2_icon, _, cm2_class = get_traffic_light_status(cm2_pct, EXECUTIVE_THRESHOLDS['cm2_margin'])
            committed_icon, _, committed_class = get_traffic_light_status(
                committed_ratio, EXECUTIVE_THRESHOLDS['committed_vs_budget'], reverse=True
            )
            
            # FIX #6: POC velocity using corrected calculation
            poc_velocity = calculate_poc_velocity(poc_current, poc_previous)        
            poc_icon_raw, _, poc_class = get_traffic_light_status(poc_velocity, EXECUTIVE_THRESHOLDS['poc_velocity'])
            poc_icon_adjusted, poc_status_adjusted, poc_class_adjusted = get_poc_velocity_status_with_maturity(poc_velocity, poc_current)

            # Store the raw icon before potentially overwriting it
            poc_icon = poc_icon_raw
            if poc_icon != poc_icon_adjusted:
                poc_icon = poc_icon_adjusted  # Use adjusted icon
                poc_class = poc_class_adjusted  # Use adjusted class

            poc_display = f"{poc_velocity:+.1f}% {poc_icon_adjusted}"
            if poc_icon_raw != poc_icon_adjusted:
                # Add indicator that status was adjusted for maturity
                poc_display += " 📊"  # Chart icon indicates maturity-adjusted
            
            # FIX #4: Cost performance using CPI directly
            cpi_icon, _, _ = get_traffic_light_status(cpi, EXECUTIVE_THRESHOLDS['cost_performance_index'])
            
            # FIX #4: Schedule performance using SPI directly
            spi_icon, _, _ = get_traffic_light_status(spi, EXECUTIVE_THRESHOLDS['schedule_performance_index'])
            
            projects_summary.append({
                'Project': project_id,
                'Name': project['name'][:30] + "..." if len(project['name']) > 30 else project['name'],
                'Contract Value': format_currency_millions(contract_value),
                'POC %': f"{poc_current:.1f}%",
                'POC Velocity': f"{poc_velocity:+.1f}% {poc_icon}",
                'CM1 %': f"{cm1_pct:.1f}%",
                'CM2 %': f"{cm2_pct:.1f}%",
                'CM2 Status': f"{cm2_icon}",
                'Cost Perf (CPI)': f"{cpi:.2f} {cpi_icon}",
                'Schedule (SPI)': f"{spi:.2f} {spi_icon}",
                'Committed Ratio': f"{committed_ratio:.2f} {committed_icon}",
                'Cost Variance': f"{cost_variance_pct:+.1f}%",
                'Risk Score': f"{risk_score}/10" if risk_score <= 10 else "10+",
                'Overall Status': get_overall_project_status(cm2_class,committed_class,poc_class,poc_current,poc_velocity)
            })
            
        except Exception as e:
            st.warning(f"Error processing project {project_id}: {str(e)}")
            continue
    
    if projects_summary:
        df_summary = pd.DataFrame(projects_summary)
        st.dataframe(df_summary, use_container_width=True)
        
        with st.expander("ℹ️ Understanding POC Velocity Indicators", expanded=False):
            st.markdown("""
            **POC Velocity** shows the rate of progress completion per month:
            - 🟢 **Green**: Meeting or exceeding expected velocity for project maturity
            - 🟡 **Yellow**: Below expected velocity but within acceptable range
            - 🔴 **Red**: Significantly below expected velocity
            - 📊 **Chart Icon**: Indicates status was adjusted based on project maturity
        
            **Expected Velocity by Project Maturity:**
            - 0-40% Complete: ~10%+ per month
            - 40-60% Complete: ~7% per month
            - 60-80% Complete: ~5% per month
            - 80-90% Complete: ~3% per month
            - 90-95% Complete: ~2% per month
            - 95%+ Complete: ~1% per month
        
            Projects naturally slow down as they approach completion due to:
            - Final testing and quality assurance
            - Documentation and handover activities
            - Punch list items and minor corrections
            """)

        # Enhanced executive insights
        render_comprehensive_executive_insights(projects_summary)
    else:
        st.warning("📊 No project data available for executive summary.")

def render_comprehensive_executive_insights(projects_summary):
    """Render comprehensive executive insights and strategic recommendations"""
    st.markdown("### 💡 Executive Insights & Strategic Recommendations")
    
    total_projects = len(projects_summary)
    excellent_projects = len([p for p in projects_summary if "🟢" in p['Overall Status']])
    critical_projects = len([p for p in projects_summary if "🔴" in p['Overall Status']])
    high_risk_projects = len([p for p in projects_summary if int(p['Risk Score'].split('/')[0]) >= 5])
    
    col1, col2, col3 = st.columns(3)
    
    with col1:
        # Portfolio health overview
        st.markdown(f"""
        <div class="exec-summary">
            <h4>📊 Portfolio Health Dashboard</h4>
            <ul>
                <li><strong>Total Projects:</strong> {total_projects}</li>
                <li><strong>Excellent Performance:</strong> {excellent_projects} ({excellent_projects/total_projects*100:.0f}%)</li>
                <li><strong>Critical Attention:</strong> {critical_projects} ({critical_projects/total_projects*100:.0f}%)</li>
                <li><strong>High Risk Projects:</strong> {high_risk_projects} ({high_risk_projects/total_projects*100:.0f}%)</li>
                <li><strong>Portfolio Health Score:</strong> {((excellent_projects * 4 + (total_projects - excellent_projects - critical_projects) * 2 + critical_projects * 1) / (total_projects * 4) * 100):.0f}%</li>
            </ul>
        </div>
        """, unsafe_allow_html=True)
    
    with col2:
        # Performance analysis
        avg_cm2 = np.mean([float(p['CM2 %'].replace('%', '')) for p in projects_summary if '%' in p['CM2 %']])
        avg_cpi = np.mean([float(p['Cost Perf (CPI)'].split()[0]) for p in projects_summary])
        avg_committed_ratio = np.mean([float(p['Committed Ratio'].split()[0]) for p in projects_summary])
        
        st.markdown(f"""
        <div class="exec-summary">
            <h4>📈 Performance Metrics</h4>
            <ul>
                <li><strong>Average CM2 Margin:</strong> {avg_cm2:.1f}%</li>
                <li><strong>Average Cost Performance:</strong> {avg_cpi:.2f}</li>
                <li><strong>Average Committed Ratio:</strong> {avg_committed_ratio:.2f}</li>
                <li><strong>Portfolio Efficiency:</strong> {(100 - (avg_committed_ratio - 1) * 100):.1f}%</li>
            </ul>
        </div>
        """, unsafe_allow_html=True)
    
    with col3:
        # Strategic recommendations
        recommendations = []
        
        if critical_projects > 0:
            recommendations.append(f"🔴 **Immediate Executive Action:** {critical_projects} projects need intervention")
        
        if critical_projects / total_projects > 0.25:
            recommendations.append("⚠️ **Portfolio Risk Alert:** >25% critical - resource reallocation needed")
        
        if high_risk_projects / total_projects > 0.3:
            recommendations.append("🎯 **Risk Management:** Enhanced monitoring for high-risk projects")
        
        if avg_cm2 < 10:
            recommendations.append("💰 **Margin Improvement:** Portfolio CM2 below target - cost optimization required")
        
        if avg_cpi < 0.95:
            recommendations.append("🏗️ **Cost Control:** Portfolio cost performance below standard")
        
        if excellent_projects / total_projects > 0.7:
            recommendations.append("✅ **Best Practices:** Share success factors across portfolio")
        
        recommendations.append("📊 **Regular Reviews:** Maintain weekly executive dashboard monitoring")
        
        st.markdown(f"""
        <div class="exec-summary">
            <h4>🎯 Strategic Actions</h4>
            <ul>
                {''.join([f'<li>{rec}</li>' for rec in recommendations])}
            </ul>
        </div>
        """, unsafe_allow_html=True)

def create_enhanced_portfolio_summary(portfolio_data):
    """Create comprehensive portfolio summary with all KPIs"""
    if not portfolio_data:
        return None
    
    metrics = {
        'total_contract_value_n': 0, 'total_contract_value_n1': 0,
        'total_revenue_n': 0, 'total_revenue_n1': 0,
        'total_cash_in_n': 0, 'total_cash_in_n1': 0,
        'total_cash_out_n': 0, 'total_cash_out_n1': 0,
        'total_cm1_value': 0, 'total_cm2_value': 0,
        'total_ec': 0, 'total_ic': 0,
        'total_committed': 0, 'total_as_sold': 0,
        'valid_projects': 0, 'weighted_poc_n': 0, 'weighted_poc_n1': 0,
        'total_cpi': 0, 'total_spi': 0, 'total_risk_score': 0,
        # Add contingency tracking
        'total_contingency_as_sold': 0,
        'total_contingency_fct_n': 0,
        'weighted_poc_for_contingency': 0,
        'projects_with_contingency': 0
    }
    
    for project_id, project in portfolio_data.items():
        data = project['data']
        if not data:
            continue
            
        # Basic financial metrics
        contract_n = safe_get_value(data, 'revenues', 'Contract Price', 'n_ptd')
        contract_n1 = safe_get_value(data, 'revenues', 'Contract Price', 'n1_ptd')
        revenue_n = safe_get_value(data, 'revenues', 'Revenues', 'n_ptd')
        revenue_n1 = safe_get_value(data, 'revenues', 'Revenues', 'n1_ptd')
        cash_in_n = safe_get_value(data, 'revenues', 'Cash IN', 'n_ptd')
        cash_in_n1 = safe_get_value(data, 'revenues', 'Cash IN', 'n1_ptd')
        cash_out_n = safe_get_value(data, 'revenues', 'Cash OUT', 'n_ptd')
        cash_out_n1 = safe_get_value(data, 'revenues', 'Cash OUT', 'n1_ptd')
        poc_n = safe_get_value(data, 'revenues', 'POC%', 'n_ptd')
        poc_n1 = safe_get_value(data, 'revenues', 'POC%', 'n1_ptd')
        
        # Enhanced cost metrics
        cost_analysis = data.get('cost_analysis', {})
        earned_value = data.get('earned_value', {})
        risk_factors = data.get('risk_factors', [])
        work_packages = data.get('work_packages', {})
        
        if contract_n > 0:
            metrics['total_contract_value_n'] += contract_n
            metrics['total_contract_value_n1'] += contract_n1
            metrics['total_revenue_n'] += revenue_n
            metrics['total_revenue_n1'] += revenue_n1
            metrics['total_cash_in_n'] += cash_in_n
            metrics['total_cash_in_n1'] += cash_in_n1
            metrics['total_cash_out_n'] += cash_out_n
            metrics['total_cash_out_n1'] += cash_out_n1
            metrics['weighted_poc_n'] += (poc_n * contract_n)
            metrics['weighted_poc_n1'] += (poc_n1 * contract_n1)
            
            # Enhanced cost and performance metrics
            metrics['total_cm1_value'] += cost_analysis.get('cm1_value_fct_n', 0)
            metrics['total_cm2_value'] += cost_analysis.get('cm2_value_fct_n', 0)
            metrics['total_ec'] += cost_analysis.get('ec_total_fct_n', 0)
            metrics['total_ic'] += cost_analysis.get('ic_total_fct_n', 0)
            metrics['total_committed'] += cost_analysis.get('total_committed', 0)
            metrics['total_as_sold'] += cost_analysis.get('total_as_sold', 0)
            
            # Performance indices
            metrics['total_cpi'] += earned_value.get('cost_performance_index', 1.0)
            metrics['total_spi'] += earned_value.get('schedule_performance_index', 1.0)
            
            # Risk assessment
            risk_score = len([r for r in risk_factors if r['severity'] in ['Critical', 'High']]) * 2 + \
                        len([r for r in risk_factors if r['severity'] == 'Medium'])
            metrics['total_risk_score'] += risk_score
            
            # Calculate contingency metrics
            contingency_metrics = calculate_contingency_metrics(work_packages, poc_n)
            if contingency_metrics['has_contingency']:
                metrics['total_contingency_as_sold'] += contingency_metrics['contingency_as_sold']
                metrics['total_contingency_fct_n'] += contingency_metrics['contingency_fct_n']
                metrics['weighted_poc_for_contingency'] += (poc_n * contract_n)
                metrics['projects_with_contingency'] += 1
            
            metrics['valid_projects'] += 1
    
    if metrics['valid_projects'] == 0:
        return None
    
    # Calculate portfolio contingency efficiency
    portfolio_contingency_efficiency = None
    if metrics['projects_with_contingency'] > 0 and metrics['total_contingency_as_sold'] > 0:
        portfolio_contingency_consumed = metrics['total_contingency_as_sold'] - metrics['total_contingency_fct_n']
        portfolio_contingency_consumed_pct = (portfolio_contingency_consumed / metrics['total_contingency_as_sold'] * 100)
        portfolio_avg_poc = (metrics['weighted_poc_for_contingency'] / metrics['total_contract_value_n']) if metrics['total_contract_value_n'] > 0 else 0
        
        if portfolio_avg_poc > 0:
            portfolio_contingency_efficiency = (2 - (portfolio_contingency_consumed_pct / portfolio_avg_poc)) * 100
            portfolio_contingency_efficiency = max(0, min(200, portfolio_contingency_efficiency))
    
    # Calculate comprehensive portfolio summary
    portfolio_summary = {
        # Financial metrics
        'total_contract_value': metrics['total_contract_value_n'],
        'total_revenue': metrics['total_revenue_n'],
        'total_cm1': metrics['total_cm1_value'],
        'total_cm2': metrics['total_cm2_value'],
        'net_cash_flow': metrics['total_cash_in_n'] - metrics['total_cash_out_n'],
        'project_count': metrics['valid_projects'],
        
        # Previous period comparisons
        'total_contract_value_n1': metrics['total_contract_value_n1'],
        'total_revenue_n1': metrics['total_revenue_n1'],
        'net_cash_flow_n1': metrics['total_cash_in_n1'] - metrics['total_cash_out_n1'],
        
        # Variance calculations
        'contract_variance': calculate_period_variance(metrics['total_contract_value_n'], metrics['total_contract_value_n1']),
        'revenue_variance': calculate_period_variance(metrics['total_revenue_n'], metrics['total_revenue_n1']),
        'net_cash_variance': calculate_period_variance(
            metrics['total_cash_in_n'] - metrics['total_cash_out_n'],
            metrics['total_cash_in_n1'] - metrics['total_cash_out_n1']
        ),
        
        # Enhanced margin metrics
        'avg_cm1_pct': (metrics['total_cm1_value'] / metrics['total_contract_value_n'] * 100) if metrics['total_contract_value_n'] > 0 else 0,
        'avg_cm2_pct': (metrics['total_cm2_value'] / metrics['total_contract_value_n'] * 100) if metrics['total_contract_value_n'] > 0 else 0,
        'portfolio_ec_pct': (metrics['total_ec'] / metrics['total_contract_value_n'] * 100) if metrics['total_contract_value_n'] > 0 else 0,
        'portfolio_ic_pct': (metrics['total_ic'] / metrics['total_contract_value_n'] * 100) if metrics['total_contract_value_n'] > 0 else 0,
        'portfolio_committed_ratio': (metrics['total_committed'] / metrics['total_as_sold']) if metrics['total_as_sold'] > 0 else 0,
        
        # Performance indices
        'avg_cost_performance_index': metrics['total_cpi'] / metrics['valid_projects'],
        'avg_schedule_performance_index': metrics['total_spi'] / metrics['valid_projects'],
        
        # POC metrics
        'weighted_poc_n': (metrics['weighted_poc_n'] / metrics['total_contract_value_n']) if metrics['total_contract_value_n'] > 0 else 0,
        'weighted_poc_n1': (metrics['weighted_poc_n1'] / metrics['total_contract_value_n1']) if metrics['total_contract_value_n1'] > 0 else 0,
        
        # Risk metrics
        'average_risk_score': metrics['total_risk_score'] / metrics['valid_projects'],
        'total_portfolio_value_at_risk': 0,  # Will be calculated separately
        
        # Cash flow efficiency
        'cash_flow_efficiency': (metrics['total_cash_in_n'] / metrics['total_cash_out_n']) if metrics['total_cash_out_n'] > 0 else 1.0,
        
        # Contingency metrics
        'portfolio_contingency_efficiency': portfolio_contingency_efficiency,
        'total_contingency_as_sold': metrics['total_contingency_as_sold'],
        'total_contingency_fct_n': metrics['total_contingency_fct_n'],
        'projects_with_contingency': metrics['projects_with_contingency']
    }
    
    # Calculate POC velocity using corrected calculation
    portfolio_summary['weighted_poc_velocity'] = calculate_poc_velocity(
        portfolio_summary['weighted_poc_n'], 
        portfolio_summary['weighted_poc_n1']
    )
    
    return portfolio_summary

# ================================================================================
# ENHANCED PROJECT ANALYSIS FUNCTIONS
# ================================================================================

def render_project_analysis_page():
    """Render simplified project analysis page - CLEAN VERSION"""
    st.header("🔍 Project Analysis")
    
    # Check for data
    if not st.session_state.projects_data:
        st.warning("📋 No project data uploaded. Please upload Excel files first.")
        return
    
    # Project selection
    project_options = {f"{k} - {v['name']}": k for k, v in st.session_state.projects_data.items()}
    selected_project = st.selectbox("Select Project for Analysis", list(project_options.keys()))
    
    if not selected_project:
        return
    
    # Get project data
    project_id = project_options[selected_project]
    project = st.session_state.projects_data[project_id]
    project_data = project['data']
    
    if not project_data:
        st.error("❌ No data available for selected project.")
        return
    
    # Render project analysis sections
    try:
        # Project header
        render_enhanced_project_header(project, project_id, project_data)
        
        st.markdown("---")
        # Project metrics
        render_comprehensive_project_metrics(project_data)
        
        st.markdown("---")
        # Financial analysis
        render_financial_performance_analysis(project_data)
        
        st.markdown("---")
        # Cost analysis (simplified)
        render_simplified_cost_structure_analysis(project_data)
        
        st.markdown("---")
        # Cash flow (quarterly only)
        render_simplified_cash_flow_timeline(project_data)
        
        st.markdown("---")
        # Earned value analysis
        render_earned_value_analysis(project_data)
        
        st.markdown("---")
        # Risk assessment - ONLY CALL ONCE
        render_project_risk_assessment(project_data)
        
        st.markdown("---")
        # Performance trends
        render_performance_trends_analysis(project_data)
        
    except Exception as e:
        st.error(f"❌ Error rendering project analysis: {str(e)}")
        st.exception(e)
        st.info("Some analysis sections may not be available. Please check your project data.")

def render_enhanced_project_header(project, project_id, project_data):
    """Render enhanced project header with status indicators"""
    col1, col2, col3, col4 = st.columns([3, 1, 1, 1])
    
    with col1:
        st.subheader(f"📊 {project['name']}")
        st.caption(f"Project ID: {project_id}")
        
        # Project info details
        project_info = project_data.get('project_info', {})
        if project_info:
            st.write(f"**Manager:** {project_info.get('Project Manager', 'N/A')}")
            st.write(f"**Phase:** {project_info.get('Project Phase', 'N/A')}")
    
    with col2:
        status = project_data['project_info'].get('Project Status', 'Unknown')
        status_color = "success" if status == "Active" else "warning" if status == "On Hold" else "error"
        st.markdown(f"**Status:** :{status_color}[{status}]")
        
        quarter = project_data['project_info'].get('Reporting Quarter', 'N/A')
        st.write(f"**Quarter:** {quarter}")
    
    with col3:
        template_version = project.get('template_version', 'Unknown')
        st.info(f"**Template:** {template_version}")
        
        upload_time = project.get('upload_time', datetime.datetime.now())
        st.caption(f"**Uploaded:** {upload_time.strftime('%Y-%m-%d %H:%M')}")
    
    with col4:
        # Overall project health indicator
        risk_factors = project_data.get('risk_factors', [])
        critical_risks = len([r for r in risk_factors if r['severity'] == 'Critical'])
        high_risks = len([r for r in risk_factors if r['severity'] == 'High'])
        
        if critical_risks > 0:
            health_status = "🔴 Critical"
            health_color = "error"
        elif high_risks > 2:
            health_status = "🟠 High Risk" 
            health_color = "warning"
        elif high_risks > 0:
            health_status = "🟡 Medium Risk"
            health_color = "warning"
        else:
            health_status = "🟢 Healthy"
            health_color = "success"
        
        st.markdown(f"**Health:** :{health_color}[{health_status}]")
        st.write(f"**Risks:** {critical_risks}C / {high_risks}H")

def render_comprehensive_project_metrics(project_data):
    """Render comprehensive project KPI dashboard with improved visual hierarchy"""
    st.subheader("📈 Project Performance Dashboard")
    
    # Add custom CSS for better styling
    st.markdown("""
    <style>
    .metric-card {
        background-color: #f8f9fa;
        padding: 1rem;
        border-radius: 10px;
        box-shadow: 0 2px 4px rgba(0,0,0,0.1);
        border-left: 4px solid #1f77b4;
        height: 100%;
    }
    .metric-card-excellent {
        border-left-color: #28a745;
        background: linear-gradient(135deg, #f8f9fa 0%, #e8f5e9 100%);
    }
    .metric-card-good {
        border-left-color: #17a2b8;
        background: linear-gradient(135deg, #f8f9fa 0%, #e3f2fd 100%);
    }
    .metric-card-warning {
        border-left-color: #ffc107;
        background: linear-gradient(135deg, #f8f9fa 0%, #fff3e0 100%);
    }
    .metric-card-critical {
        border-left-color: #dc3545;
        background: linear-gradient(135deg, #f8f9fa 0%, #ffebee 100%);
    }
    .primary-metric {
        font-size: 2rem;
        font-weight: bold;
        color: #1f4e79;
        margin: 0.5rem 0;
    }
    .metric-label {
        font-size: 0.9rem;
        color: #666;
        font-weight: 600;
        text-transform: uppercase;
        letter-spacing: 0.5px;
    }
    .metric-trend {
        font-size: 1.1rem;
        font-weight: 600;
        margin-top: 0.5rem;
    }
    .section-header {
        background: linear-gradient(90deg, #1f4e79 0%, #2d5aa0 100%);
        color: white;
        padding: 0.5rem 1rem;
        border-radius: 5px;
        margin: 1rem 0 0.5rem 0;
        font-weight: 600;
    }
    </style>
    """, unsafe_allow_html=True)
    
    # Financial metrics
    contract_val_n = safe_get_value(project_data, 'revenues', 'Contract Price', 'n_ptd')
    contract_val_n1 = safe_get_value(project_data, 'revenues', 'Contract Price', 'n1_ptd')
    revenue_n = safe_get_value(project_data, 'revenues', 'Revenues', 'n_ptd')
    revenue_n1 = safe_get_value(project_data, 'revenues', 'Revenues', 'n1_ptd')
    poc_n = safe_get_value(project_data, 'revenues', 'POC%', 'n_ptd')
    poc_n1 = safe_get_value(project_data, 'revenues', 'POC%', 'n1_ptd')
    
    # Enhanced metrics
    cost_analysis = project_data.get('cost_analysis', {})
    earned_value = project_data.get('earned_value', {})
    work_packages = project_data.get('work_packages', {})
    
    # Calculate contingency metrics
    contingency_metrics = calculate_contingency_metrics(work_packages, poc_n)
    
    # PRIMARY METRICS - Top Row (Most Important)
    st.markdown('<div class="section-header">🎯 Primary Performance Indicators</div>', unsafe_allow_html=True)
    
    col1, col2, col3, col4 = st.columns(4)
    
    with col1:
        contract_var = calculate_period_variance(contract_val_n, contract_val_n1)
        var_color = "success" if contract_var >= 0 else "error"
        st.markdown(f"""
        <div class="metric-card metric-card-{'excellent' if contract_var >= 0 else 'warning'}">
            <div class="metric-label">Contract Value</div>
            <div class="primary-metric">{format_currency_millions(contract_val_n)}</div>
            <div class="metric-trend" style="color: {'#28a745' if contract_var >= 0 else '#dc3545'}">
                {'📈' if contract_var >= 0 else '📉'} {contract_var:+.1f}%
            </div>
        </div>
        """, unsafe_allow_html=True)
    
    with col2:
        cm2_pct = cost_analysis.get('cm2_pct_fct_n', 0)
        cm2_icon, cm2_status, cm2_class = get_traffic_light_status(cm2_pct, EXECUTIVE_THRESHOLDS['cm2_margin'])
        
        cm2_excellent = EXECUTIVE_THRESHOLDS['cm2_margin']['excellent']
        cm2_good = EXECUTIVE_THRESHOLDS['cm2_margin']['good']
        cm2_warning = EXECUTIVE_THRESHOLDS['cm2_margin']['warning']
        
        if cm2_pct >= cm2_excellent:
            target_text = f"Excellent (≥{cm2_excellent}%)"
        elif cm2_pct >= cm2_good:
            target_text = f"Target: {cm2_excellent}%"
        elif cm2_pct >= cm2_warning:
            target_text = f"Target: {cm2_good}%"
        else:
            target_text = f"Min Required: {cm2_warning}%"
        
        st.markdown(f"""
        <div class="metric-card metric-card-{cm2_class}">
            <div class="metric-label">CM2 Margin</div>
            <div class="primary-metric">{cm2_pct:.1f}%</div>
            <div class="metric-trend">{cm2_icon} {cm2_status}</div>
            <div style="font-size: 0.8rem; color: #666; margin-top: 0.2rem;">
                {target_text}
            </div>
        </div>
        """, unsafe_allow_html=True)
    
    with col3:
        poc_velocity = calculate_poc_velocity(poc_n, poc_n1)
        
        poc_icon_raw, poc_status_raw, poc_class_raw = get_traffic_light_status(
            poc_velocity, EXECUTIVE_THRESHOLDS['poc_velocity']
        )
        poc_icon_adjusted, poc_status_adjusted, poc_class_adjusted = get_poc_velocity_status_with_maturity(
            poc_velocity, poc_n
        )
        
        st.markdown(f"""
        <div class="metric-card metric-card-{poc_class_adjusted}">
            <div class="metric-label">POC Progress</div>
            <div class="primary-metric">{poc_n:.1f}%</div>
            <div class="metric-trend">{poc_icon_adjusted} {poc_velocity:+.1f}% /month</div>
            <div style="font-size: 0.8rem; color: #666; margin-top: 0.2rem;">
                Expected: ~{calculate_expected_poc_velocity(poc_n):.1f}% /month
            </div>
        </div>
        """, unsafe_allow_html=True)
    
    with col4:
        # Contingency Efficiency
        if contingency_metrics['has_contingency']:
            st.markdown(f"""
            <div class="metric-card metric-card-{contingency_metrics['status_color']}">
                <div class="metric-label">Contingency Efficiency</div>
                <div class="primary-metric">{contingency_metrics['efficiency']:.0f}%</div>
                <div class="metric-trend">{contingency_metrics['status_icon']} {contingency_metrics['status']}</div>
                <div style="font-size: 0.8rem; color: #666; margin-top: 0.2rem;">
                    {contingency_metrics['consumed_percentage']:.1f}% consumed
                </div>
            </div>
            """, unsafe_allow_html=True)
        else:
            st.markdown(f"""
            <div class="metric-card">
                <div class="metric-label">Contingency Efficiency</div>
                <div class="primary-metric">N/A</div>
                <div class="metric-trend">➖ No Contingency</div>
                <div style="font-size: 0.8rem; color: #666; margin-top: 0.2rem;">
                    No risk budget allocated
                </div>
            </div>
            """, unsafe_allow_html=True)
    
    # PERFORMANCE METRICS - Second Row
    st.markdown('<div class="section-header">📊 Performance Metrics</div>', unsafe_allow_html=True)
    
    col5, col6, col7, col8 = st.columns(4)
    
    with col5:
        revenue_var = calculate_period_variance(revenue_n, revenue_n1)
        st.markdown(f"""
        <div class="metric-card">
            <div class="metric-label">Revenue PTD</div>
            <div class="primary-metric" style="font-size: 1.5rem;">{format_currency_millions(revenue_n)}</div>
            <div class="metric-trend" style="color: {'#28a745' if revenue_var >= 0 else '#dc3545'}">
                {revenue_var:+.1f}% vs previous
            </div>
        </div>
        """, unsafe_allow_html=True)
    
    with col6:
        cpi = earned_value.get('cost_performance_index', 1.0)
        cpi_icon, _, cpi_class = get_traffic_light_status(cpi, EXECUTIVE_THRESHOLDS['cost_performance_index'])
        st.markdown(f"""
        <div class="metric-card metric-card-{cpi_class}">
            <div class="metric-label">Cost Performance (CPI)</div>
            <div class="primary-metric" style="font-size: 1.5rem;">{cpi:.2f}</div>
            <div class="metric-trend">{cpi_icon} {'Efficient' if cpi >= 1.0 else 'Over Budget'}</div>
        </div>
        """, unsafe_allow_html=True)
    
    with col7:
        spi = earned_value.get('schedule_performance_index', 1.0)
        spi_icon, _, spi_class = get_traffic_light_status(spi, EXECUTIVE_THRESHOLDS['schedule_performance_index'])
        st.markdown(f"""
        <div class="metric-card metric-card-{spi_class}">
            <div class="metric-label">Schedule Performance (SPI)</div>
            <div class="primary-metric" style="font-size: 1.5rem;">{spi:.2f}</div>
            <div class="metric-trend">{spi_icon} {'On Time' if spi >= 1.0 else 'Behind Schedule'}</div>
        </div>
        """, unsafe_allow_html=True)
    
    with col8:
        cash_in_pct = safe_get_value(project_data, 'revenues', 'Cash In %', 'n_ptd')
        cash_icon = "🟢" if cash_in_pct >= poc_n else "🟡" if cash_in_pct >= poc_n * 0.8 else "🔴"
        st.markdown(f"""
        <div class="metric-card">
            <div class="metric-label">Cash Collection</div>
            <div class="primary-metric" style="font-size: 1.5rem;">{cash_in_pct:.1f}%</div>
            <div class="metric-trend">{cash_icon} vs POC: {poc_n:.1f}%</div>
        </div>
        """, unsafe_allow_html=True)
    
    # Add spacing
    st.markdown("<br>", unsafe_allow_html=True)
    
    # OPERATIONAL METRICS - Third Row
    st.markdown('<div class="section-header">⚙️ Operational Metrics</div>', unsafe_allow_html=True)
    
    col9, col10, col11, col12 = st.columns(4)
    
    with col9:
        committed_ratio = cost_analysis.get('committed_ratio', 0)
        committed_icon, _, committed_class = get_traffic_light_status(committed_ratio, EXECUTIVE_THRESHOLDS['committed_vs_budget'], reverse=True)
        st.markdown(f"""
        <div class="metric-card metric-card-{committed_class}">
            <div class="metric-label">Committed Ratio</div>
            <div class="primary-metric" style="font-size: 1.5rem;">{committed_ratio:.2f}</div>
            <div class="metric-trend">{committed_icon} {'Good' if committed_ratio <= 1.0 else 'Overcommitted'}</div>
        </div>
        """, unsafe_allow_html=True)
    
    with col10:
        cost_variance_pct = cost_analysis.get('cost_variance_pct', 0)
        cost_var_icon, _, cost_var_class = get_traffic_light_status(abs(cost_variance_pct), EXECUTIVE_THRESHOLDS['cost_variance'], reverse=True)
        st.markdown(f"""
        <div class="metric-card metric-card-{cost_var_class}">
            <div class="metric-label">Cost Variance</div>
            <div class="primary-metric" style="font-size: 1.5rem;">{cost_variance_pct:+.1f}%</div>
            <div class="metric-trend">{cost_var_icon} from baseline</div>
        </div>
        """, unsafe_allow_html=True)
    
    with col11:
        commodity_wps = [
            wp for wp in work_packages.values() 
            if not ('risk' in wp.get('description', '').lower() and 'contingenc' in wp.get('description', '').lower())
        ]
        
        wp_at_risk = len([wp for wp in commodity_wps if wp.get('variance_pct', 0) > 15])
        risk_contingencies = [
            wp for wp in work_packages.values() 
            if 'risk' in wp.get('description', '').lower() and 'contingenc' in wp.get('description', '').lower()
        ]
        wp_opportunities = len([wp for wp in risk_contingencies if wp.get('variance_pct', 0) > 0])
        
        total_wp_display = len(commodity_wps)
        wp_icon = "🟢" if wp_at_risk == 0 else "🟡" if wp_at_risk <= 2 else "🔴"
        
        trend_text = f"{wp_icon} {wp_at_risk} at risk"
        if wp_opportunities > 0:
            trend_text += f" | 💚 {wp_opportunities} opportunities"
        
        st.markdown(f"""
        <div class="metric-card">
            <div class="metric-label">Work Packages</div>
            <div class="primary-metric" style="font-size: 1.5rem;">{total_wp_display}</div>
            <div class="metric-trend">{trend_text}</div>
        </div>
        """, unsafe_allow_html=True)
    
    with col12:
        # Time to completion estimate
        if poc_velocity > 0:
            months_to_complete = (100 - poc_n) / poc_velocity
            completion_icon = "🟢" if months_to_complete <= 12 else "🟡" if months_to_complete <= 24 else "🔴"
            completion_text = f"{months_to_complete:.0f} months"
        else:
            completion_icon = "🔴"
            completion_text = "Unknown"
        
        st.markdown(f"""
        <div class="metric-card">
            <div class="metric-label">Est. Time to Complete</div>
            <div class="primary-metric" style="font-size: 1.5rem;">{completion_text}</div>
            <div class="metric-trend">{completion_icon} at current velocity</div>
        </div>
        """, unsafe_allow_html=True)
    
    # CONTINGENCY DETAILS - Fourth Row (if applicable)
    if contingency_metrics['has_contingency']:
        st.markdown('<div class="section-header">💰 Contingency Management Details</div>', unsafe_allow_html=True)
        
        col13, col14, col15, col16 = st.columns(4)
        
        with col13:
            st.metric(
                "Original Contingency",
                format_currency_thousands(contingency_metrics['contingency_as_sold']),
                "Baseline allocation"
            )
        
        with col14:
            st.metric(
                "Remaining Contingency",
                format_currency_thousands(contingency_metrics['remaining_amount']),
                f"{contingency_metrics['remaining_percentage']:.1f}% of original"
            )
        
        with col15:
            st.metric(
                "Consumption Trend",
                contingency_metrics['trend'],
                f"{contingency_metrics['trend_icon']} vs previous period"
            )
        
        with col16:
            # Project when contingency will be depleted
            if contingency_metrics['recent_consumption'] > 0 and poc_velocity > 0:
                months_to_depletion = contingency_metrics['remaining_amount'] / (contingency_metrics['recent_consumption'] / 1)  # Assuming monthly
                depletion_poc = poc_n + (poc_velocity * months_to_depletion)
                
                if depletion_poc < 100:
                    st.metric(
                        "Depletion Risk",
                        f"@ {depletion_poc:.0f}% POC",
                        f"⚠️ In {months_to_depletion:.0f} months"
                    )
                else:
                    st.metric(
                        "Depletion Risk",
                        "Low",
                        "✅ Sufficient to completion"
                    )
            else:
                st.metric(
                    "Depletion Risk",
                    "N/A",
                    "➖ No recent consumption"
                )
    
    # Visual Performance Summary Chart (existing code continues...)
    with st.expander("📊 Performance Trend Visualization", expanded=False):
        fig = go.Figure()
        
        # Create a radar chart for quick performance overview
        categories = ['Cost Efficiency', 'Schedule', 'Margin', 'Cash Flow', 'Progress', 'Contingency']
        values = [
            cpi * 100,
            spi * 100,
            min(cm2_pct * 5, 100),  # Scale CM2% to 0-100
            min(cash_in_pct / poc_n * 100, 100) if poc_n > 0 else 100,
            poc_n,
            contingency_metrics['efficiency'] if contingency_metrics['has_contingency'] else 100
        ]
        
        fig.add_trace(go.Scatterpolar(
            r=values,
            theta=categories,
            fill='toself',
            name='Current Performance',
            fillcolor='rgba(31, 119, 180, 0.5)',
            line=dict(color='rgb(31, 119, 180)', width=2)
        ))
        
        # Add target line
        fig.add_trace(go.Scatterpolar(
            r=[100, 100, 100, 100, 100, 100],
            theta=categories,
            name='Target',
            line=dict(color='green', width=1, dash='dash')
        ))
        
        fig.update_layout(
            polar=dict(
                radialaxis=dict(
                    visible=True,
                    range=[0, 120]
                )),
            showlegend=True,
            title="Performance Radar",
            height=400
        )
        
        st.plotly_chart(fig, use_container_width=True)

def render_financial_performance_analysis(project_data):
    """Render detailed financial performance analysis"""
    st.subheader("💰 Financial Performance Analysis")
    
    # Revenue trend analysis
    col1, col2 = st.columns(2)
    
    with col1:
        st.markdown("#### 📊 Revenue Performance")
        
        revenue_data = []
        for period in ['n1_ptd', 'n_ptd', 'n_mtd']:
            period_name = {'n1_ptd': 'Previous Period', 'n_ptd': 'Current Period', 'n_mtd': 'Month to Date'}[period]
            
            contract = safe_get_value(project_data, 'revenues', 'Contract Price', period)
            revenue = safe_get_value(project_data, 'revenues', 'Revenues', period)
            poc = safe_get_value(project_data, 'revenues', 'POC%', period)
            cash_in = safe_get_value(project_data, 'revenues', 'Cash IN', period)
            
            revenue_data.append({
                'Period': period_name,
                'Contract': contract/1000,
                'Revenue': revenue/1000,
                'POC%': poc,
                'Cash IN': cash_in/1000
            })
        
        df_revenue = pd.DataFrame(revenue_data)
        
        # Revenue trend chart
        fig = go.Figure()
        
        fig.add_trace(go.Bar(
            name='Contract Value',
            x=df_revenue['Period'],
            y=df_revenue['Contract'],
            marker_color='lightblue',
            opacity=0.7
        ))
        
        fig.add_trace(go.Bar(
            name='Revenue',
            x=df_revenue['Period'],
            y=df_revenue['Revenue'],
            marker_color='darkblue'
        ))
        
        fig.add_trace(go.Scatter(
            name='POC %',
            x=df_revenue['Period'],
            y=df_revenue['POC%'],
            mode='lines+markers',
            yaxis='y2',
            line=dict(color='red', width=3)
        ))
        
        fig.update_layout(
            title='Revenue and POC Trend',
            yaxis=dict(title='Value (CHF Thousands)'),
            yaxis2=dict(title='POC %', overlaying='y', side='right'),
            height=400
        )
        
        st.plotly_chart(fig, use_container_width=True)
    
    with col2:
        st.markdown("#### 💸 Cash Flow Performance")
        
        cash_data = []
        for period in ['n1_ptd', 'n_ptd', 'n_mtd']:
            period_name = {'n1_ptd': 'Previous', 'n_ptd': 'Current', 'n_mtd': 'MTD'}[period]
            
            cash_in = safe_get_value(project_data, 'revenues', 'Cash IN', period)
            cash_out = safe_get_value(project_data, 'revenues', 'Cash OUT', period)
            net_cash = cash_in - cash_out
            
            cash_data.append({
                'Period': period_name,
                'Cash IN': cash_in/1000,
                'Cash OUT': cash_out/1000,
                'Net Cash': net_cash/1000
            })
        
        df_cash = pd.DataFrame(cash_data)
        
        # Cash flow chart
        fig = go.Figure()
        
        fig.add_trace(go.Bar(
            name='Cash IN',
            x=df_cash['Period'],
            y=df_cash['Cash IN'],
            marker_color='green',
            opacity=0.7
        ))
        
        fig.add_trace(go.Bar(
            name='Cash OUT',
            x=df_cash['Period'],
            y=df_cash['Cash OUT'],
            marker_color='red',
            opacity=0.7
        ))
        
        fig.add_trace(go.Scatter(
            name='Net Cash Flow',
            x=df_cash['Period'],
            y=df_cash['Net Cash'],
            mode='lines+markers',
            line=dict(color='blue', width=3),
            marker=dict(size=10)
        ))
        
        fig.update_layout(
            title='Cash Flow Analysis',
            yaxis_title='Cash Flow (CHF Thousands)',
            height=400
        )
        
        st.plotly_chart(fig, use_container_width=True)
    
    # Financial summary table
    st.markdown("#### 📋 Financial Summary")
    
    financial_summary = []
    for period in ['Previous Period', 'Current Period', 'Month to Date']:
        period_key = {'Previous Period': 'n1_ptd', 'Current Period': 'n_ptd', 'Month to Date': 'n_mtd'}[period]
        
        contract = safe_get_value(project_data, 'revenues', 'Contract Price', period_key)
        revenue = safe_get_value(project_data, 'revenues', 'Revenues', period_key)
        poc = safe_get_value(project_data, 'revenues', 'POC%', period_key)
        cash_in = safe_get_value(project_data, 'revenues', 'Cash IN', period_key)
        cash_out = safe_get_value(project_data, 'revenues', 'Cash OUT', period_key)
        cash_in_pct = safe_get_value(project_data, 'revenues', 'Cash In %', period_key)
        
        financial_summary.append({
            'Period': period,
            'Contract Value': format_currency_thousands(contract),
            'Revenue': format_currency_thousands(revenue),
            'POC %': f"{poc:.1f}%",
            'Cash IN': format_currency_thousands(cash_in),
            'Cash OUT': format_currency_thousands(cash_out),
            'Net Cash': format_currency_thousands(cash_in - cash_out),
            'Collection %': f"{cash_in_pct:.1f}%"
        })
    
    df_financial = pd.DataFrame(financial_summary)
    st.dataframe(df_financial, use_container_width=True)

def render_simplified_cost_structure_analysis(project_data):
    """Render simplified cost structure analysis focusing on available data"""
    st.subheader("🏗️ Cost Structure Analysis")
    
    cost_analysis = project_data.get('cost_analysis', {})
    
    if not cost_analysis:
        st.warning("📊 No cost analysis data available.")
        return
    
    # Cost structure overview
    col1, col2, col3 = st.columns(3)
    
    with col1:
        # Total cost breakdown
        st.markdown("#### 💼 Cost Summary")
        
        total_as_sold = cost_analysis.get('total_as_sold', 0)
        total_committed = cost_analysis.get('total_committed', 0)
        total_fct_n = cost_analysis.get('total_fct_n', 0)
        
        st.metric("Total As Sold", format_currency_thousands(total_as_sold))
        st.metric("Total Committed", format_currency_thousands(total_committed))
        st.metric("Total FCT (n)", format_currency_thousands(total_fct_n))
    
    with col2:
        # Margin analysis
        st.markdown("#### 📊 Margin Performance")
        
        contract_value = safe_get_value(project_data, 'revenues', 'Contract Price', 'n_ptd')
        cm1_pct = cost_analysis.get('cm1_pct_fct_n', 0)
        cm2_pct = cost_analysis.get('cm2_pct_fct_n', 0)
        
        cm1_icon, _, _ = get_traffic_light_status(cm1_pct, EXECUTIVE_THRESHOLDS['cm1_margin'])
        cm2_icon, _, _ = get_traffic_light_status(cm2_pct, EXECUTIVE_THRESHOLDS['cm2_margin'])
        
        st.metric("CM1 Margin", f"{cm1_pct:.1f}%", f"{cm1_icon}")
        st.metric("CM2 Margin", f"{cm2_pct:.1f}%", f"{cm2_icon}")
    
    with col3:
        # Performance metrics
        st.markdown("#### 🎯 Performance")
        
        committed_ratio = cost_analysis.get('committed_ratio', 0)
        cost_variance_pct = cost_analysis.get('cost_variance_pct', 0)
        
        committed_icon, _, _ = get_traffic_light_status(committed_ratio, EXECUTIVE_THRESHOLDS['committed_vs_budget'], reverse=True)
        variance_icon = "🟢" if abs(cost_variance_pct) <= 10 else "🟡" if abs(cost_variance_pct) <= 20 else "🔴"
        
        st.metric("Committed Ratio", f"{committed_ratio:.2f}", f"{committed_icon}")
        st.metric("Cost Variance", f"{cost_variance_pct:+.1f}%", f"{variance_icon}")
    
    # Cost structure visualization
    if contract_value > 0 and total_fct_n > 0:
        st.markdown("#### 📊 Cost Structure Breakdown")
        
        cost_breakdown = {
            'CM2 (Profit)': max(0, contract_value - total_fct_n)/1000,
            'Total Costs': total_fct_n/1000
        }
        
        fig = go.Figure(data=[
            go.Pie(
                labels=list(cost_breakdown.keys()),
                values=list(cost_breakdown.values()),
                hole=0.4,
                marker_colors=['darkgreen', 'red']
            )
        ])
        
        fig.update_layout(
            title=f'Cost Structure (CM2: {cm2_pct:.1f}%)',
            height=400
        )
        
        st.plotly_chart(fig, use_container_width=True)

def render_simplified_cash_flow_timeline(project_data):
    """Render simplified quarterly cash flow analysis with practical KPIs and insights"""
    st.subheader("💰 Quarterly Cash Flow Analysis & Insights")
    
    quarterly_data = project_data.get('cash_flow_quarterly', {})
    
    if not quarterly_data:
        st.warning("📊 No quarterly cash flow data available.")
        return
    
    # Process quarterly data
    quarters = sorted(quarterly_data.keys())
    quarter_metrics = []
    
    total_as_sold = 0
    total_fct_n = 0
    positive_quarters = 0
    negative_quarters = 0
    
    for quarter in quarters:
        data = quarterly_data[quarter]
        as_sold = data.get('as_sold', 0)
        fct_n1 = data.get('fct_n1', 0)
        fct_n = data.get('fct_n', 0)
        
        # For PADOVA project, handle the case where as_sold can be negative
        if as_sold != 0:
            variance_vs_plan = ((fct_n - as_sold) / abs(as_sold)) * 100
        else:
            variance_vs_plan = 0
            
        if fct_n1 != 0:
            variance_vs_previous = ((fct_n - fct_n1) / abs(fct_n1)) * 100
        else:
            variance_vs_previous = 0
        
        # Determine quarter status - adjust logic for negative cash flows
        if as_sold < 0:  # For negative expected cash flows (outflows)
            if fct_n >= as_sold:  # If actual outflow is less than or equal to planned
                status = "🟢 On/Better than Plan"
            elif variance_vs_plan >= -10:
                status = "🟡 Slight Variance"
            elif variance_vs_plan >= -20:
                status = "🟠 Concerning"
            else:
                status = "🔴 Critical"
        else:  # For positive expected cash flows (inflows)
            if fct_n >= as_sold:
                status = "🟢 On/Above Plan"
            elif variance_vs_plan >= -10:
                status = "🟡 Slight Variance"
            elif variance_vs_plan >= -20:
                status = "🟠 Concerning"
            else:
                status = "🔴 Critical"
        
        quarter_metrics.append({
            'quarter': quarter,
            'as_sold': as_sold,
            'fct_n1': fct_n1,
            'fct_n': fct_n,
            'variance_vs_plan': variance_vs_plan,
            'variance_vs_previous': variance_vs_previous,
            'status': status,
            'cumulative_fct': total_fct_n + fct_n
        })
        
        total_as_sold += as_sold
        total_fct_n += fct_n
        
        if fct_n > 0:
            positive_quarters += 1
        elif fct_n < 0:
            negative_quarters += 1
    
    # Key Performance Indicators
    st.markdown("### 🎯 Cash Flow KPIs")
    
    col1, col2, col3, col4, col5 = st.columns(5)
    
    with col1:
        # Handle the case where both totals might be negative
        if total_as_sold != 0:
            overall_variance = ((total_fct_n - total_as_sold) / abs(total_as_sold)) * 100
        else:
            overall_variance = 0
        variance_icon = "🟢" if abs(overall_variance) <= 5 else "🟡" if abs(overall_variance) <= 15 else "🔴"
        st.metric("Overall Performance", f"{overall_variance:+.1f}%", f"{variance_icon}")
    
    with col2:
        st.metric("Total Planned", format_currency_millions(total_as_sold))
    
    with col3:
        st.metric("Total Forecast", format_currency_millions(total_fct_n))
    
    with col4:
        total_quarters = len(quarters)
        if total_quarters > 0:
            cash_flow_consistency = (positive_quarters / total_quarters * 100)
        else:
            cash_flow_consistency = 0
        consistency_icon = "🟢" if cash_flow_consistency >= 80 else "🟡" if cash_flow_consistency >= 60 else "🔴"
        st.metric("Positive Cash Flow Quarters", f"{cash_flow_consistency:.0f}%", f"{consistency_icon}")
    
    with col5:
        # Calculate cash flow trend
        if len(quarter_metrics) >= 2:
            early_quarters = quarter_metrics[:len(quarter_metrics)//2]
            late_quarters = quarter_metrics[len(quarter_metrics)//2:]
            
            early_avg = np.mean([q['fct_n'] for q in early_quarters])
            late_avg = np.mean([q['fct_n'] for q in late_quarters])
            
            if early_avg != 0:
                trend = ((late_avg - early_avg) / abs(early_avg)) * 100
            else:
                trend = 100 if late_avg > 0 else -100 if late_avg < 0 else 0
            trend_icon = "📈" if trend > 5 else "📊" if trend > -5 else "📉"
        else:
            trend = 0
            trend_icon = "📊"
        
        st.metric("Cash Flow Trend", f"{trend:+.1f}%", f"{trend_icon}")
    
    # Quarterly Cash Flow Visualization
    st.markdown("### 📊 Quarterly Cash Flow Performance")
    
    fig = make_subplots(
        rows=2, cols=1,
        subplot_titles=('Cash Flow by Quarter', 'Variance Analysis'),
        vertical_spacing=0.1,
        specs=[[{"secondary_y": True}], [{"secondary_y": False}]]
    )
    
    # Main cash flow chart
    fig.add_trace(go.Bar(
        name='As Sold Plan',
        x=[q['quarter'] for q in quarter_metrics],
        y=[q['as_sold']/1000 for q in quarter_metrics],
        marker_color='lightblue',
        opacity=0.7
    ), row=1, col=1)
    
    fig.add_trace(go.Bar(
        name='FCT (n)',
        x=[q['quarter'] for q in quarter_metrics],
        y=[q['fct_n']/1000 for q in quarter_metrics],
        marker_color=['green' if abs(q['variance_vs_plan']) <= 10 else 'orange' if abs(q['variance_vs_plan']) <= 20 else 'red' for q in quarter_metrics],
        opacity=0.8
    ), row=1, col=1)
    
    # Add cumulative line
    fig.add_trace(go.Scatter(
        name='Cumulative FCT',
        x=[q['quarter'] for q in quarter_metrics],
        y=[sum([qm['fct_n'] for qm in quarter_metrics[:i+1]])/1000 for i in range(len(quarter_metrics))],
        mode='lines+markers',
        line=dict(color='purple', width=3),
        yaxis='y2'
    ), row=1, col=1, secondary_y=True)
    
    # Variance chart
    fig.add_trace(go.Bar(
        name='Variance vs Plan',
        x=[q['quarter'] for q in quarter_metrics],
        y=[q['variance_vs_plan'] for q in quarter_metrics],
        marker_color=['green' if abs(v) <= 10 else 'orange' if abs(v) <= 20 else 'red' for v in [q['variance_vs_plan'] for q in quarter_metrics]],
        showlegend=False
    ), row=2, col=1)
    
    # Add threshold lines
    fig.add_hline(y=0, line_dash="dash", line_color="black", row=2, col=1)
    fig.add_hline(y=-10, line_dash="dot", line_color="orange", row=2, col=1)
    fig.add_hline(y=-20, line_dash="dot", line_color="red", row=2, col=1)
    fig.add_hline(y=10, line_dash="dot", line_color="orange", row=2, col=1)
    fig.add_hline(y=20, line_dash="dot", line_color="red", row=2, col=1)
    
    fig.update_layout(
        height=700,
        title_text="Quarterly Cash Flow Analysis"
    )
    
    # Update axes
    fig.update_xaxes(title_text="Quarter", row=1, col=1)
    fig.update_yaxes(title_text="Cash Flow (CHF Thousands)", row=1, col=1)
    fig.update_yaxes(title_text="Cumulative (CHF Thousands)", secondary_y=True, row=1, col=1)
    fig.update_xaxes(title_text="Quarter", row=2, col=1)
    fig.update_yaxes(title_text="Variance %", row=2, col=1)
    
    st.plotly_chart(fig, use_container_width=True)
    
    # Quarterly Performance Table
    st.markdown("### 📋 Quarterly Performance Summary")
    
    quarterly_summary = []
    for q in quarter_metrics:
        quarterly_summary.append({
            'Quarter': q['quarter'],
            'Plan (As Sold)': format_currency_millions(q['as_sold']),
            'Forecast (FCT)': format_currency_millions(q['fct_n']),
            'vs Plan': f"{q['variance_vs_plan']:+.1f}%",
            'vs Previous FCT': f"{q['variance_vs_previous']:+.1f}%",
            'Status': q['status']
        })
    
    df_quarterly = pd.DataFrame(quarterly_summary)
    st.dataframe(df_quarterly, use_container_width=True)
    
    # Fixed Cash Flow Insights & Recommendations
    st.markdown("### 💡 Cash Flow Insights & Strategic Recommendations")
    
    col1, col2 = st.columns(2)
    
    with col1:
        # Performance insights
        excellent_quarters = len([q for q in quarter_metrics if "🟢" in q['status']])
        critical_quarters = len([q for q in quarter_metrics if "🔴" in q['status']])
        total_quarters_count = len(quarters)
        
        # Calculate success rate safely
        if total_quarters_count > 0:
            success_rate = ((total_quarters_count - critical_quarters) / total_quarters_count * 100)
        else:
            success_rate = 0
            
        # Calculate efficiency - handle negative values
        if total_as_sold != 0:
            efficiency = (total_fct_n / total_as_sold * 100)
        else:
            efficiency = 100 if total_fct_n == 0 else 0
            
        # Calculate forecast accuracy
        forecast_accuracy = (100 - abs(overall_variance))
        
        st.markdown(f"""
        <div class="exec-summary">
            <h4>📊 Performance Analysis</h4>
            <ul>
                <li><strong>Total Quarters:</strong> {total_quarters_count}</li>
                <li><strong>Excellent Performance:</strong> {excellent_quarters} quarters</li>
                <li><strong>Critical Performance:</strong> {critical_quarters} quarters</li>
                <li><strong>Success Rate:</strong> {success_rate:.0f}%</li>
                <li><strong>Cash Flow Efficiency:</strong> {efficiency:.1f}% of plan</li>
                <li><strong>Forecast Accuracy:</strong> {forecast_accuracy:.1f}%</li>
                <li><strong>Net Position:</strong> {format_currency_millions(total_fct_n)}</li>
            </ul>
        </div>
        """, unsafe_allow_html=True)
    
    with col2:
        # Strategic recommendations based on PADOVA's specific cash flow pattern
        recommendations = []
        
        # Check if this is an investment phase (negative early cash flows, positive later)
        early_cash_flow = sum([q['fct_n'] for q in quarter_metrics[:len(quarter_metrics)//2]])
        late_cash_flow = sum([q['fct_n'] for q in quarter_metrics[len(quarter_metrics)//2:]])
        
        if early_cash_flow < 0 and late_cash_flow > 0:
            recommendations.append("💼 **Investment Phase Project:** Early outflows followed by returns")
            recommendations.append("📊 **Monitor Payback:** Track when cumulative cash flow turns positive")
        
        if abs(overall_variance) <= 5:
            recommendations.append("✅ **Excellent Forecast:** Cash flows tracking very close to plan")
        elif abs(overall_variance) <= 15:
            recommendations.append("🟡 **Good Performance:** Minor variances from plan")
        else:
            recommendations.append("🔴 **Variance Alert:** Significant deviation from planned cash flows")
        
        if negative_quarters > positive_quarters:
            recommendations.append("💸 **Funding Focus:** More outflow quarters than inflow - ensure adequate financing")
        
        if critical_quarters > 0:
            recommendations.append(f"⚠️ **Risk Areas:** {critical_quarters} quarters with critical variances")
        
        # General recommendations
        recommendations.extend([
            "📊 **Regular Reviews:** Continue monthly cash flow monitoring",
            "🎯 **Milestone Tracking:** Ensure major inflows align with deliverables"
        ])
        
        st.markdown(f"""
        <div class="exec-summary">
            <h4>🎯 Strategic Recommendations</h4>
            <ul>
                {''.join([f'<li>{rec}</li>' for rec in recommendations])}
            </ul>
        </div>
        """, unsafe_allow_html=True)
    
    # Cash Flow Health Score
    st.markdown("### 🏥 Cash Flow Health Assessment")
    
    # Calculate health score - adjusted for projects with negative cash flows
    variance_score = max(0, 100 - abs(overall_variance))
    consistency_score = cash_flow_consistency
    
    # Trend score adjusted for negative values
    if trend > 0:
        trend_score = min(100, 50 + trend/2)
    else:
        trend_score = max(0, 50 + trend/2)
    
    overall_health = (variance_score + consistency_score + trend_score) / 3
    
    health_color = "success" if overall_health >= 80 else "warning" if overall_health >= 60 else "error"
    health_status = "Excellent" if overall_health >= 80 else "Good" if overall_health >= 60 else "Needs Attention"
    
    col1, col2, col3, col4 = st.columns(4)
    
    with col1:
        st.metric("Variance Score", f"{variance_score:.0f}/100")
    
    with col2:
        st.metric("Consistency Score", f"{consistency_score:.0f}/100")
    
    with col3:
        st.metric("Trend Score", f"{trend_score:.0f}/100")
    
    with col4:
        st.markdown(f"**Overall Health:** :{health_color}[{overall_health:.0f}/100 - {health_status}]")

def render_earned_value_analysis(project_data):
    """Render comprehensive earned value management analysis"""
    st.subheader("📊 Earned Value Management Analysis")
    
    earned_value = project_data.get('earned_value', {})
    
    if not earned_value:
        st.warning("📊 No earned value data available.")
        return
    
    # EVM Key metrics
    col1, col2, col3, col4 = st.columns(4)
    
    with col1:
        planned_value = earned_value.get('planned_value', 0)
        st.metric("Planned Value (PV)", format_currency_millions(planned_value))
    
    with col2:
        earned_value_amount = earned_value.get('earned_value', 0)
        st.metric("Earned Value (EV)", format_currency_millions(earned_value_amount))
    
    with col3:
        actual_cost = earned_value.get('actual_cost', 0)
        st.metric("Actual Cost (AC)", format_currency_millions(actual_cost))
    
    with col4:
        estimate_at_completion = earned_value.get('estimate_at_completion', 0)
        st.metric("Estimate at Completion", format_currency_millions(estimate_at_completion))
    
    # Performance indices
    col5, col6, col7, col8 = st.columns(4)
    
    with col5:
        cpi = earned_value.get('cost_performance_index', 1.0)
        cpi_status = "🟢" if cpi >= 1.0 else "🟡" if cpi >= 0.9 else "🔴"
        st.metric("Cost Performance Index", f"{cpi:.2f}", f"{cpi_status}")
    
    with col6:
        spi = earned_value.get('schedule_performance_index', 1.0)
        spi_status = "🟢" if spi >= 1.0 else "🟡" if spi >= 0.9 else "🔴"
        st.metric("Schedule Performance Index", f"{spi:.2f}", f"{spi_status}")
    
    with col7:
        cost_variance = earned_value.get('cost_variance', 0)
        cv_status = "🟢" if cost_variance >= 0 else "🟡" if cost_variance >= -planned_value*0.1 else "🔴"
        st.metric("Cost Variance", format_currency_thousands(cost_variance), f"{cv_status}")
    
    with col8:
        schedule_variance = earned_value.get('schedule_variance', 0)
        sv_status = "🟢" if schedule_variance >= 0 else "🟡" if schedule_variance >= -planned_value*0.1 else "🔴"
        st.metric("Schedule Variance", format_currency_thousands(schedule_variance), f"{sv_status}")
    
    # EVM Chart
    evm_data = {
        'Planned Value': planned_value/1000,
        'Earned Value': earned_value_amount/1000,
        'Actual Cost': actual_cost/1000
    }
    
    fig = go.Figure()
    
    fig.add_trace(go.Bar(
        name='EVM Metrics',
        x=list(evm_data.keys()),
        y=list(evm_data.values()),
        marker_color=['blue', 'green', 'red'],
        opacity=0.7
    ))
    
    fig.update_layout(
        title='Earned Value Management Overview',
        yaxis_title='Value (CHF Thousands)',
        height=400
    )
    
    st.plotly_chart(fig, use_container_width=True)


def render_project_risk_assessment(project_data):
    """Render detailed project risk assessment"""
    st.subheader("⚠️ Project Risk Assessment")
    
    # Ensure project_data is not None and has the expected structure
    if not project_data or not isinstance(project_data, dict):
        st.warning("⚠️ Unable to load risk assessment data.")
        return
    
    # Safely get risk_factors with proper error handling
    risk_factors = project_data.get('risk_factors', [])
    
    # Debug option (optional - can be removed in production)
    with st.expander("🔍 Risk Assessment Debug", expanded=False):
        st.write(f"Total risk factors in data: {len(risk_factors)}")
        if st.checkbox("Show raw risk data", value=False, key="risk_debug_checkbox"):
            st.json(risk_factors)
    
    # Ensure risk_factors is a list
    if not isinstance(risk_factors, list):
        risk_factors = []
    
    if not risk_factors or len(risk_factors) == 0:
        st.success("✅ No significant risk factors identified.")
        return
    
    # Risk summary
    risk_counts = {'Critical': 0, 'High': 0, 'Medium': 0, 'Low': 0}
    for risk in risk_factors:
        if isinstance(risk, dict) and 'severity' in risk:
            severity = risk.get('severity', 'Low')
            if severity in risk_counts:
                risk_counts[severity] += 1
    
    # Display risk count cards
    cols = st.columns(4)
    
    severity_configs = [
        ('Critical', '🔴', 'risk-critical', risk_counts['Critical']),
        ('High', '🟠', 'risk-high', risk_counts['High']),
        ('Medium', '🟡', 'risk-medium', risk_counts['Medium']),
        ('Low', '🟢', 'risk-low', risk_counts['Low'])
    ]
    
    for col, (severity, icon, css_class, count) in zip(cols, severity_configs):
        with col:
            st.markdown(f"""
            <div class="risk-card {css_class}">
                <h4>{icon} {severity}</h4>
                <h2>{count}</h2>
            </div>
            """, unsafe_allow_html=True)
    
    # Risk details table
    st.markdown("### 📋 Risk Register")
    
    risk_details = []
    for i, risk in enumerate(risk_factors):
        if isinstance(risk, dict):
            severity = risk.get('severity', 'Unknown')
            severity_icon = {"Critical": "🔴", "High": "🟠", "Medium": "🟡", "Low": "🟢"}.get(severity, "⚪")
            impact = risk.get('impact', 'Unknown')
            impact_icon = {"High": "⚡", "Medium": "⚡", "Low": "💧"}.get(impact, "💧")
            
            risk_details.append({
                'Type': risk.get('type', 'Unknown'),
                'Severity': f"{severity_icon} {severity}",
                'Impact': f"{impact_icon} {impact}",
                'Description': risk.get('description', 'No description')[:100] + '...' if len(risk.get('description', '')) > 100 else risk.get('description', 'No description'),
                'Recommendation': risk.get('recommendation', 'No recommendation')[:100] + '...' if len(risk.get('recommendation', '')) > 100 else risk.get('recommendation', 'No recommendation')
            })
    
    if risk_details:
        df_risks = pd.DataFrame(risk_details)
        st.dataframe(
            df_risks, 
            use_container_width=True,
            hide_index=True
        )
    else:
        st.info("No risk details available.")
    
def render_performance_trends_analysis(project_data):
    """Render performance trends and forecasting analysis"""
    st.subheader("📈 Performance Trends & Forecasting")
    
    # Historical performance comparison
    col1, col2 = st.columns(2)
    
    with col1:
        st.markdown("#### 📊 Period-over-Period Analysis")
        
        # Create performance comparison
        metrics = ['Contract Price', 'Revenues', 'POC%', 'Cash IN', 'Cash OUT']
        comparison_data = []
        
        for metric in metrics:
            current = safe_get_value(project_data, 'revenues', metric, 'n_ptd')
            previous = safe_get_value(project_data, 'revenues', metric, 'n1_ptd')
            
            # FIX #6: Special handling for POC velocity
            if metric == 'POC%':
                variance = calculate_poc_velocity(current, previous)
            else:
                variance = calculate_period_variance(current, previous)
            
            comparison_data.append({
                'Metric': metric,
                'Previous': previous,
                'Current': current,
                'Variance %': variance
            })
        
        df_comparison = pd.DataFrame(comparison_data)
        
        # Performance variance chart
        fig = go.Figure()
        
        fig.add_trace(go.Bar(
            name='Variance %',
            x=df_comparison['Metric'],
            y=df_comparison['Variance %'],
            marker_color=['green' if v >= 0 else 'red' for v in df_comparison['Variance %']]
        ))
        
        fig.add_hline(y=0, line_dash="dash", line_color="black")
        fig.update_layout(
            title='Performance Variance Analysis',
            yaxis_title='Variance %',
            height=400
        )
        
        st.plotly_chart(fig, use_container_width=True)
    
    with col2:
        st.markdown("#### 🎯 Forecasting & Projections")
        
        # Project completion forecast
        current_poc = safe_get_value(project_data, 'revenues', 'POC%', 'n_ptd')
        previous_poc = safe_get_value(project_data, 'revenues', 'POC%', 'n1_ptd')
        # FIX #6: Use corrected POC velocity calculation
        poc_velocity = calculate_poc_velocity(current_poc, previous_poc)
        
        if poc_velocity > 0:
            months_to_completion = (100 - current_poc) / poc_velocity if poc_velocity > 0 else 999
        else:
            months_to_completion = 999
        
        st.metric("Current POC", f"{current_poc:.1f}%")
        st.metric("POC Velocity", f"{poc_velocity:.1f}%/month")
        
        if months_to_completion < 100:
            completion_status = "🟢" if months_to_completion <= 12 else "🟡" if months_to_completion <= 24 else "🔴"
            st.metric("Est. Completion", f"{months_to_completion:.1f} months", f"{completion_status}")
        else:
            st.metric("Est. Completion", "Unable to forecast", "🔴")
        
        # Performance efficiency
        earned_value = project_data.get('earned_value', {})
        cpi = earned_value.get('cost_performance_index', 1.0)
        spi = earned_value.get('schedule_performance_index', 1.0)
        
        efficiency_score = (cpi + spi) / 2 * 100
        efficiency_status = "🟢" if efficiency_score >= 100 else "🟡" if efficiency_score >= 90 else "🔴"
        
        st.metric("Efficiency Score", f"{efficiency_score:.1f}%", f"{efficiency_status}")
    
    # Quarterly forecast table
    quarterly_data = project_data.get('quarterly', {})
    if quarterly_data:
        st.markdown("#### 📅 Quarterly Performance Forecast")
        
        quarterly_forecast = []
        for quarter in ['Q1', 'Q2', 'Q3', 'Q4']:
            q_data = quarterly_data.get(quarter, {})
            if q_data:
                actuals = q_data.get('actuals', 0)
                budget = q_data.get('budget', 0)
                delta_pct = q_data.get('delta_pct', 0)
                
                status = "🟢" if delta_pct >= 0 else "🟡" if delta_pct >= -10 else "🔴"
                
                quarterly_forecast.append({
                    'Quarter': quarter,
                    'Budget': format_currency_thousands(budget),
                    'Actuals': format_currency_thousands(actuals),
                    'Delta %': f"{delta_pct:+.1f}%",
                    'Status': status
                })
        
        if quarterly_forecast:
            df_quarterly = pd.DataFrame(quarterly_forecast)
            st.dataframe(df_quarterly, use_container_width=True)

# ================================================================================
# PAGE RENDERING FUNCTIONS
# ================================================================================

def render_data_upload_page():
    """Render the data upload page"""
    st.header("📁 Data Upload")
    
    st.markdown("""
    <div class="exec-summary">
        <h4>📋 Template Support</h4>
        <ul>
            <li><strong>Compatible with:</strong> Template v2.3 and v2.4</li>
            <li><strong>Required Sheets:</strong> Project Info, Revenues, Cost Breakdown, Cash Flow</li>
            <li><strong>Focus Areas:</strong> Quarterly cash flow performance, margin analysis, risk assessment</li>
            <li><strong>Key Features:</strong> Executive KPIs, cost performance tracking, earned value management</li>
        </ul>
    </div>
    """, unsafe_allow_html=True)
    
    uploaded_files = st.file_uploader(
        "Choose Excel files",
        type=['xlsx'],
        accept_multiple_files=True,
        help="Upload project files in Template format"
    )
    
    if uploaded_files:
        st.success(f"📁 Uploaded {len(uploaded_files)} file(s)")
        
        progress_bar = st.progress(0)
        
        for i, uploaded_file in enumerate(uploaded_files):
            st.write(f"**Processing:** {uploaded_file.name}")
            
            project_data = parse_excel_template_v24(uploaded_file)
            
            if project_data:
                project_name = project_data['project_info'].get('Project Name', uploaded_file.name)
                project_no = project_data['project_info'].get('Project No.', f'Unknown_{i+1}')
                
                st.session_state.projects_data[project_no] = {
                    'name': project_name,
                    'filename': uploaded_file.name,
                    'data': project_data,
                    'upload_time': datetime.datetime.now(),
                    'template_version': 'v2.3/v2.4'
                }
                
                st.success(f"✅ Successfully processed: {project_name} ({project_no})")
            else:
                st.error(f"❌ Failed to process: {uploaded_file.name}")
            
            progress_bar.progress((i + 1) / len(uploaded_files))

def render_portfolio_overview_page():
    """Render the comprehensive portfolio overview page - UPDATED VERSION"""
    st.header("🎯 Executive Portfolio Overview")
    
    if not st.session_state.projects_data:
        st.warning("📋 No project data uploaded. Please go to 'Data Upload' to upload Excel files.")
        return
    
    try:
        portfolio_data = {k: v for k, v in st.session_state.projects_data.items() 
                         if 'data' in v and v['data'] is not None}
        
        if not portfolio_data:
            st.error("❌ No valid project data found.")
            return
        
        portfolio_summary = create_enhanced_portfolio_summary(portfolio_data)
        
        if not portfolio_summary:
            st.error("❌ Unable to create portfolio summary.")
            return
        
        # Render comprehensive dashboard components
        render_executive_kpi_dashboard(portfolio_summary)
        
        # EXISTING MARGIN ANALYSIS
        render_enhanced_margin_analysis(portfolio_data)
        
        # NEW: ADD MARGIN VARIABILITY ANALYSIS
        st.markdown("---")
        render_margin_variability_analysis(portfolio_data)

        # ADD NEW REVENUE ANALYTICS HERE
        st.markdown("---")
        render_portfolio_revenue_analytics(portfolio_data)  # <-- NEW SECTION
        
        # CONTINUE WITH EXISTING SECTIONS
        render_work_package_analysis(portfolio_data)
        render_quarterly_cash_flow_analysis(portfolio_data)
        render_comprehensive_risk_assessment(portfolio_data)
        render_executive_project_table(portfolio_data)
        
    except Exception as e:
        st.error(f"❌ Error loading portfolio overview: {str(e)}")
        st.exception(e)

# ================================================================================
# MAIN APPLICATION
# ================================================================================

def main():
    """Main application function - COMPLETELY CLEAN AND MINIMAL"""
    
    # Initialize session state
    if 'projects_data' not in st.session_state:
        st.session_state.projects_data = {}
    
    # Header
    st.markdown("""
    <div class="main-header">
        <h1>📊 Executive Portfolio Dashboard</h1>
        <p>Enhanced Template v2.3/v2.4 Compatible | All Issues Fixed</p>
    </div>
    """, unsafe_allow_html=True)
    
    # Navigation
    st.sidebar.title("🧭 Dashboard Navigation")
    page = st.sidebar.selectbox(
        "Select View",
        ["📊 Portfolio Overview", "🔍 Project Analysis", "📁 Data Upload"],
        help="Navigate between dashboard sections"
    )
    

    st.sidebar.markdown("---")
    st.sidebar.markdown("### ⚙️ Settings")

    with st.sidebar.expander("🎯 Thresholds"):
        # CM2 Thresholds
        st.markdown("**CM2 Margin Thresholds**")
        cm2_excellent = st.number_input("CM2 Excellent (%)", value=15.0, step=1.0, key="cm2_excellent")
        cm2_good = st.number_input("CM2 Good (%)", value=10.0, step=1.0, key="cm2_good")
        cm2_warning = st.number_input("CM2 Warning (%)", value=5.0, step=1.0, key="cm2_warning")
    
        st.markdown("---")
    
    # Update all thresholds
        EXECUTIVE_THRESHOLDS['cm2_margin']['excellent'] = cm2_excellent
        EXECUTIVE_THRESHOLDS['cm2_margin']['good'] = cm2_good
        EXECUTIVE_THRESHOLDS['cm2_margin']['warning'] = cm2_warning
        EXECUTIVE_THRESHOLDS['cm2_margin']['critical'] = 0  # Always 0

# Add this right after the threshold settings
#    if st.sidebar.checkbox("Show Active Thresholds", value=False):
#        st.sidebar.markdown("**Active Threshold Values:**")
#        st.sidebar.json({
#            "CM2": EXECUTIVE_THRESHOLDS['cm2_margin'],
#        })

    # Project status
    if st.session_state.projects_data:
        st.sidebar.markdown("---")
        st.sidebar.markdown("### 📋 Projects")
        
        total_projects = len(st.session_state.projects_data)
        st.sidebar.info(f"**Total:** {total_projects}")
        
        if st.sidebar.button("🗑️ Clear Data"):
            st.session_state.projects_data = {}
            st.sidebar.success("Data cleared!")
            st.rerun()
    
    # Page routing with individual error handling
    if "Data Upload" in page:
    # Clear rendered sections when switching pages
        if 'rendered_sections' in st.session_state:
            st.session_state.rendered_sections.clear()
        try:
            render_data_upload_page()
        except Exception as e:
            st.error(f"Error in Data Upload: {str(e)}")

    elif "Portfolio Overview" in page:
        # Clear rendered sections when switching pages
        if 'rendered_sections' in st.session_state:
            st.session_state.rendered_sections.clear()
        try:
            render_portfolio_overview_page()
        except Exception as e:
            st.error(f"Error in Portfolio Overview: {str(e)}")
            st.exception(e)

    elif "Project Analysis" in page:
        # Clear rendered sections when switching pages
        if 'rendered_sections' in st.session_state:
            st.session_state.rendered_sections.clear()
        try:
            render_project_analysis_page()
        except Exception as e:
            st.error(f"Error in Project Analysis: {str(e)}")
    
    # Footer
    st.markdown("---")
    st.markdown("""
    <div style="text-align: center; color: #666; padding: 1rem;">
        <strong>Executive Portfolio Dashboard v5.0</strong><br>
        Enhanced with All Fixes Applied | Error-Free
    </div>
    """, unsafe_allow_html=True)

if __name__ == "__main__":
    main()
